"""
每周渠道报价表解析与匹配（JTT物流每周报价单）

解析对象：`/Users/admin/报价工具/JTT物流每周报价单（26年x月第x周）...xlsx`，
格式基本固定：
  - sheet「汇总价格表」：按国家分区（美国渠道/加拿大渠道/欧线渠道/英国渠道），
    每区数据行为 渠道(系统下单渠道名称, 向下填充) × 仓点(仓库/邮编) × 重量段单价。
    重量段列名形如「价格(12KG+)」「价格（50KG+)」「价格(334KG+)」，各子表段位不同，
    段位定义在区头行或子表头行（C列=「系统下单渠道名称」）读取。
  - sheet「美国专线普船直送渠道报价表」：渠道 × 覆盖仓点 × 100KG+ 单段价。

匹配能力：
  - 渠道：归一化（去注释/空白、含税↔包税等价）+ 别名表 + 子串 + 关键词打分
  - 仓点：代码列表 / 范围(YYZ1-9) / 代码+邮编(DTM2-44145) / 前缀(97.98.99开头) /
          国家(德国/波兰、法国) / FBA仓兜底
  - 重量段：计费重取最大阈值 ≤ 计费重的段价
"""
import openpyxl
import re

SECTION_NAMES = {'美国渠道', '加拿大渠道', '欧线渠道', '英国渠道'}
SUB_HEADER_MARK = '系统下单渠道名称'

# 渠道别名（原始渠道名 -> 每周表渠道名），用于个别名称顺序/叫法不一致的情况
CHANNEL_ALIASES = {
    '欧洲含税-铁路快铁卡派': '欧洲铁路包税-快铁卡派',
    '欧洲含税-铁路普铁卡派': '欧洲铁路包税-普铁卡派',
}

# 渠道关键词打分用
CHANNEL_KEYWORDS = [
    '欧洲', '美国', '英国', '加拿大', '美转加',
    '空派', '海运', '铁路', '普铁', '快铁', '卡车', '直送',
    '卡派', '快递派', '超快线', '快线', 'OA',
    '带电', '普货', '自税', '包税', '递延', '整柜', '专线',
]

# 仓点代码前缀 -> 国家（用于「德国/波兰、法国」等按国家匹配）
COUNTRY_PREFIX = {
    'WRO': 'PL', 'POZ': 'PL', 'GDN': 'PL', 'KRK': 'PL',
    'DTM': 'DE', 'HAJ': 'DE', 'BER': 'DE', 'KSF': 'DE',
    'FRA': 'DE', 'LEJ': 'DE', 'HAM': 'DE', 'CGN': 'DE',
    'BHX': 'UK', 'LBA': 'UK', 'MAN': 'UK', 'LHR': 'UK', 'EMA': 'UK',
    'YVR': 'CA', 'YYC': 'CA', 'YYZ': 'CA', 'YOW': 'CA', 'YEG': 'CA',
    'YHM': 'CA', 'YXU': 'CA', 'YXX': 'CA', 'YGK': 'CA', 'YUL': 'CA',
    'YOO': 'CA', 'XYY': 'CA',
}
COUNTRY_NAMES = {
    '德国': 'DE', '德': 'DE',
    '波兰': 'PL', '法国': 'FR', '波兰、法国': 'PL,FR', '波兰、法国、': 'PL,FR',
}


# ── 解析 ──

def _clean_channel(v):
    """渠道名清洗：去括号注释、空白/换行"""
    s = str(v or '').strip()
    s = re.sub(r'[（(][^（）()]*[）)]', '', s)  # 去（…）与 (…)
    s = re.sub(r'\s+', '', s)
    return s


def _norm_channel(v):
    """渠道名归一化（用于匹配）：清洗 + 含税↔包税等价"""
    s = _clean_channel(v)
    s = s.replace('含税', '包税')
    return s


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s == '/':
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_tier_cols(ws, row, start_col=5, max_col=17):
    """从某行的表头单元格提取 重量段阈值 -> 列号 映射，如 {12:5, 50:6, 100:7}"""
    tiers = {}
    for c in range(start_col, max_col + 1):
        h = str(ws.cell(row=row, column=c).value or '').strip()
        m = re.search(r'(\d+)\s*KG', h)
        if m:
            tiers[int(m.group(1))] = c
    return tiers


def parse_weekly_quotation(filepath):
    """解析每周报价表，返回 entries 列表：

    [{'section', 'channel', 'channel_raw', 'wh_pattern', 'tiers': {threshold: price}}]
    """
    if not filepath or not __import__('os').path.exists(filepath):
        return []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    entries = []

    try:
        # ── sheet 1: 汇总价格表 ──
        sheet = None
        if '汇总价格表' in wb.sheetnames:
            sheet = wb['汇总价格表']
        else:
            for sn in wb.sheetnames:
                if '价格表' in sn or '汇总' in sn:
                    sheet = wb[sn]
                    break
        if sheet is None and wb.sheetnames:
            sheet = wb[wb.sheetnames[0]]

        if sheet is not None:
            ws = sheet
            # 找各区起点
            sec_starts = []
            for r in range(1, ws.max_row + 1):
                a = str(ws.cell(row=r, column=1).value or '').strip()
                if a in SECTION_NAMES:
                    sec_starts.append(r)

            for idx, start in enumerate(sec_starts):
                end = (sec_starts[idx + 1] if idx + 1 < len(sec_starts) else ws.max_row + 1) - 1
                section = str(ws.cell(row=start, column=1).value or '').strip()
                hdr_row = start + 1
                tiers = _parse_tier_cols(ws, hdr_row)
                cur_channel_raw = ''
                cur_channel = ''
                for r in range(start + 2, end + 1):
                    c_val = str(ws.cell(row=r, column=3).value or '').strip()
                    d_val = str(ws.cell(row=r, column=4).value or '').strip()
                    # 子表头：重定义重量段
                    if SUB_HEADER_MARK in c_val or d_val == '仓库/邮编':
                        new_tiers = _parse_tier_cols(ws, r)
                        if new_tiers:
                            tiers = new_tiers
                        continue
                    if c_val:
                        cur_channel_raw = _clean_channel(c_val)
                        cur_channel = _norm_channel(c_val)
                    if not d_val or not cur_channel:
                        continue
                    prices = {}
                    for thr, col in tiers.items():
                        prices[thr] = _to_float(ws.cell(row=r, column=col).value)
                    entries.append({
                        'section': section,
                        'channel': cur_channel,
                        'channel_raw': cur_channel_raw,
                        'wh_pattern': d_val,
                        'tiers': prices,
                    })

        # ── sheet 2: 美国专线普船直送渠道报价表 ──
        for sn in wb.sheetnames:
            if '普船直送' in sn:
                ws2 = wb[sn]
                hdr = 2
                cur_ch_raw = ''
                cur_ch = ''
                for r in range(hdr + 1, ws2.max_row + 1):
                    b = str(ws2.cell(row=r, column=2).value or '').strip()
                    d = str(ws2.cell(row=r, column=4).value or '').strip()
                    e = ws2.cell(row=r, column=5).value
                    if b:
                        cur_ch_raw = _clean_channel(b.split('\n')[0] if '\n' in b else b)
                        cur_ch = _norm_channel(cur_ch_raw)
                    if not d or not cur_ch:
                        continue
                    price = _to_float(e)
                    for wh_code in re.split(r'[、，,/\s]+', d):
                        wh_code = wh_code.strip()
                        if wh_code:
                            entries.append({
                                'section': '美国渠道',
                                'channel': cur_ch,
                                'channel_raw': cur_ch_raw,
                                'wh_pattern': wh_code,
                                'tiers': {100: price},
                            })
    finally:
        wb.close()

    return entries


# ── 仓点匹配 ──

def _extract_prefixes(pattern):
    """提取邮编/开头前缀，如 "97.98.99开头" -> ['97','98','99']；"FBA（8/9邮编）" -> ['8','9']"""
    return re.findall(r'\d{1,3}', pattern)


def _is_country_pattern(pattern):
    return any(name in pattern for name in COUNTRY_NAMES)


def _country_codes(pattern):
    codes = set()
    for name, cc in COUNTRY_NAMES.items():
        if name in pattern:
            for c in cc.split(','):
                codes.add(c.strip())
    return codes


def _warehouse_country(w):
    for prefix, cc in COUNTRY_PREFIX.items():
        if w.startswith(prefix):
            return cc
    return ''


def _expand_token(tok):
    """展开单个仓点 token，返回匹配集合。

    支持：代码+邮编 "DTM2-44145" -> {DTM2, 44145}；范围 "YYZ1-9" -> YYZ1..YYZ9；
    纯代码 "FTW1" -> {FTW1}；纯数字 "8" -> {'8'}
    """
    tok = tok.strip()
    if not tok:
        return set()
    t = tok.upper()
    # 代码+邮编：字母+数字-5位邮编
    m = re.fullmatch(r'([A-Z]+\d+)-(\d{5})', t)
    if m:
        return {m.group(1), m.group(2)}
    # 范围：可选字母前缀 + 数字-数字
    m = re.fullmatch(r'([A-Z]*)(\d+)-(\d+)', t)
    if m:
        letters, s, e = m.group(1), int(m.group(2)), int(m.group(3))
        if s <= e <= s + 20:  # 防超长区间
            return {letters + str(i) for i in range(s, e + 1)}
        return {t}
    return {t}


def _split_tokens(pattern):
    """按常见分隔符拆仓点 token（保留 token 内连字符以便范围/邮编解析）"""
    toks = re.split(r'[,，、/\\;·\s]+', pattern)
    return [t for t in toks if t.strip()]


def match_warehouse(pattern, warehouse):
    """判断每周报价表仓点模式是否匹配给定仓库代码"""
    w = str(warehouse or '').strip().upper()
    p = str(pattern or '').strip()
    if not w or not p:
        return False
    # 1. 邮编/开头 前缀模式
    if '邮编' in p or '开头' in p:
        prefixes = _extract_prefixes(p)
        if prefixes and any(w.startswith(pr) for pr in prefixes):
            return True
    # 2. 国家模式（德国/波兰、法国…）
    if _is_country_pattern(p):
        return _warehouse_country(w) in _country_codes(p)
    # 3. 代码/范围/列表精确匹配
    expanded = set()
    for tok in _split_tokens(p):
        expanded |= _expand_token(tok)
    if w in expanded:
        return True
    # 4. FBA 兜底
    if 'FBA' in p.upper():
        return True
    return False


# ── 渠道匹配 ──

def _channel_token_score(norm_pick, norm_weekly):
    return sum(1 for kw in CHANNEL_KEYWORDS if kw in norm_pick and kw in norm_weekly)


def match_channel(channel, entries):
    """按渠道匹配，返回候选 entries（最优在前）。

    顺序：别名/精确 → 子串（取最长） → 关键词打分（降序）
    """
    if not channel or not entries:
        return []
    raw = str(channel).strip()
    norm_pick = _norm_channel(raw)

    # 1. 别名
    alias_norm = None
    if raw in CHANNEL_ALIASES:
        alias_norm = _norm_channel(CHANNEL_ALIASES[raw])
    elif norm_pick in CHANNEL_ALIASES.values():
        alias_norm = norm_pick

    if alias_norm:
        exact = [e for e in entries if e['channel'] == alias_norm]
        if exact:
            return exact

    # 2. 精确
    exact = [e for e in entries if e['channel'] == norm_pick]
    if exact:
        return exact

    # 3. 子串：pick ⊆ weekly 或 weekly ⊆ pick
    sub = []
    for e in entries:
        ec = e['channel']
        if norm_pick and (norm_pick in ec or ec in norm_pick):
            sub.append(e)
    if sub:
        sub.sort(key=lambda e: -len(e['channel']))  # 最长/最具体优先
        return sub

    # 4. 关键词打分
    scored = []
    for e in entries:
        s = _channel_token_score(norm_pick, e['channel'])
        if s > 0:
            scored.append((s, e))
    scored.sort(key=lambda x: (-x[0], -len(x[1]['channel'])))
    return [e for _, e in scored]


# ── 重量段 ──

def pick_tier_price(tiers, weight):
    """取最大阈值 ≤ 计费重的段价；低于最小段取最小段；weight 未知取最小段价"""
    prices = {t: p for t, p in (tiers or {}).items() if p is not None}
    if not prices:
        return None
    if weight is None:
        return prices[min(prices.keys())]
    chosen = None
    for t in sorted(prices.keys()):
        if t <= weight:
            chosen = t
    if chosen is None:
        chosen = min(prices.keys())
    return prices[chosen]


# ── 计费重复刻（与模板 AD 列公式一致） ──

def compute_chargeable_weight(row, channel):
    """复刻参考值模版 AD 列计费重：ROUND(MAX(总实重, 总材积重, [快递派:M*12]), 0)

    优先用参考尺寸 V/W/X/Y；无参考用实际尺寸 N/O/P/Q；都没有 → 返回 None。
    """
    def num(v):
        try:
            return float(v) if v not in (None, '') else 0.0
        except (ValueError, TypeError):
            return 0.0

    M = num(row.get('box_count'))
    if M <= 0:
        return None
    is_exp = '快递派' in (channel or '')

    ref_actual = num(row.get('ref_w'))            # V 参考实重
    ref_l = num(row.get('ref_l'))                  # W 参考长
    ref_wid = num(row.get('ref_wid'))              # X 参考宽
    ref_h = num(row.get('ref_h'))                  # Y 参考高
    if ref_actual or (ref_l and ref_wid and ref_h):
        total_actual = ref_actual * M
        total_vol = (ref_l * ref_wid * ref_h / 6000.0) * M if (ref_l and ref_wid and ref_h) else 0.0
        if is_exp:
            return round(max(total_vol, M * 12))
        return round(max(total_actual, total_vol))

    actual = num(row.get('weight'))
    l = num(row.get('length'))
    w = num(row.get('width'))
    h = num(row.get('height'))
    if not (actual or (l and w and h)):
        return None
    total_actual = actual * M
    total_vol = (l * w * h / 6000.0) * M if (l and w and h) else 0.0
    if is_exp:
        return round(max(total_vol, M * 12))
    return round(max(total_actual, total_vol))


# ── 组合查找 ──

def find_weekly_price(entries, channel, warehouse, weight):
    """按 渠道+仓点+计费重 匹配每周报价，返回 (price, matched_channel_raw, matched_pattern)。

    匹配不到 → (None, '', '')。
    """
    for e in match_channel(channel, entries):
        if not match_warehouse(e['wh_pattern'], warehouse):
            continue
        p = pick_tier_price(e['tiers'], weight)
        if p is not None:
            return p, e['channel_raw'], e['wh_pattern']
    return None, '', ''
