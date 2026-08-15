"""
拣货数据导出工具 v2（2026-06-10 新规则）

输入：
  1. 发票 (.xlsx)  — 头部含服务/仓库代码/箱数，数据区含货箱编号+重量+尺寸+品名
  2. 系统导出拣货数据 (.xlsx) — 运单号+扩展箱号对照
  3. 箱规历史数据库 (.xlsx) — 品名+客户箱规 → 标准箱规
  4. 内部拣货数据参考值模版 (.xlsx) — 输出格式+公式

输出：
  内部拣货数据参考值 (.xlsx) — 按 SO 号归组，匹配历史箱规

匹配逻辑：
  - 发票每行货箱编号 → 前12位 = FBA ID → 匹配系统导出的扩展箱号 → 取运单号(SO号)
  - 箱规历史：品名+重量/尺寸匹配 → 取标准箱规(V/W/X/Y)
  - 无历史匹配 → V/W/X/Y 留空，标红

用法：
  python3 export_picking_data.py
  python3 export_picking_data.py <发票文件> <系统导出文件> [输出文件]
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
import os
import re
import sys

# ── 配置 ──
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_FILE = os.path.join(DATA_DIR, "箱规历史数据库.xlsx")
TEMPLATE_FILE = os.path.join(DATA_DIR, "内部拣货数据参考值模版.xlsx")

QUOTATION_FILE = os.path.join(DATA_DIR, "报价表.xlsx")
DEFAULT_COUNTRY_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# ── 辅助函数 ──

def extract_fba_id(box_no):
    """从货箱编号提取前12位 FBA ID"""
    box_no = str(box_no).strip()
    return box_no[:12] if len(box_no) >= 12 else box_no


def calc_box_count(box_no):
    """计算箱数：从货箱编号 U{start}-{end} 或 U{num} 格式"""
    box_no = str(box_no).strip()
    # 匹配 U 后面跟数字，可选 -{数字}
    m = re.search(r'U(\d+)(?:-(\d+))?$', box_no)
    if m:
        start = int(m.group(1))
        end = int(m.group(2)) if m.group(2) else start
        return end - start + 1
    return 1


def country_from_warehouse(warehouse):
    """根据仓库代码推断国家（仅供参考，最终手动确认）"""
    w = str(warehouse or '').upper().strip()
    us_warehouses = {'RFD2', 'IAH3', 'LGB8', 'LAX9', 'SMF3', 'ONT8', 'FTW1', 'MEM6', 'MDW2'}
    if w in us_warehouses or w.startswith(('RFD', 'IAH', 'LGB', 'LAX', 'SMF', 'ONT', 'FTW', 'MEM', 'MDW')):
        return 'US'
    # 欧洲常见仓库前缀
    eu_prefixes = ('DTM', 'WRO', 'HAJ', 'FBA', 'POZ', 'AVP', 'YVR')
    if any(w.startswith(p) for p in eu_prefixes):
        return 'DE'  # 默认欧洲-德国
    return ''


# ── 解析函数 ──

def parse_invoice(filepath):
    """解析发票文件，返回 (data_rows, service, warehouse)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sn = wb.sheetnames
    # 查找主 Sheet
    main_sheet = None
    for name in sn:
        if '发票' in name or 'page' in name.lower():
            main_sheet = name
            break
    if not main_sheet:
        main_sheet = sn[0]

    ws = wb[main_sheet]

    # ── 提取头部信息 ──
    service = ''
    warehouse = ''
    for r in range(1, 25):
        a = str(ws.cell(row=r, column=1).value or '').strip()
        b = ws.cell(row=r, column=2).value
        if a == '服务':
            service = str(b or '').strip()
        elif a == '收件人姓名':
            warehouse = str(b or '').strip()

    # ── 查找数据表头行 ──
    data_start = None
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(row=r, column=1).value or '').strip()
        if '货箱编号' in a:
            data_start = r
            break

    if data_start is None:
        wb.close()
        return [], service, warehouse

    # ── 读取数据行 ──
    data_rows = []
    for r in range(data_start + 1, ws.max_row + 1):
        box_no = str(ws.cell(row=r, column=1).value or '').strip()
        if not box_no:
            break
        row = {
            'box_no': box_no,
            'weight': ws.cell(row=r, column=2).value,    # B
            'length': ws.cell(row=r, column=3).value,     # C
            'width': ws.cell(row=r, column=4).value,       # D
            'height': ws.cell(row=r, column=5).value,      # E
            'en_name': ws.cell(row=r, column=6).value,     # F
            'cn_name': ws.cell(row=r, column=7).value,     # G
            'warehouse': warehouse,  # 每行携带自己的仓库代码
            'service': service,      # 每行携带自己的物流渠道
        }
        data_rows.append(row)

    wb.close()
    return data_rows, service, warehouse


def parse_system_export(filepath):
    """解析系统导出拣货数据，返回 {FBA前缀 → SO号} 映射"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    prefix_to_so = {}
    for r in range(2, ws.max_row + 1):
        ext_box = str(ws.cell(row=r, column=3).value or '').strip()
        so_no = str(ws.cell(row=r, column=2).value or '').strip()
        if not ext_box or not so_no:
            continue
        prefix = ext_box[:12]
        if prefix not in prefix_to_so:
            prefix_to_so[prefix] = so_no
    wb.close()
    return prefix_to_so


def parse_history(filepath):
    """解析箱规历史数据库，返回记录列表 [{name, w, l, wid, h, rw, rl, rwid, rh}]"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    records = []
    for r in range(3, ws.max_row + 1):
        name = str(ws.cell(row=r, column=1).value or '').strip()
        if not name:
            continue
        try:
            rec = {
                'name': name,
                'weight': _to_float(ws.cell(row=r, column=2).value),
                'length': _to_float(ws.cell(row=r, column=3).value),
                'width': _to_float(ws.cell(row=r, column=4).value),
                'height': _to_float(ws.cell(row=r, column=5).value),
                'rw': _to_float(ws.cell(row=r, column=6).value),
                'rl': _to_float(ws.cell(row=r, column=7).value),
                'rwid': _to_float(ws.cell(row=r, column=8).value),
                'rh': _to_float(ws.cell(row=r, column=9).value),
            }
        except (ValueError, TypeError):
            continue
        if all(v is not None for v in [rec['rw'], rec['rl'], rec['rwid'], rec['rh']]):
            records.append(rec)
    wb.close()
    return records


def _to_float(v):
    """安全转 float"""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def merge_duplicate_picking_rows(output_rows):
    """合并完全重复的数据行（海外仓快递派等"单行单箱"发票）。

    同一 运单号+渠道+仓库 下，FBA ID 完全一致时，再按 中/英文品名、实重、
    长宽高区分唯一性；完全一致的行合并为一行，总箱数(CTN)累加。
    不同 SO/渠道/仓库 或 FBA ID 不同、品名/尺寸不同 → 不合并。
    """
    merged = {}
    for row in output_rows:
        key = (
            row.get('so_no', ''), row.get('service', ''), row.get('warehouse', ''),
            row.get('fba_id', ''), row.get('cn_name', ''), row.get('en_name', ''),
            row.get('weight'), row.get('length'), row.get('width'), row.get('height'),
        )
        if key in merged:
            merged[key]['box_count'] = (merged[key].get('box_count') or 0) + (row.get('box_count') or 0)
        else:
            merged[key] = dict(row)
    return list(merged.values())


def parse_quotation(filepath):
    """解析报价表，返回 {warehouse: {e_price, f_price, supplier_ch}} 映射

    报价表格式：
        A列=物流渠道, B列=后台仓库, C列=应收单价, G列=供应商渠道, H列=应付单价

    匹配逻辑：以 后台仓库(B列) 为键，同一仓库有多条记录时取最后一条。
    """
    if not filepath or not os.path.exists(filepath):
        return {}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # 查找第一个有数据的 sheet（排除空表）
    ws = None
    for sn in wb.sheetnames:
        s = wb[sn]
        if s.max_row >= 3 and s.max_column >= 8:
            ws = s
            break
    if ws is None:
        wb.close()
        return {}
    quotation = {}
    for r in range(2, ws.max_row + 1):
        warehouse = str(ws.cell(row=r, column=2).value or '').strip()
        if not warehouse:
            continue
        e_price = _to_float(ws.cell(row=r, column=3).value)
        f_price = _to_float(ws.cell(row=r, column=8).value)   # H列=应付单价
        supplier_ch = str(ws.cell(row=r, column=7).value or '').strip()  # G列=供应商渠道
        quotation[warehouse] = {
            'e_price': e_price if e_price is not None else '',
            'f_price': f_price if f_price is not None else '',
            'supplier_ch': supplier_ch,
        }
    wb.close()
    return quotation


def find_history_match(records, name, weight, length, width, height):
    """在箱规历史数据库中查找匹配记录，返回 {rw, rl, rwid, rh} 或 None

    支持长宽高 6 种排列组合（实际数据中尺寸顺序常不一致）。
    """
    if not name:
        return None
    name = name.strip()

    # 生成所有 (l, w, h) 排列
    dims_permutations = [
        (length, width, height),
        (length, height, width),
        (width, length, height),
        (width, height, length),
        (height, length, width),
        (height, width, length),
    ]

    candidates = []
    for rec in records:
        if rec['name'] != name:
            continue
        if not _dims_match(rec['weight'], weight, 0.5):
            continue
        # 尝试各种排列
        best_score = -1
        for pl, pw, ph in dims_permutations:
            l_ok = _dims_match(rec['length'], pl, 1.0)
            w_ok = _dims_match(rec['width'], pw, 1.0)
            h_ok = _dims_match(rec['height'], ph, 1.0)
            if l_ok and w_ok and h_ok:
                score = 0
                if _dims_match(rec['length'], pl, 0.5): score += 1
                if _dims_match(rec['width'], pw, 0.5): score += 1
                if _dims_match(rec['height'], ph, 0.5): score += 1
                if score > best_score:
                    best_score = score
        if best_score >= 0:
            # weight 精确度加分
            if _dims_match(rec['weight'], weight, 0.1):
                best_score += 1
            candidates.append((best_score, rec))

    if not candidates:
        return None
    candidates.sort(key=lambda x: -x[0])
    best = candidates[0][1]
    return {'rw': best['rw'], 'rl': best['rl'], 'rwid': best['rwid'], 'rh': best['rh']}


def _dims_match(a, b, tolerance):
    """检查两个数值是否在允许误差内（任一为 None 则跳过该维度）"""
    if a is None or b is None:
        return True  # 无法比较时视为匹配
    return abs(a - b) < tolerance


def parse_invoice_merge(invoice_files):
    """解析多份发票，合并数据行

    参数:
        invoice_files: 发票文件路径列表
    返回:
        (merged_data_rows, service, warehouse)
    """
    all_rows = []
    service = ''
    warehouse = ''
    for f in invoice_files:
        rows, svc, wh = parse_invoice(f)
        all_rows.extend(rows)
        if not service and svc:
            service = svc
        if not warehouse and wh:
            warehouse = wh
    return all_rows, service, warehouse


# ── 输出生成 ──

def generate_picking_output(invoice_file, system_file, output_path,
                              history_file=None, template_file=None,
                              quotation_file=None):
    """核心入口：生成内部拣货数据参考值"""
    if history_file is None:
        history_file = HISTORY_FILE
    if template_file is None:
        template_file = TEMPLATE_FILE
    if quotation_file is None:
        quotation_file = QUOTATION_FILE

    # ── 解析输入 ──
    data_rows, service, warehouse = parse_invoice(invoice_file)
    prefix_to_so = parse_system_export(system_file)
    history_records = parse_history(history_file)
    quotation_data = parse_quotation(quotation_file)

    if not data_rows:
        raise ValueError("发票中未找到有效数据行")

    print(f"  📄 发票数据: {len(data_rows)} 行")
    print(f"  🔗 系统SO映射: {len(prefix_to_so)} 个FBA前缀")
    print(f"  📚 箱规历史: {len(history_records)} 条记录")
    print(f"  💰 报价单: {len(quotation_data)} 条")

    # ── 构建输出行数据 ──
    output_rows = []
    missing_history = []  # 记录无历史匹配的行
    for row in data_rows:
        box_no = row['box_no']
        fba_id = extract_fba_id(box_no)
        box_count = calc_box_count(box_no)
        so_no = prefix_to_so.get(fba_id, '')

        cn_name = str(row['cn_name'] or '').strip()
        en_name = str(row.get('en_name') or '').strip()
        weight = _to_float(row['weight'])
        length = _to_float(row['length'])
        width = _to_float(row['width'])
        height = _to_float(row['height'])

        hm = find_history_match(history_records, cn_name, weight, length, width, height)

        # 使用每行自己的仓库代码匹配报价单（多发票时各行仓库可能不同）
        row_wh = row.get('warehouse', warehouse)
        row_svc = row.get('service', service)
        q_info = quotation_data.get(row_wh, {})

        out = {
            'so_no': so_no,
            'service': row_svc,
            'country': country_from_warehouse(row_wh),
            'warehouse': row_wh,
            'e_price': q_info.get('e_price', ''),
            'f_price': q_info.get('f_price', ''),
            'supplier_ch': q_info.get('supplier_ch', ''),
            'ship_name': '',  # 手动输入
            'customs': '',    # 手动输入
            'fba_id': fba_id,
            'cn_name': cn_name,
            'en_name': en_name,
            'pickup_fee': '', # 手动输入
            'box_count': box_count,
            'weight': weight,
            'length': length,
            'width': width,
            'height': height,
            'history_note': '', # 手动输入
            'ref_w': hm['rw'] if hm else None,
            'ref_l': hm['rl'] if hm else None,
            'ref_wid': hm['rwid'] if hm else None,
            'ref_h': hm['rh'] if hm else None,
        }
        output_rows.append(out)
        if hm is None:
            missing_history.append(out)

    total_boxes = sum(r['box_count'] for r in output_rows)
    print(f"  📊 输出行: {len(output_rows)} 行, 总箱数: {total_boxes}")
    print(f"  ⚠️  无历史匹配: {len(missing_history)} 行")

    # ── 合并重复行（海外仓快递派"单行单箱"：FBA一致时按品名/实重/长宽高区分，重复合并）──
    output_rows = merge_duplicate_picking_rows(output_rows)

    # ── 写入模板 ──
    _write_output_to_template(output_rows, template_file, output_path)
    return output_path, total_boxes


# ── 样式常量（参考标准版格式） ──
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
DATA_FONT = Font(name='微软雅黑', size=10, color=Color(indexed=8))
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_WRAP = Alignment(vertical='center', wrap_text=True)

# 需要居中对齐的数值/公式列
CENTER_COLS = {3, 4, 5, 6, 7, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30}
# 需要左对齐的文本列
LEFT_COLS = {1, 2, 8, 9, 10, 11}

# 列宽（匹配标准版）
COL_WIDTHS = {
    'A': 18.54, 'B': 21.46, 'C': 10.56, 'D': 9.54,
    'E': 8.43, 'F': 8.43, 'G': 8.16, 'H': 27.44,
    'I': 21.78, 'J': 9.54, 'K': 17.61, 'L': 19.0,
    'M': 12.16, 'N': 7.07, 'O': 8.39, 'P': 8.39,
    'Q': 8.39, 'R': 7.61, 'S': 9.0, 'T': 6.33,
    'U': 6.11, 'V': 10.89, 'W': 12.46, 'X': 8.43,
    'Y': 7.61, 'Z': 5.93, 'AA': 8.07, 'AB': 9.54,
    'AC': 8.07, 'AD': 9.0, 'AE': 9.54,
}


def _style_data_cell(cell):
    """对数据单元格应用标准样式"""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    col = cell.column
    if col in CENTER_COLS:
        cell.alignment = ALIGN_CENTER
    elif col in LEFT_COLS:
        if col == 8:  # 供应商渠道 → 允许换行
            cell.alignment = ALIGN_WRAP
        else:
            cell.alignment = ALIGN_LEFT
def _write_output_to_template(output_rows, template_file, output_path):
    """将输出行数据写入模板并保存"""
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    merges_to_remove = [str(m) for m in list(ws.merged_cells.ranges) if m.min_row >= 2]
    for m_str in merges_to_remove:
        ws.unmerge_cells(m_str)
    max_existing = ws.max_row
    if max_existing >= 2:
        ws.delete_rows(2, max_existing - 1)

    groups = []
    cur = []
    for row in output_rows:
        if not cur:
            cur = [row]
        elif (cur[0]['so_no'] == row['so_no']
              and cur[0]['service'] == row['service']
              and cur[0]['warehouse'] == row['warehouse']):
            cur.append(row)
        else:
            groups.append(cur)
            cur = [row]
    if cur:
        groups.append(cur)

    current_row = 2
    for group in groups:
        start_row = current_row
        c1 = ws.cell(row=start_row, column=1)
        c1.value = group[0]['so_no']
        _style_data_cell(c1)
        c2 = ws.cell(row=start_row, column=2)
        c2.value = group[0]['service']
        _style_data_cell(c2)
        c3 = ws.cell(row=start_row, column=3)
        c3.value = group[0]['country']
        _style_data_cell(c3)
        c4 = ws.cell(row=start_row, column=4)
        c4.value = group[0]['warehouse']
        _style_data_cell(c4)
        for row_data in group:
            r = current_row
            # E列=应收单价, F列=应付单价, G列=供应商渠道（从报价单读取）
            for col, key in [(5, 'e_price'), (6, 'f_price'), (7, 'supplier_ch')]:
                cell = ws.cell(row=r, column=col)
                cell.value = row_data.get(key, '')
                _style_data_cell(cell)
            for col, key in [(10, 'fba_id'), (11, 'cn_name')]:
                cell = ws.cell(row=r, column=col)
                cell.value = row_data[key]
                _style_data_cell(cell)
            for col, key in [(13, 'box_count'), (14, 'weight'), (15, 'length'),
                              (16, 'width'), (17, 'height')]:
                cell = ws.cell(row=r, column=col)
                cell.value = row_data[key]
                _style_data_cell(cell)
            # R列=材积重（公式）
            cell_r = ws.cell(row=r, column=18)
            cell_r.value = f'=O{r}*P{r}*Q{r}/6000'
            cell_r.number_format = '#,##0.00'
            _style_data_cell(cell_r)
            # S列=单箱材积重差异
            cell_s = ws.cell(row=r, column=19)
            cell_s.value = f'=R{r}-Z{r}'
            _style_data_cell(cell_s)
            # T列=周长差异
            cell_t = ws.cell(row=r, column=20)
            cell_t.value = f'=O{r}+P{r}+Q{r}-Y{r}-X{r}-W{r}'
            _style_data_cell(cell_t)
            # V/W/X/Y = 参考长/宽/高/实重
            v_val, w_val, x_val, y_val = row_data['ref_w'], row_data['ref_l'], row_data['ref_wid'], row_data['ref_h']
            for col, val in [(22, v_val), (23, w_val), (24, x_val), (25, y_val)]:
                cell = ws.cell(row=r, column=col)
                cell.value = val
                _style_data_cell(cell)
            # 无历史匹配 → 标红
            if all(v is None for v in [v_val, w_val, x_val, y_val]):
                for col in [22, 23, 24, 25]:
                    ws.cell(row=r, column=col).fill = red_fill

            elif v_val is None: ws.cell(row=r, column=22).fill = red_fill
            elif w_val is None: ws.cell(row=r, column=23).fill = red_fill
            elif x_val is None: ws.cell(row=r, column=24).fill = red_fill
            elif y_val is None: ws.cell(row=r, column=25).fill = red_fill
            # Y列=材积重（公式）
            cell_y = ws.cell(row=r, column=26)
            cell_y.value = f'=W{r}*X{r}*Y{r}/6000'
            cell_y.number_format = '#,##0.00'
            _style_data_cell(cell_y)
            # Z列=体积（公式）
            cell_z = ws.cell(row=r, column=27)
            cell_z.value = f'=W{r}*X{r}*Y{r}*M{r}/1000000'
            _style_data_cell(cell_z)
            # AA列=总实重（公式）
            cell_aa = ws.cell(row=r, column=28)
            cell_aa.value = f'=V{r}*M{r}'
            _style_data_cell(cell_aa)
            # AB列=总材积重（公式）
            cell_ab = ws.cell(row=r, column=29)
            cell_ab.value = f'=Z{r}*M{r}'
            cell_ab.number_format = '#,##0.00'
            _style_data_cell(cell_ab)
            # AC列=计费重（公式）
            cell_ac = ws.cell(row=r, column=30)
            cell_ac.value = f'=ROUND(MAX(AB{r}:AC{r}),0)'
            _style_data_cell(cell_ac)
            # 行高
            ws.row_dimensions[r].height = 30

            current_row += 1
        if start_row < current_row - 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
    # ── 设置列宽 ──
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    # ── 设置表头行高（如果模板未设置） ──
    if ws.row_dimensions[1].height is None or ws.row_dimensions[1].height < 28:
        ws.row_dimensions[1].height = 28

    wb.save(output_path)
    print(f"  ✅ 已保存: {output_path}")


def generate_picking_output_multi(invoice_files, system_file, output_path,
                                   history_file=None, template_file=None,
                                   quotation_file=None):
    """支持多份发票合并输出一份拣货数据

    参数:
        invoice_files: 发票文件路径列表
        system_file: 系统导出拣货数据文件路径
        output_path: 输出文件路径
        history_file: 箱规历史数据库路径（可选）
        template_file: 模板文件路径（可选）
        quotation_file: 报价表文件路径（可选）
    返回:
        (output_path, total_boxes)
    """
    if history_file is None:
        history_file = HISTORY_FILE
    if template_file is None:
        template_file = TEMPLATE_FILE
    if quotation_file is None:
        quotation_file = QUOTATION_FILE

    # ── 解析多份发票 ──
    data_rows, service, warehouse = parse_invoice_merge(invoice_files)
    prefix_to_so = parse_system_export(system_file)
    history_records = parse_history(history_file)
    quotation_data = parse_quotation(quotation_file)

    if not data_rows:
        raise ValueError("发票中未找到有效数据行")

    print(f"  📄 发票数据（合并）: {len(invoice_files)} 个文件, {len(data_rows)} 行")
    print(f"  🔗 系统SO映射: {len(prefix_to_so)} 个FBA前缀")
    print(f"  📚 箱规历史: {len(history_records)} 条记录")
    print(f"  💰 报价单: {len(quotation_data)} 条")

    # ── 构建输出行数据 ──
    output_rows = []
    missing_history = []
    for row in data_rows:
        box_no = row['box_no']
        fba_id = extract_fba_id(box_no)
        box_count = calc_box_count(box_no)
        so_no = prefix_to_so.get(fba_id, '')

        cn_name = str(row['cn_name'] or '').strip()
        en_name = str(row.get('en_name') or '').strip()
        weight = _to_float(row['weight'])
        length = _to_float(row['length'])
        width = _to_float(row['width'])
        height = _to_float(row['height'])

        hm = find_history_match(history_records, cn_name, weight, length, width, height)

        # 使用每行自己的仓库代码匹配报价单（多发票时各行仓库可能不同）
        row_wh = row.get('warehouse', warehouse)
        row_svc = row.get('service', service)
        q_info = quotation_data.get(row_wh, {})

        out = {
            'so_no': so_no,
            'service': row_svc,
            'country': country_from_warehouse(row_wh),
            'warehouse': row_wh,
            'e_price': q_info.get('e_price', ''),
            'f_price': q_info.get('f_price', ''),
            'supplier_ch': q_info.get('supplier_ch', ''),
            'ship_name': '', 'customs': '',
            'fba_id': fba_id,
            'cn_name': cn_name,
            'en_name': en_name,
            'pickup_fee': '',
            'box_count': box_count,
            'weight': weight, 'length': length, 'width': width, 'height': height,
            'history_note': '',
            'ref_w': hm['rw'] if hm else None,
            'ref_l': hm['rl'] if hm else None,
            'ref_wid': hm['rwid'] if hm else None,
            'ref_h': hm['rh'] if hm else None,
        }
        output_rows.append(out)
        if hm is None:
            missing_history.append(out)

    total_boxes = sum(r['box_count'] for r in output_rows)
    print(f"  📊 输出行: {len(output_rows)} 行, 总箱数: {total_boxes}")
    print(f"  ⚠️  无历史匹配: {len(missing_history)} 行")

    # ── 合并重复行（海外仓快递派"单行单箱"：FBA一致时按品名/实重/长宽高区分，重复合并）──
    output_rows = merge_duplicate_picking_rows(output_rows)

    # ── 写入模板（复用内部的写入逻辑）──
    _write_output_to_template(output_rows, template_file, output_path)
    return output_path, total_boxes


# ── CLI 入口 ──

def main(invoice_file=None, system_file=None, output_file=None):
    """CLI 主入口"""
    if invoice_file is None:
        files = sorted([
            os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR)
            if f.startswith("(测试用发票)") and f.endswith('.xlsx')
        ])
        if not files:
            print("❌ 未找到测试发票文件")
            return None, None, None
        invoice_file = files[-1]

    if system_file is None:
        files = sorted([
            os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR)
            if f.startswith("(测试用)导出拣货数据") and f.endswith('.xlsx')
        ])
        if not files:
            print("❌ 未找到系统导出拣货数据文件")
            return None, None, None
        system_file = files[-1]

    # 先处理，获取总箱数（用临时文件名）
    import re
    date_str = ''
    base = os.path.basename(system_file)
    m = re.search(r'(\d{4}-\d{2}-\d{2})', base)
    if m:
        date_str = m.group(1)

    tmp_path = os.path.join(DATA_DIR, '_temp_output.xlsx')
    result_path, total_boxes = generate_picking_output(invoice_file, system_file, tmp_path)

    # 生成带日期+箱数的文件名并重命名
    if output_file is None:
        output_name = f"内部拣货数据参考值_{date_str}_{total_boxes}箱.xlsx"
        output_file = os.path.join(DATA_DIR, output_name)

    if os.path.exists(result_path):
        os.rename(result_path, output_file)
        print(f"  ✅ 最终文件: {output_file}")

    return invoice_file, system_file, output_file

    print(f"📂 发票: {os.path.basename(invoice_file)}")
    print(f"📂 系统导出拣货数据: {os.path.basename(system_file)}")
    print(f"📂 输出: {os.path.basename(output_file)}")
    print()

    result = generate_picking_output(invoice_file, system_file, output_file)
    return invoice_file, system_file, result


if __name__ == '__main__':
    if len(sys.argv) > 1:
        if len(sys.argv) >= 4:
            main(sys.argv[1], sys.argv[2], sys.argv[3])
        elif len(sys.argv) >= 3:
            main(sys.argv[1], sys.argv[2])
        else:
            main(sys.argv[1])
    else:
        main()
