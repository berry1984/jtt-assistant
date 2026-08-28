#!/usr/bin/env python3
"""
TR发票 → 供应商发票 转换工具

功能：将TR/思锐/赛诺吉下单发票(.xlsx)转换为以下供应商格式：
  1. 天图 (--to 天图)
  2. 航乐英国 (--to 航乐-uk)
  3. 航乐欧洲 (--to 航乐-eu)
  4. 美琦美线 (--to 美琦)

用法：
  python3 convert_invoice.py <TR发票.xlsx> --to 天图 [输出路径]
  python3 convert_invoice.py <TR发票.xlsx> --to 航乐-uk [输出路径]
  python3 convert_invoice.py <TR发票.xlsx> --to 航乐-eu [输出路径]
  python3 convert_invoice.py <TR发票.xlsx> --to 美琦 [输出路径]

源文件兼容：
  - TR系统下单发票  ✅
  - 思锐客户下单发票 ✅
  - 赛诺吉发票      ✅

TR源文件结构：Sheet=Page1, 头部Row 1-16, 数据Row 18+
"""

import sys
import os
import re
import shutil
import argparse
import zipfile
import tempfile
import xml.etree.ElementTree as ET
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill
)
from openpyxl.utils import get_column_letter
from io import BytesIO


# ═══════════════════════════════════════════════════════════════
#  常量
# ═══════════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

TIANTU_TEMPLATE = os.path.join(SCRIPT_DIR, '天图单票专用模板20260601.xlsx')
HANGLE_UK_TEMPLATE = os.path.join(SCRIPT_DIR, '航乐-客户名称 客户单号 英国发票模板9.9更新.xls')
HANGLE_EU_TEMPLATE = os.path.join(SCRIPT_DIR, '航乐-客户单号- 欧州发票模板2.26更新.xls')
HANGLE_UK_TEMPLATE_XLSX = os.path.join(SCRIPT_DIR, '航乐-客户名称 客户单号 英国发票模板9.9更新.xlsx')
HANGLE_EU_TEMPLATE_XLSX = os.path.join(SCRIPT_DIR, '航乐-客户单号- 欧州发票模板2.26更新.xlsx')
MEIQI_TEMPLATE = os.path.join(SCRIPT_DIR, '美琦美线发票模版.xlsx')

# 美琦渠道映射：JTT/客户渠道名称 → 美琦服务渠道名称（服务渠道 sheet B 列下拉清单）
# 未命中映射的渠道保留源名称，并自动追加到下拉清单。
MEIQI_CHANNEL_MAP = {
    # '美国超快线-卡派': 'Match15-卡派',   # 按需补充已知对应关系
}

# TR发票列映射 (Row 17+ header)
TR_COL = {
    'A': '箱号', 'B': '重量_KG', 'C': '长度_CM', 'D': '宽度_CM', 'E': '高度_CM',
    'F': '英文品名', 'G': '中文品名', 'H': '单价_USD', 'I': '数量',
    'J': '材质', 'K': '海关编码', 'L': '用途', 'M': '品牌', 'N': '型号',
    'O': '链接', 'P': '销售价格', 'Q': '产品图片', 'R': '产品重量_KG',
    'S': 'ASIN', 'T': 'FNSKU', 'U': 'SKU', 'V': 'PO_Number',
}

# ═══════════════════════════════════════════════════════════════
#  1. TR发票读取
# ═══════════════════════════════════════════════════════════════

class TRInvoice:
    """解析TR/思锐/赛诺吉下单发票

    支持格式：
      - TR系统下单发票：工作表 Page1，Row 1-16 头部，Row 18+ 数据
      - 思锐/下单发票：工作表 发票，Row 1-26 头部，Row 28+ 数据
      自动检测格式。
    """

    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path, data_only=True)
        # 自动检测工作表
        if 'Page1' in self.wb.sheetnames:
            self.ws = self.wb['Page1']
        elif '发票' in self.wb.sheetnames:
            self.ws = self.wb['发票']
        else:
            self.ws = self.wb[self.wb.sheetnames[0]]
        self.header = {}       # 头部字段 dict
        self.data_rows = []    # 数据行列表
        self.images = {}       # {行号: image_bytes} 提取的嵌入图片
        self._parse()

    def _cell_str(self, row, col):
        v = self.ws.cell(row=row, column=col).value
        if v is None:
            return ''
        return str(v).strip()

    def _cell_val(self, row, col):
        return self.ws.cell(row=row, column=col).value

    def _parse_header_pair(self, row, label_col, val_col, label_key, val_key):
        """解析一对 标签列:值列，如 A:B, E:F"""
        label = self._cell_str(row, label_col)
        val = self._cell_val(row, val_col)
        if label:
            self.header[label.rstrip('*:')] = val
        return val

    def _parse(self):
        """自动检测发票格式并调用对应解析器"""
        # 检测格式：发票/下单格式的 A1 单元格为 '服务*'
        first_label = self._cell_str(1, 1)
        if first_label == '服务*':
            self._parse_fapiao_format()
        else:
            self._parse_tr_format()

    def _parse_tr_format(self):
        """TR系统格式：工作表 Page1，Row 1-16 头部，Row 18+ 数据"""
        ws = self.ws

        # ── 头部 Row 1-16 ──
        # 左侧 A/B 列
        for r in range(1, 17):
            a = self._cell_str(r, 1)
            b = self._cell_val(r, 2)
            if a:
                self.header[a.rstrip('*:')] = b

        # 右侧 E/F 列 (覆盖写入，优先级高于左侧)
        for r in range(1, 17):
            e = self._cell_str(r, 5)
            f = self._cell_val(r, 6)
            if e:
                self.header[e.rstrip('*:')] = f

        # 右侧 I/J 列
        for r in range(1, 17):
            i = self._cell_str(r, 9)
            j = self._cell_val(r, 10)
            if i:
                self.header[i.rstrip('*:')] = j

        # ── 数据行 Row 18+ ──
        max_row = ws.max_row
        for r in range(18, max_row + 1):
            box_no = self._cell_val(r, 1)
            if box_no is None or str(box_no).strip() == '':
                continue
            row_data = {
                'A': box_no,                     # 箱号
                'B': self._cell_val(r, 2),       # 重量
                'C': self._cell_val(r, 3),       # 长度
                'D': self._cell_val(r, 4),       # 宽度
                'E': self._cell_val(r, 5),       # 高度
                'F': self._cell_val(r, 6),       # 英文品名
                'G': self._cell_val(r, 7),       # 中文品名
                'H': self._cell_val(r, 8),       # 单价
                'I': self._cell_val(r, 9),       # 数量
                'J': self._cell_val(r, 10),      # 材质
                'K': self._cell_val(r, 11),      # 海关编码
                'L': self._cell_val(r, 12),      # 用途
                'M': self._cell_val(r, 13),      # 品牌
                'N': self._cell_val(r, 14),      # 型号
                'O': self._cell_val(r, 15),      # 链接
                'P': self._cell_val(r, 16),      # 销售价格
                'Q': self._cell_val(r, 17),      # 产品图片
                'R': self._cell_val(r, 18),      # 产品重量
                'S': self._cell_val(r, 19),      # ASIN
                'T': self._cell_val(r, 20),      # FNSKU
                'U': self._cell_val(r, 21),      # SKU
                'V': self._cell_val(r, 22),      # PO Number
                '_row': r,                       # 源文件行号（图片映射用）
            }
            self.data_rows.append(row_data)

        # ── 提取嵌入图片 (来自Q列/col 17) ──
        if hasattr(ws, '_images') and ws._images:
            for img in ws._images:
                try:
                    anchor = img.anchor
                    if hasattr(anchor, '_from'):
                        img_col = anchor._from.col + 1
                        img_row = anchor._from.row + 1
                        # Q列=17，数据行范围18+
                        if img_col == 17 and img_row >= 18:
                            # 获取图片字节数据
                            img_data = None
                            ref = getattr(img, 'ref', None)
                            if isinstance(ref, BytesIO):
                                img_data = ref.getvalue()
                            elif isinstance(ref, str):
                                # 通过关系ID从workbook获取
                                if hasattr(self.wb, '_images') and ref in self.wb._images:
                                    part = self.wb._images[ref]
                                    if hasattr(part, '_blob'):
                                        img_data = part._blob
                            if img_data is None:
                                data = img._data()
                                if isinstance(data, bytes):
                                    img_data = data
                                elif isinstance(data, BytesIO):
                                    img_data = data.getvalue()
                            if img_data:
                                self.images[img_row] = img_data
                except Exception:
                    pass

    # ─────────────────────────────────────────────────────────
    #  发票/下单 格式解析
    #  工作表：发票，Row 1-26 头部，Row 27 列头，Row 28+ 数据
    # ─────────────────────────────────────────────────────────

    def _parse_fapiao_format(self):
        """解析 发票/下单 格式"""
        ws = self.ws

        # ── 头部 Row 1-26：A列=标签，B列=值 ──
        import re as _re
        for r in range(1, 27):
            label = self._cell_str(r, 1)
            val = self._cell_val(r, 2)
            if label:
                # 清理标签：去掉尾部 *:、去掉尾部（中文括号注释）
                clean = _re.sub(r'[（(][^）)]*[）)]$', '', label).strip().rstrip('*:')
                self.header[clean] = val

        # 补充：Row 14 C列可能是客户订单号（本文件为空，保留）
        # 数据行中的 PO Number (R列) 会被提取为 dr['V']

        # ── 数据行 Row 28+（Row 27 是列标题） ──
        max_row = ws.max_row
        for r in range(28, max_row + 1):
            box_no = self._cell_val(r, 1)
            if box_no is None or str(box_no).strip() == '':
                continue

            # 列映射：发票格式列位置 → TR 内部字段 A-V
            # A(1)=货箱编号, B(2)=英文品名, C(3)=中文品名, D(4)=数量,
            # E(5)=单价, F(6)=重量, G(7)=长, H(8)=宽, I(9)=高,
            # J(10)=海关编码, K(11)=品牌, L(12)=材质, M(13)=型号,
            # N(14)=用途, O(15)=链接, P(16)=图片(DISPIMG), Q(17)=SKU,
            # R(18)=PO, S(19)=ASIN
            row_data = {
                'A': box_no,                                # 货箱编号
                'B': self._cell_val(r, 6),                  # F: 重量
                'C': self._cell_val(r, 7),                  # G: 长度
                'D': self._cell_val(r, 8),                  # H: 宽度
                'E': self._cell_val(r, 9),                  # I: 高度
                'F': self._cell_val(r, 2),                  # B: 英文品名
                'G': self._cell_val(r, 3),                  # C: 中文品名
                'H': self._cell_val(r, 5),                  # E: 单价
                'I': self._cell_val(r, 4),                  # D: 数量
                'J': self._cell_val(r, 12),                 # L: 材质
                'K': self._cell_val(r, 10),                 # J: 海关编码
                'L': self._cell_val(r, 14),                 # N: 用途
                'M': self._cell_val(r, 11),                 # K: 品牌
                'N': self._cell_val(r, 13),                 # M: 型号
                'O': self._cell_val(r, 15),                 # O: 链接
                'P': None,                                  # P: 销售价格(本格式无此列)
                'Q': self._cell_val(r, 16),                 # P: 产品图片(DISPIMG公式)
                'R': None,                                  # 产品重量(本格式无此列)
                'S': self._cell_val(r, 19),                 # S: ASIN
                'T': None,                                  # FNSKU(本格式无此列)
                'U': self._cell_val(r, 17),                 # Q: SKU
                'V': self._cell_val(r, 18),                 # R: PO Number
                '_row': r,
            }
            self.data_rows.append(row_data)

        # ── 提取嵌入图片 (WPS cellimages.xml 格式) ──
        self._extract_wps_cell_images()

    # ─────────────────────────────────────────────────────────
    #  WPS 图片提取（cellimages.xml 格式，配合 DISPIMG 公式）
    # ─────────────────────────────────────────────────────────

    def _extract_wps_cell_images(self):
        """从 WPS cellimages.xml 格式提取嵌入图片

        WPS Office 使用 xl/cellimages.xml + xl/_rels/cellimages.xml.rels
        存储"放置在单元格中"的图片。DISPIMG 公式引用图片 ID，此方法
        将 ID 映射回实际图片字节并存入 self.images。
        """
        import zipfile
        from xml.etree import ElementTree as ET

        try:
            with zipfile.ZipFile(self.path, 'r') as z:
                if 'xl/cellimages.xml' not in z.namelist():
                    return
                if 'xl/_rels/cellimages.xml.rels' not in z.namelist():
                    return

                # 读取关系映射: rId → media 图片文件
                rels_xml = z.read('xl/_rels/cellimages.xml.rels')
                rels_root = ET.fromstring(rels_xml)
                rid_to_target = {}
                for rel in rels_root:
                    rid = rel.get('Id', '')
                    target = rel.get('Target', '')
                    if rid and target:
                        rid_to_target[rid] = target

                # 读取 cellimages.xml → 图片 ID → rId → 实际文件
                ci_xml = z.read('xl/cellimages.xml')
                ci_root = ET.fromstring(ci_xml)
                NS_PIC = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
                NS_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'

                # 收集 DISPIMG ID → image_bytes 的映射
                wps_images = {}  # {dispimg_id: bytes}
                for ci in ci_root.iter(f'{{{NS_PIC}}}pic'):
                    nvPicPr = ci.find(f'{{{NS_PIC}}}nvPicPr')
                    if nvPicPr is None:
                        continue
                    cNvPr = nvPicPr.find(f'{{{NS_PIC}}}cNvPr')
                    if cNvPr is None:
                        continue
                    img_name = cNvPr.get('name', '')  # e.g., ID_A7244BAA5ACE43...

                    blipFill = ci.find(f'{{{NS_PIC}}}blipFill')
                    if blipFill is None:
                        continue
                    blip = blipFill.find(f'{{{NS_A}}}blip')
                    if blip is None:
                        continue
                    embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
                    if not embed:
                        continue

                    img_target = rid_to_target.get(embed)
                    if not img_target:
                        continue

                    img_path = f'xl/{img_target}' if not img_target.startswith('xl/') else img_target
                    try:
                        img_bytes = z.read(img_path)
                        wps_images[img_name] = img_bytes
                    except KeyError:
                        continue

                # 将图片映射到数据行
                for dr in self.data_rows:
                    q_val = dr.get('Q', '') or ''
                    q_str = str(q_val)
                    # DISPIMG 公式形如: =DISPIMG("ID_xxx",1)
                    # 或 =_xlfn.DISPIMG("ID_xxx",1)
                    import re
                    m = re.search(r'DISPIMG\("([^"]+)"', q_str)
                    if not m:
                        continue
                    dispimg_id = m.group(1)
                    if dispimg_id in wps_images:
                        src_row = dr.get('_row')
                        if src_row:
                            self.images[src_row] = wps_images[dispimg_id]
        except Exception as e:
            print(f'  ⚠️  提取 WPS 图片失败: {e}')

    def get(self, key, default=None):
        """获取头部字段值"""
        return self.header.get(key, default)

    def __repr__(self):
        info = f'TRInvoice({os.path.basename(self.path)})\n'
        info += f'  订单号: {self.get("客户订单号")}\n'
        info += f'  服务: {self.get("服务")}\n'
        info += f'  收件人: {self.get("收件人姓名")}\n'
        info += f'  箱数: {self.get("箱数")}\n'
        info += f'  数据行: {len(self.data_rows)} 行\n'
        return info


# ═══════════════════════════════════════════════════════════════
#  2. 天图 转换
# ═══════════════════════════════════════════════════════════════

def convert_to_tiantu(tr, output_path, image_url_base=None, order_list_path=None):
    """
    TR发票 → 天图格式
    规则详见 TR转天图发票_转换规则说明.md

    image_url_base: 如果提供，则为 IMAGE 公式的 URL 前缀（如 http://host/temp/xxx），
                    实现 Excel 365 "放置在单元格中"。
                    如果不提供，回退到文本链接。
    order_list_path: 如果提供，按对应订单号匹配订单列表——运单号回填 B14 客户订单号，
                     并抓取「供应商服务」回填 B1 服务（美琦同名规则）。
    """
    print(f'📄 天图模板: {TIANTU_TEMPLATE}')
    if not os.path.exists(TIANTU_TEMPLATE):
        print('❌ ERROR: 天图模板文件不存在!')
        return False

    # 输出模板版本日志
    tpl_size = os.path.getsize(TIANTU_TEMPLATE)
    print(f'  模板大小: {tpl_size} bytes')
    # 检查新模板特征：Sheet2行数
    try:
        tmp_wb = load_workbook(TIANTU_TEMPLATE, data_only=True)
        s2_rows = tmp_wb['Sheet2'].max_row
        print(f'  模板Sheet2服务数: {s2_rows} 条')
        tmp_wb.close()
    except:
        pass

    # 复制模板
    shutil.copy(TIANTU_TEMPLATE, output_path)
    wb = load_workbook(output_path)
    ws = wb['Sheet1']

    # ── 样式定义 ──
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    font_label_big = Font(name='微软雅黑', size=11, bold=True)
    font_label_small = Font(name='微软雅黑', size=10, bold=True)
    font_value = Font(name='微软雅黑', size=10)
    font_data = Font(name='微软雅黑', size=10)
    font_header = Font(name='微软雅黑', size=11, bold=True)
    font_total = Font(name='微软雅黑', size=11, bold=True)
    font_red = Font(name='微软雅黑', size=12, bold=True, color='FF0000')

    fill_header = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')

    # ── 辅助函数 ──
    def set_cell(col, row, value, font=None, align=None, border=None, number_format=None):
        cell = ws[f'{col}{row}']
        cell.value = value
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format
        return cell

    # ── 头部 Row 1-28 ──

    # B1: 服务 = 订单列表对应订单号的「供应商服务」（渠道抓取），未命中回退 TR 服务字段
    # 同时确保该服务在 Sheet2 下拉列表中
    service_name = _match_supplier_service(tr, order_list_path) or tr.get('服务', '')
    if service_name:
        # 检查 Sheet2 中是否已有该服务，没有则追加
        ws2 = wb['Sheet2']
        found = False
        for r in range(1, ws2.max_row + 1):
            if ws2.cell(row=r, column=1).value == service_name:
                found = True
                break
        if not found:
            next_row = ws2.max_row + 1
            ws2.cell(row=next_row, column=1).value = service_name
    set_cell('B', 1, service_name, font_value)

    # B2: 仓库代码 = TR B4 (收件人姓名，如 YVR4/FTW1/POZ1/RFD2)
    wh_code = tr.get('收件人姓名', '')
    set_cell('B', 2, wh_code, font_value)

    # B3: 收件人姓名 = TR B4
    set_cell('B', 3, wh_code, font_value)

    # B4: 收件人公司 = TR B5
    set_cell('B', 4, tr.get('收件人公司', ''), font_value)

    # B5: 收件人地址一 = TR B6
    set_cell('B', 5, tr.get('收件人地址一', ''), font_value)

    # B6: 收件人地址二 → 留空
    set_cell('B', 6, '', font_value)

    # B7: 收件人地址三 → 留空
    set_cell('B', 7, '', font_value)

    # B8: 收件人城市 = TR B9
    set_cell('B', 8, tr.get('收件人城市', ''), font_value)

    # B9: 收件人省份/州 = TR B10
    set_cell('B', 9, tr.get('收件人省份/州', ''), font_value)

    # B10: 收件人邮编 = TR B11
    set_cell('B', 10, tr.get('收件人邮编', ''), font_value)

    # B11: 收件人国家代码 = TR B12, 空则默认US
    country = tr.get('收件人国家代码(二字代码)', '') or tr.get('收件人国家代码', '') or 'US'
    set_cell('B', 11, country, font_value)

    # B12: 收件人电话 = TR B13
    set_cell('B', 12, tr.get('收件人电话', ''), font_value)

    # B13: 收件人邮箱 = TR B14
    set_cell('B', 13, tr.get('收件人邮箱', ''), font_value)

    # B14: 客户订单号 = TR B1；若提供订单列表，按地址库编码查运单号填入
    order_no = tr.get('客户订单号', '') or ''
    order_no = _match_waybill(tr, order_list_path) or order_no
    set_cell('B', 14, order_no, font_value)

    # B15: Amazon Reference ID → 留空
    set_cell('B', 15, '', font_value)

    # B16: 带电 → "带电"/"不带电"
    has_battery = tr.get('带电', '否')
    set_cell('B', 16, '带电' if has_battery == '是' else '不带电', font_value)

    # B17: 带磁 → "带磁"/"不带磁"
    has_magnet = tr.get('带磁', '否')
    set_cell('B', 17, '带磁' if has_magnet == '是' else '不带磁', font_value)

    # B18: 交税方式 = TR F8
    set_cell('B', 18, tr.get('交税方式', ''), font_value)

    # B19: 报关方式 = TR F6
    set_cell('B', 19, tr.get('报关方式', ''), font_value)

    # B20: 清关方式 = TR F7
    set_cell('B', 20, tr.get('清关方式', ''), font_value)

    # B21: VAT号 = TR E10
    set_cell('B', 21, tr.get('VAT号', ''), font_value)

    # B22: 总箱数 = TR B16
    set_cell('B', 22, tr.get('箱数', ''), font_value)

    # B23: 备注 → 留空
    set_cell('B', 23, '', font_value)

    # B24: 国内报关抬头信息补充 (A24:B24 合并, 保留原标题)
    # keep original

    # B25: 公司名称 (TR中没有，留空)
    set_cell('B', 25, '', font_value)

    # B26: 公司税号 (TR中没有，留空)
    set_cell('B', 26, '', font_value)

    # B27: 真实出口货值 (TR中没有，留空)
    set_cell('B', 27, '', font_value)

    # B28: 币种 = TR J14 (申报币种) 或 TR E14 (币种)
    currency = tr.get('申报币种', '') or tr.get('币种', '')
    if not currency:
        # 根据国家推断
        if country in ('GB', 'UK'):
            currency = '英镑 GBP'
        elif country == 'US':
            currency = '美元 USD'
        elif country in ('DE', 'FR', 'IT', 'ES', 'NL', 'BE', 'PL', 'CZ'):
            currency = '欧元 EUR'
        else:
            currency = '美元 USD'
    set_cell('B', 28, currency, font_value)

    # ── 数据行 Row 30+ ← TR Row 18+ ──
    # 先清除模板中已有的示例数据行
    for r in range(30, 100):
        for c in range(1, 19):  # A-R
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.border = Border()

    # 收集待嵌入的图片 (cell_ref → image_bytes)
    pending_images = {}

    # 写入数据
    data_start_row = 30
    for i, dr in enumerate(tr.data_rows):
        r = data_start_row + i
        ws.row_dimensions[r].height = 80  # 产品图片需要足够高度

        # A: 货箱编号
        set_cell('A', r, dr['A'], font_data, center_align, thin_border)

        # B: PO Number (优先行级V列，其次头部PO)
        po = dr.get('V') or tr.get('PO Number', '')
        set_cell('B', r, po, font_data, center_align, thin_border)

        # C: 产品英文品名
        set_cell('C', r, dr['F'], font_data, center_align, thin_border)

        # D: 产品中文品名
        set_cell('D', r, dr['G'], font_data, center_align, thin_border)

        # E: 产品申报单价(USD)
        set_cell('E', r, dr['H'] if dr['H'] is not None else '',
                 font_data, center_align, thin_border, '#,##0.00')

        # F: 产品单箱申报数量
        set_cell('F', r, dr['I'] if dr['I'] is not None else '',
                 font_data, center_align, thin_border, '0')

        # G: 产品单箱申报总价(USD) = E × F (计算值)
        e_val = dr['H'] or 0
        f_val = dr['I'] or 0
        g_val = e_val * f_val
        set_cell('G', r, g_val, font_data, center_align, thin_border, '#,##0.00')

        # H: 产品材质
        set_cell('H', r, dr['J'], font_data, center_align, thin_border)

        # I: 产品海关编码
        set_cell('I', r, dr['K'], font_data, center_align, thin_border)

        # J: 产品用途
        set_cell('J', r, dr['L'], font_data, center_align, thin_border)

        # K: 产品品牌
        set_cell('K', r, dr['M'], font_data, center_align, thin_border)

        # L: 产品型号
        set_cell('L', r, dr['N'], font_data, center_align, thin_border)

        # M: 产品图片 (优先嵌入图片，其次放链接文本)
        src_row = dr.get('_row')
        img_bytes = tr.images.get(src_row) if src_row else None
        if img_bytes:
            # 收集图片，保存后会通过 IMAGE() 公式嵌入单元格
            pending_images[f'M{r}'] = img_bytes
            # 占位文本（Excel 365 会以 IMAGE 公式结果覆盖显示）
            set_cell('M', r, '[图片]', font_data, center_align, thin_border)
        q_val = dr.get('Q', '') or ''
        if q_val and not img_bytes:
            set_cell('M', r, q_val, font_data, center_align, thin_border)

        # N: 产品销售链接
        set_cell('N', r, dr['O'] if dr['O'] else '',
                 font_data, center_align, thin_border)

        # O: 货箱重量(KG)
        set_cell('O', r, dr['B'] if dr['B'] is not None else '',
                 font_data, center_align, thin_border, '0.0')

        # P: 货箱长度(CM)
        set_cell('P', r, dr['C'] if dr['C'] is not None else '',
                 font_data, center_align, thin_border, '0.0')

        # Q: 货箱宽度(CM)
        set_cell('Q', r, dr['D'] if dr['D'] is not None else '',
                 font_data, center_align, thin_border, '0.0')

        # R: 货箱高度(CM)
        set_cell('R', r, dr['E'] if dr['E'] is not None else '',
                 font_data, center_align, thin_border, '0.0')

    num_data_rows = len(tr.data_rows)

    # ── 合计行 ──
    total_row = data_start_row + num_data_rows + 1  # 空一行 + 合计行

    # Row 29是表头，保持不动

    set_cell('A', total_row, '合计', font_total, center_align, thin_border)

    # F: 总数量
    total_qty = sum(dr['I'] or 0 for dr in tr.data_rows)
    set_cell('F', total_row, total_qty, font_total, center_align, thin_border, '0')

    # G: 总价
    total_value = sum((dr['H'] or 0) * (dr['I'] or 0) for dr in tr.data_rows)
    set_cell('G', total_row, total_value, font_total, center_align, thin_border, '#,##0.00')

    # 其他列加边框
    for col_l in ['B', 'C', 'D', 'E', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        cell = ws[f'{col_l}{total_row}']
        cell.border = thin_border
        cell.font = font_total
        cell.alignment = center_align

    # ── 清理模板中的VLOOKUP公式 (B3-B5, B8-B10, B12) ──
    # 已经在前面被覆盖了

    # ── 保存 ──
    wb.save(output_path)

    # 后处理：将图片嵌入为单元格 IMAGE() 公式（Excel 365 "放置在单元格中"）
    if pending_images:
        _embed_images_as_cell_images(output_path, pending_images, image_url_base)

    print(f'✅ 天图发票已生成: {os.path.basename(output_path)}')
    print(f'   数据: {num_data_rows} 行, 合计: {total_qty} 件, ${total_value:.2f}')
    if pending_images:
        print(f'   嵌入图片: {len(pending_images)} 张')
    return True


# ═══════════════════════════════════════════════════════════════
#  3. 航乐 转换 (通用)
# ═══════════════════════════════════════════════════════════════

def convert_to_hangle(tr, output_path, region='uk', image_url_base=None, order_list_path=None):
    """
    TR发票 → 航乐格式 (UK 或 EU)

    基于实际模板 (.xlsx 转换版)，复制后填充数据。
    保留模板的所有格式、合并单元格、渠道参考列表等。

    image_url_base: 如果提供，则嵌入 IMAGE() 公式实现"放置在单元格中"图片效果。
    order_list_path: 如果提供，按对应订单号匹配订单列表——运单号用于输出文件名（app.py 拼装），
                     并抓取「供应商服务」回填 I7 渠道（美琦同名规则）。
                     航乐模板正文无客户订单号字段。
    """
    # ── 选择模板 ──
    if region == 'uk':
        template_path = HANGLE_UK_TEMPLATE_XLSX
        currency_label = 'GBP'
        region_label = '英国'
    else:
        template_path = HANGLE_EU_TEMPLATE_XLSX
        currency_label = 'EUR'
        region_label = '欧洲'

    if not os.path.exists(template_path):
        # 回退到旧版 .xls 路径提示
        print(f'❌ 航乐{region_label}模板 (.xlsx) 不存在: {template_path}')
        print(f'   请先运行: python3 convert_template_xls_to_xlsx.py')
        return False

    print(f'📄 航乐{region_label}模板: {template_path}')

    # ── 获取 TR 数据 ──
    country = tr.get('收件人国家代码(二字代码)', '') or tr.get('收件人国家代码', '') or 'US'
    country_map = {
        'GB': 'UK', 'UK': 'UK', 'US': 'US', 'DE': 'DE', 'FR': 'FR',
        'IT': 'IT', 'ES': 'ES', 'NL': 'NL', 'BE': 'BE', 'PL': 'PL',
        'CZ': 'CZ', 'CA': 'CA',
    }
    country_code = country_map.get(country.upper(), country.upper())
    has_battery = tr.get('带电', '否')
    has_magnet = tr.get('带磁', '否')

    # ── 解析箱号中的箱数（如 "FBA15LV645QTU000001-5" → 5 箱）──
    import re
    def parse_box_count(box_no):
        if not box_no:
            return 1
        m = re.search(r'-(\d+)$', str(box_no))
        return int(m.group(1)) if m else 1

    # ── 复制模板 ──
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['Packing list装箱单发票']

    # 按美琦运单号匹配规则匹配运单号（航乐模板正文无客户订单号字段，结果供 app.py 拼文件名用）
    _match_waybill(tr, order_list_path)

    # ── 样式定义（匹配航乐模板样式）──
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    val_font = Font(name='微软雅黑', size=11)
    bold_font = Font(name='微软雅黑', size=10, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def wcell(r, c, val, font=None, align=None, border=None, nf=None):
        cell = ws.cell(row=r, column=c)
        cell.value = val
        if font: cell.font = font
        if align: cell.alignment = align
        if border: cell.border = border
        if nf: cell.number_format = nf
        return cell

    # ══════════════════════════════════════════════
    #  头部数据填充——模板已含标签，只需填值并匹配格式
    # ══════════════════════════════════════════════

    # ── Row 3: ADD / COUNTRY / SHIPMENT ID / AMAZON REF ──
    # 注：不传font以保留模板原有格式（宋体红色等）
    # B3:E3 (merged) = ADD(收件地址) 值
    ws['B3'] = tr.get('收件人姓名', '')
    # G3 = COUNTRY 值
    ws['G3'] = country_code
    # H3:J3 (merged) = SHIPMENT ID 标签（保留模板标签，值不填）
    # K3:M3 (merged) = AMAZON REF 标签（保留模板标签）

    # ── Row 4: ZIP / COMPANY / TEL / 报关类型 / 交货仓库 ──
    ws['B4'] = tr.get('收件人邮编', '')
    # D4:E4 (merged) = COMPANY
    ws['D4'] = tr.get('收件人公司', '')
    # G4:H4 (merged) = TEL
    ws['G4'] = tr.get('收件人电话', '')
    # I4:J4 (merged) = 报关类型标签（保留模板标签，值写到 Row5 I5）
    # K4:M4 (merged) = 交货仓库标签（保留模板标签，值写到 Row5 K5）

    # ── Row 5: CITY / ATTN / EMAIL / 报关退税 / 交货仓库 ──
    ws['B5'] = tr.get('收件人城市', '')
    # D5:E5 (merged) = ATTN
    ws['D5'] = tr.get('收件人姓名', '')
    # G5:H5 (merged) = EMAIL
    ws['G5'] = tr.get('收件人邮箱', '')
    # I5:J5 (merged) = 报关类型值
    ws['I5'] = tr.get('报关方式', '')
    # K5:M5 (merged) = 交货仓库值
    ws['K5'] = ''  # 按需填写

    # ── Row 7-10: VAT / 渠道 / 包税 / 物品属性 ──
    # B7:H7 = VAT公司名称/公司名称
    ws['B7'] = tr.get('VAT公司英文名', '')
    # B8:H8 = VAT号
    ws['B8'] = tr.get('VAT号', '')
    # B9:H9 = EORI号
    ws['B9'] = tr.get('EORI号', '')
    # B10:H10 = VAT注册地址
    ws['B10'] = tr.get('VAT注册地址', '')
    # I7:K10 (merged) = 渠道（服务名）—— 渠道抓取：订单列表对应订单号「供应商服务」优先，未命中回退 TR 服务
    ws['I7'] = _match_supplier_service(tr, order_list_path) or tr.get('服务', '')
    # L7:L10 (merged) = 是否包税
    ws['L7'] = tr.get('交税方式', '')
    # M7:M10 (merged) = 物品属性
    category = _guess_product_category(tr)
    ws['M7'] = category

    # 收集待嵌入图片
    pending_images = {}

    # ══════════════════════════════════════════════
    #  数据行 (Row 12+)
    # ══════════════════════════════════════════════

    # 清空模板数据区域 (row 12-23)
    for r in range(12, 50):
        for c in range(1, 30):
            cell = ws.cell(row=r, column=c)
            # 保留公式提示文字（合计行标签）
            if r == 24 and c == 3 and str(cell.value or '').strip() == 'Total No.of Boxes and weight':
                cell.value = None
            else:
                cell.value = None

    # 添加 PO Number 表头（模板不含，正确版本有此列）—— 匹配模板表头格式
    y11_cell = ws['Y11']
    y11_cell.value = 'PO Number*\n（Reference ID）'
    y11_cell.font = Font(name='宋体', size=15, color='FFFF0000')
    y11_cell.fill = PatternFill(patternType='solid', fgColor='FFFFFF00')
    y11_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_start = 12
    for i, dr in enumerate(tr.data_rows):
        r = data_start + i
        ws.row_dimensions[r].height = 30

        qty = dr['I'] or 0
        box_wt = dr['B'] or 0
        unit_price = dr['H'] or 0
        box_count = parse_box_count(dr['A'])
        total_qty = qty * box_count

        # 尺寸 — 取整（匹配正确版本做法）
        length_cm = int(dr['C']) if dr['C'] else 0
        width_cm = int(dr['D']) if dr['D'] else 0
        height_cm = int(dr['E']) if dr['E'] else 0

        # 材重除数：欧洲 6000，英国 5000
        vol_div = 6000 if region == 'eu' else 5000
        raw_vol_wt = (length_cm * width_cm * height_cm / vol_div) if length_cm and width_cm and height_cm else None
        raw_cbm = (length_cm * width_cm * height_cm / 1000000) if length_cm and width_cm and height_cm else None
        # 净重 = 实重 / 单箱数量（不是总数量）
        net_wt = round(box_wt / qty, 4) if box_wt and qty else None

        data = {
            1: str(dr['A']) if dr['A'] is not None else '',                      # A: Box No
            2: str(dr['G']) if dr['G'] else '',                                  # B: 品名中文
            3: str(dr['F']) if dr['F'] else '',                                  # C: 品名英文
            4: str(dr['J']) if dr['J'] else '',                                  # D: 材质中文
            5: str(dr['J']) if dr['J'] else '',                                  # E: 材质英文
            6: str(int(dr['K'])) if dr['K'] is not None else '',                 # F: 海关编码
            7: str(dr['N']) if dr['N'] else '无',                                # G: 型号
            8: str(dr['M']) if dr['M'] else '无品牌',                            # H: 品牌
            9: str(dr['L']) if dr['L'] else '',                                  # I: 用途
            10: qty,                                                             # J: 单箱数量
            11: total_qty,                                                       # K: 总数量
            12: net_wt,                                                          # L: 净重
            13: box_wt,                                                          # M: 实重
            14: unit_price,                                                      # N: 单价
            15: round(unit_price * total_qty, 2),                                # O: 总价
            16: length_cm,                                                       # P: 长
            17: width_cm,                                                        # Q: 宽
            18: height_cm,                                                       # R: 高
            19: raw_vol_wt,                                                       # S: 材重
            20: raw_cbm,                                                          # T: CBM
            21: '是' if has_battery == '是' else '否',                            # U: 是否带电
            22: '是' if has_magnet == '是' else '否',                             # V: 是否带磁
            23: str(dr.get('Q', '') or ''),                                      # W: 产品图片（嵌入）
            24: str(dr['O']) if dr['O'] else '',                                 # X: 链接
            25: str(dr.get('V', '') or ''),                                      # Y: PO Number
        }

        nf_map = {10: '0', 11: '0', 12: '0.0000', 13: '0.0',
                  14: '0.00', 15: '0.00', 19: '0.000', 20: '0.000000'}

        for col_idx, val in data.items():
            if val is None or val == '':
                continue
            wcell(r, col_idx, val, val_font, center_align, thin_border,
                  nf_map.get(col_idx, None))

        # 收集图片（W列 = 产品图片）
        src_row = dr.get('_row')
        img_bytes = tr.images.get(src_row) if src_row else None
        if img_bytes:
            col_w = chr(64 + 23)  # W
            pending_images[f'{col_w}{r}'] = img_bytes

    num_data = len(tr.data_rows)

    # ══════════════════════════════════════════════
    #  合计行
    # ══════════════════════════════════════════════
    total_row = data_start + num_data + 3  # 空3行间隙

    # 计算合计
    total_qty = sum((dr['I'] or 0) * parse_box_count(dr['A']) for dr in tr.data_rows)
    total_weight = sum(dr['B'] or 0 for dr in tr.data_rows)
    total_price = sum((dr['H'] or 0) * (dr['I'] or 0) * parse_box_count(dr['A']) for dr in tr.data_rows)

    vol_div = 6000 if region == 'eu' else 5000
    vol_wts = []
    cbms = []
    for dr in tr.data_rows:
        l = int(dr['C']) if dr['C'] else 0
        w = int(dr['D']) if dr['D'] else 0
        h = int(dr['E']) if dr['E'] else 0
        if l and w and h:
            vol_wts.append(l * w * h / vol_div)
            cbms.append(l * w * h / 1000000)
    total_vol_wt = round(sum(vol_wts), 3) if vol_wts else 0
    total_cbm = round(sum(cbms), 6) if cbms else 0

    # 写入合计
    ws.merge_cells(start_row=total_row, start_column=3, end_row=total_row, end_column=5)
    ws.cell(row=total_row, column=3).value = 'Total No.of Boxes and weight'

    total_fields = {
        11: (total_qty, '0'),
        13: (total_weight, '0.0'),
        15: (total_price, '0.00'),
        19: (total_vol_wt, '0.000'),
        20: (total_cbm, '0.000000'),
    }
    for col_idx, (val, nf) in total_fields.items():
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value = val
        cell.font = bold_font
        cell.alignment = center_align
        cell.number_format = nf

    # 合计行加粗+边框
    for c in range(1, 26):
        cell = ws.cell(row=total_row, column=c)
        if cell.value is not None:
            if cell.font and cell.font.name:
                cell.font = Font(name=cell.font.name, size=11, bold=True)
            else:
                cell.font = Font(name='宋体', size=11, bold=True)
        cell.alignment = center_align

    # ── 保存 ──
    wb.save(output_path)

    # 后处理：嵌入产品图片（IMAGE() 公式 + xl/media/ 双重）
    if pending_images:
        _embed_images_as_cell_images(output_path, pending_images, image_url_base)

    print(f'✅ 航乐{region_label}发票已生成: {os.path.basename(output_path)}')
    print(f'   数据: {num_data} 行, 总件数: {total_qty}, 总价: {total_price:.2f} {currency_label}')
    if pending_images:
        print(f'   嵌入图片: {len(pending_images)} 张')
    return True


def _parse_box_count(box_no):
    """解析箱号尾部的箱数（如 'FBA19CW3W9W7U000001-3' → 3）"""
    if not box_no:
        return 1
    m = re.search(r'-(\d+)$', str(box_no))
    return int(m.group(1)) if m else 1


def _format_hs_code(val):
    """海关编码格式化为 XXXX.XX.XXXX（美琦模版示例格式）。

    如 9405429000 → 9405.42.9000；非 10 位数字原样返回。
    """
    if val is None:
        return ''
    s = str(val).strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) == 10:
        return f'{digits[0:4]}.{digits[4:6]}.{digits[6:10]}'
    return s


def _extract_fba_id(box_no):
    """从货箱编号提取物品 FBA ID：'U00000' 之前的 12 个字符。

    如 'FBA19CW3W9W7U000001-3' → 'FBA19CW3W9W7'；未找到返回空串。
    """
    if not box_no:
        return ''
    s = str(box_no).strip()
    idx = s.find('U00000')
    if idx >= 12:
        return s[idx - 12:idx]
    return ''


# ═══════════════════════════════════════════════════════════════
#  3.5 美琦 美线 转换
# ═══════════════════════════════════════════════════════════════


def _load_order_list_map(order_list_path):
    """从订单列表 excel 构建 {地址库编码: 运单号} 映射。

    自动定位表头行：优先找含「运单号」的表头行；地址库编码列优先
    「地址库编码」，其次「仓库代码」/「仓库」。同一编码取第一行。
    """
    wb = load_workbook(order_list_path, data_only=True)
    ws = wb.worksheets[0]

    def find_col(header_vals, keywords):
        for idx, v in enumerate(header_vals):
            for kw in keywords:
                if kw in v:
                    return idx + 1  # 1-based 列号
        return None

    # 定位表头行（含「运单号」的行；找不到则取第 1 行）
    header_row = 1
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or '').strip()
                    for c in range(1, min(ws.max_column, 60) + 1)]
        if any('运单号' in v for v in row_vals):
            header_row = r
            break

    header_vals = [str(ws.cell(row=header_row, column=c).value or '').strip()
                   for c in range(1, min(ws.max_column, 60) + 1)]
    waybill_col = find_col(header_vals, ['运单号'])
    code_col = find_col(header_vals, ['地址库编码', '仓库代码', '仓库'])
    if not waybill_col or not code_col:
        print(f'   ⚠️ 订单列表未找到「运单号」或「地址库编码」列: {os.path.basename(order_list_path)}')
        return {}

    mapping = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(row=r, column=code_col).value
        waybill = ws.cell(row=r, column=waybill_col).value
        if code is None or waybill is None:
            continue
        code_s = str(code).strip()
        waybill_s = str(waybill).strip()
        if code_s and waybill_s and code_s not in mapping:
            mapping[code_s] = waybill_s
    print(f'   📋 订单列表: {os.path.basename(order_list_path)}，地址库编码 {len(mapping)} 条')
    return mapping


def _load_order_rows(order_list_path):
    """从订单列表 excel 解析全部数据行，返回 [{'运单号','仓库代码','供应商服务'}, ...]。

    自动定位表头行（含「运单号」）；未提供「供应商服务」列时该项为空串。
    """
    wb = load_workbook(order_list_path, data_only=True)
    ws = wb.worksheets[0]

    header_row = 1
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or '').strip()
                    for c in range(1, min(ws.max_column, 60) + 1)]
        if any('运单号' in v for v in row_vals):
            header_row = r
            break

    header_vals = [str(ws.cell(row=header_row, column=c).value or '').strip()
                   for c in range(1, min(ws.max_column, 60) + 1)]

    def find_col(keywords):
        for idx, v in enumerate(header_vals):
            for kw in keywords:
                if kw in v:
                    return idx + 1  # 1-based 列号
        return None

    waybill_col = find_col(['运单号'])
    code_col = find_col(['地址库编码', '仓库代码', '仓库'])
    service_col = find_col(['供应商服务'])
    if not waybill_col:
        print(f'   ⚠️ 订单列表未找到「运单号」列: {os.path.basename(order_list_path)}')
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        waybill = ws.cell(row=r, column=waybill_col).value
        if waybill is None or not str(waybill).strip():
            continue
        rows.append({
            '运单号': str(waybill).strip(),
            '仓库代码': str(ws.cell(row=r, column=code_col).value or '').strip() if code_col else '',
            '供应商服务': str(ws.cell(row=r, column=service_col).value or '').strip() if service_col else '',
        })
    print(f'   📋 订单列表: {os.path.basename(order_list_path)}，{len(rows)} 行数据')
    return rows


def _match_order_row(tr, order_list_path):
    """从订单列表匹配 TR 发票对应的行。

    匹配优先级：
      1. 订单号匹配：TR「客户订单号」== 订单列表「运单号」
      2. 仓库代码匹配：TR「地址库编码」（为空取「收件人姓名」）== 订单列表「仓库代码」
    返回命中的行 dict（含「运单号」「供应商服务」）；未命中返回 None。
    美琦/天图/航乐 三家共用。
    """
    if not order_list_path:
        return None
    rows = _load_order_rows(order_list_path)
    if not rows:
        return None

    # 1. 订单号（运单号）匹配
    order_no = (tr.get('客户订单号', '') or '').strip()
    if order_no:
        for row in rows:
            if row.get('运单号') == order_no:
                return row

    # 2. 仓库代码（地址库编码）匹配
    wh_code = (tr.get('地址库编码', '') or tr.get('收件人姓名', '') or '').strip()
    if wh_code:
        for row in rows:
            if row.get('仓库代码') == wh_code:
                return row

    return None


def _match_waybill(tr, order_list_path):
    """美琦运单号匹配规则：从订单列表按对应订单号查运单号。

    返回匹配到的运单号；未提供订单列表/未命中 返回 None。
    美琦/天图/航乐 三家共用，客户订单号在各自模板中的落点由调用方决定。
    """
    row = _match_order_row(tr, order_list_path)
    if not row:
        wh_code = (tr.get('地址库编码', '') or tr.get('收件人姓名', '') or '').strip()
        if wh_code:
            print(f'   ⚠️ 订单列表未找到匹配 {wh_code} 的行，客户订单号保留源值')
        return None
    waybill = row.get('运单号') or None
    if waybill:
        print(f'   📋 订单列表命中: → 运单号 {waybill}')
    return waybill


def _match_supplier_service(tr, order_list_path):
    """渠道抓取规则：目标发票的「服务/渠道」从订单列表对应订单号抓取「供应商服务」。

    返回匹配到的供应商服务；未提供订单列表/未命中/该列为空 返回 None，
    调用方回退到 TR 源「服务」字段。
    """
    if not order_list_path:
        return None
    row = _match_order_row(tr, order_list_path)
    if not row:
        return None
    svc = row.get('供应商服务', '') or ''
    if svc:
        print(f'   📋 供应商服务命中: {svc}')
        return svc
    print(f'   ⚠️ 订单列表命中行无「供应商服务」列值，服务保留源值')
    return None


def convert_to_meiqi(tr, output_path, order_list_path=None):
    """
    TR发票 → 美琦美线发票格式

    新版美琦模版（美琦美线发票模版.xlsx）：
      - 头部 Row 1-17：A 列=标签，B 列=值
      - 数据列头 Row 18：A-R（货箱编号/品名英/品名中/数量/单价/重量/长宽高/
        海关编码/品牌/材质/型号/用途/产品图片/PO Number/物品箱号/物品FBA ID），
        数据行从 Row 19 起
    收件人信息从「亚马逊仓库代码」sheet 按地址库编码查表。
    产品图片参考天图从源发票提取并嵌入 O 列。
    若提供 order_list_path（订单列表 excel），按地址库编码查运单号填入 B1 客户订单号。
    """
    if not os.path.exists(MEIQI_TEMPLATE):
        print(f'❌ 美琦模板不存在: {MEIQI_TEMPLATE}')
        return False

    print(f'📄 美琦模板: {MEIQI_TEMPLATE}')

    # ── 复制模板 ──
    shutil.copy(MEIQI_TEMPLATE, output_path)
    wb = load_workbook(output_path)
    ws = wb['发票']

    # ── 构建 亚马逊仓库代码 查表: {地址编码: {...}} ──
    address_lib = {}
    ws_addr = wb['亚马逊仓库代码']
    for r in range(2, ws_addr.max_row + 1):
        code = ws_addr.cell(row=r, column=1).value
        if not code or not str(code).strip():
            continue
        address_lib[str(code).strip()] = {
            '联系人': ws_addr.cell(row=r, column=3).value,   # C: 收件人姓名
            '公司名': ws_addr.cell(row=r, column=4).value,   # D: 公司名
            '地址一': ws_addr.cell(row=r, column=5).value,   # E: 地址一
            '城市': ws_addr.cell(row=r, column=6).value,     # F: 城市
            '省洲': ws_addr.cell(row=r, column=7).value,     # G: 省/洲
            '国家': ws_addr.cell(row=r, column=8).value,     # H: 国家
            '邮编': ws_addr.cell(row=r, column=9).value,     # I: 邮编
        }

    def set_val(row, val):
        ws[f'B{row}'] = val if val is not None else ''

    # ══════════════════════════════════════════════
    #  头部 Row 1-17
    # ══════════════════════════════════════════════

    # 有效地址库编码 = 源地址库编码 或 收件人姓名（仓库代码，如 IND9）
    wh_code = (tr.get('地址库编码', '') or tr.get('收件人姓名', '') or '').strip()

    # B1: 客户订单号 = 源客户订单号；若提供订单列表，按地址库编码查运单号填入
    order_no = tr.get('客户订单号', '') or ''
    b1_val = _match_waybill(tr, order_list_path) or order_no
    set_val(1, b1_val)

    # B2: 客户参考号 = 源客户参考号（无则取客户订单号）
    set_val(2, tr.get('客户参考号', '') or order_no)

    # B3: 服务 —— 渠道抓取：订单列表对应订单号「供应商服务」优先，未命中回退 TR 服务（渠道映射；未命中保留源名称）
    service = _match_supplier_service(tr, order_list_path) or tr.get('服务', '') or ''
    meiqi_service = MEIQI_CHANNEL_MAP.get(service, service)
    set_val(3, meiqi_service)

    # 若渠道不在「服务渠道」下拉清单，追加到 B 列（B2:B132 范围内），保证下拉有效
    if meiqi_service:
        ws_svc = wb['服务渠道']
        found = False
        for r in range(2, 133):
            if (ws_svc.cell(row=r, column=2).value or '') == meiqi_service:
                found = True
                break
        if not found:
            for r in range(2, 133):
                if ws_svc.cell(row=r, column=2).value is None:
                    ws_svc.cell(row=r, column=2).value = meiqi_service
                    break

    # B4: 地址库编码
    set_val(4, wh_code)

    # B5-B10: 收件人信息 = 亚马逊仓库代码查表；查不到回退源字段
    if wh_code and wh_code in address_lib:
        lib = address_lib[wh_code]
        set_val(5, lib['联系人'])        # 收件人姓名
        set_val(6, lib['地址一'])        # 收件人地址一
        set_val(7, lib['城市'])          # 收件人城市
        set_val(8, lib['省洲'])          # 收件人省/洲
        set_val(9, lib['邮编'])          # 收件人邮编
        set_val(10, lib['国家'])         # 收件人国家代码
    else:
        set_val(5, tr.get('收件人姓名', ''))
        set_val(6, tr.get('收件人地址一', ''))
        set_val(7, tr.get('收件人城市', ''))
        set_val(8, tr.get('收件人省份/州', ''))
        set_val(9, tr.get('收件人邮编', ''))
        set_val(10, tr.get('收件人国家代码(二字代码)', '')
                 or tr.get('收件人国家代码', '') or 'US')

    # B11-B13: 收件人电话/邮箱/公司名称 → 留空（与美琦完成示例一致）
    set_val(11, '')
    set_val(12, '')
    set_val(13, '')

    # B14/B15: 带电/带磁（是/否）
    set_val(14, '是' if tr.get('带电', '否') == '是' else '否')
    set_val(15, '是' if tr.get('带磁', '否') == '是' else '否')

    # B16: 报关方式（美线规则归一化：退税报关→一般贸易，代理报关→代理报关）
    customs = tr.get('报关方式', '') or ''
    if '退税' in customs:
        customs = '一般贸易'
    elif '代理' in customs:
        customs = '代理报关'
    set_val(16, customs)

    # B17: 箱数 = 源箱数；为空则按数据行箱数合计
    box_count_total = sum(_parse_box_count(dr['A']) for dr in tr.data_rows)
    src_boxes = tr.get('箱数', '')
    set_val(17, src_boxes if src_boxes is not None and str(src_boxes).strip() != ''
            else box_count_total)

    # ══════════════════════════════════════════════
    #  数据行 Row 19+
    # ══════════════════════════════════════════════

    # 收集模板表头自带的图片（Row<=17，如左上角品牌图），嵌入时一并保留。
    # 注：img._data() 会关闭 BytesIO ref，且 openpyxl 保存时会再次序列化同一图片，
    # 故收集后清空 ws._images，统一由 _embed_images_as_cell_images 后处理写入。
    pending_images = {}
    for img in list(ws._images):
        try:
            a = img.anchor
            if hasattr(a, '_from') and a._from.row + 1 <= 17:
                col_letter = get_column_letter(a._from.col + 1)
                img_row_1 = a._from.row + 1
                img_data = img._data()
                if isinstance(img_data, bytes):
                    pending_images[f'{col_letter}{img_row_1}'] = img_data
                elif isinstance(img_data, BytesIO):
                    pending_images[f'{col_letter}{img_row_1}'] = img_data.getvalue()
        except Exception:
            pass
    ws._images = []

    # 清空模版自带示例数据 (Row 19-202, A-R)
    for r in range(19, 203):
        for c in range(1, 19):
            ws.cell(row=r, column=c).value = None

    data_start = 19
    box_cursor = 1  # 运行式箱号区间起点
    for i, dr in enumerate(tr.data_rows):
        r = data_start + i
        ws.row_dimensions[r].height = 80  # 产品图片列 O 需要足够高度

        # 每行箱数 → 货箱编号区间（如 1-3、4-6）
        box_cnt = _parse_box_count(dr['A'])
        box_start = box_cursor
        box_end = box_cursor + box_cnt - 1
        box_cursor += box_cnt
        ws.cell(row=r, column=1).value = f'{box_start}-{box_end}'   # A: 货箱编号

        ws.cell(row=r, column=2).value = dr['F'] if dr['F'] is not None else ''   # B: 品名英文
        ws.cell(row=r, column=3).value = dr['G'] if dr['G'] is not None else ''   # C: 品名中文
        ws.cell(row=r, column=4).value = dr['I'] if dr['I'] is not None else ''   # D: 申报数量（单箱）
        ws.cell(row=r, column=5).value = dr['H'] if dr['H'] is not None else ''   # E: 申报单价（美金）
        ws.cell(row=r, column=6).value = dr['B'] if dr['B'] is not None else ''   # F: 货箱重量
        ws.cell(row=r, column=7).value = dr['C'] if dr['C'] is not None else ''   # G: 货箱长度
        ws.cell(row=r, column=8).value = dr['D'] if dr['D'] is not None else ''   # H: 货箱宽度
        ws.cell(row=r, column=9).value = dr['E'] if dr['E'] is not None else ''   # I: 货箱高度
        ws.cell(row=r, column=10).value = dr['K'] if dr['K'] is not None else ''  # J: 海关编码（源原值，不改变格式/不加小数点）
        ws.cell(row=r, column=11).value = dr['M'] if dr['M'] is not None else ''  # K: 品牌
        ws.cell(row=r, column=12).value = dr['J'] if dr['J'] is not None else ''  # L: 材质
        ws.cell(row=r, column=13).value = dr['N'] if dr['N'] is not None else ''  # M: 型号
        ws.cell(row=r, column=14).value = dr['L'] if dr['L'] is not None else ''  # N: 用途

        # O: 产品图片（从源发票提取嵌入，参考天图方案）
        src_row = dr.get('_row')
        img_bytes = tr.images.get(src_row) if src_row else None
        if img_bytes:
            pending_images[f'O{r}'] = img_bytes
            ws.cell(row=r, column=15).value = '[图片]'

        # P: PO Number（源 V 列，优先行级，其次头部）
        po = dr.get('V') or tr.get('PO Number', '') or ''
        ws.cell(row=r, column=16).value = po

        # Q: 物品箱号 = 单行总箱数
        ws.cell(row=r, column=17).value = box_cnt

        # R: 物品FBA ID = 货箱编号中 U00000 前 12 位
        ws.cell(row=r, column=18).value = _extract_fba_id(dr['A'])

    num_data = len(tr.data_rows)
    wb.save(output_path)

    # 后处理：嵌入产品图片（+保留模板表头图片），参考天图 twoCellAnchor 方案
    if pending_images:
        _embed_images_as_cell_images(output_path, pending_images, None)

    print(f'✅ 美琦美线发票已生成: {os.path.basename(output_path)}')
    print(f'   数据: {num_data} 行, 箱数: {box_count_total}, 图片: {len(pending_images)}')
    return True


def _guess_product_category(tr):
    """
    从 TR 发票的多个产品中文品名推断物品属性。
    例如所有产品含"灯" → "灯类"，否则合并或不填。
    """
    import re
    all_names = [dr.get('G', '') or '' for dr in tr.data_rows if dr.get('G')]
    if not all_names:
        return ''

    # 常见分类关键词
    keywords = {
        '灯类': ['灯', '照明'],
        '服装': ['衣', '服', '裤', '裙', '袜', '帽', '鞋'],
        '电子': ['电子', '电', '机', '器', '充电', '电池', '线'],
        '家居': ['家具', '桌', '椅', '凳', '架', '柜', '收纳'],
        '玩具': ['玩具', '玩'],
        '箱包': ['包', '箱', '袋', '背包'],
        '美容': ['美容', '化妆', '护肤', '梳'],
    }

    scores = {}
    for cat, kws in keywords.items():
        score = 0
        for name in all_names:
            for kw in kws:
                if kw in name:
                    score += 1
                    break
        if score > 0:
            scores[cat] = score

    if scores:
        best = max(scores, key=scores.get)
        return best

    return '普货'


# ═══════════════════════════════════════════════════════════════
#  4. 批处理 — 目录中所有 TR 发票
# ═══════════════════════════════════════════════════════════════

def batch_convert(input_dir, output_dir, target='天图'):
    """转换目录中所有 TR 发票"""
    if not os.path.isdir(input_dir):
        print(f'❌ 输入目录不存在: {input_dir}')
        return

    os.makedirs(output_dir, exist_ok=True)

    # 找到所有TR发票文件
    files = [f for f in os.listdir(input_dir)
             if f.endswith('.xlsx') and '订单' not in f
             and '模板' not in f and '天图' not in f
             and '航乐' not in f and '美琦' not in f]
    files.sort()

    if not files:
        print(f'⚠️  未找到 .xlsx 文件 (排除模板和已转换的)')
        return

    success = 0
    for fn in files:
        in_path = os.path.join(input_dir, fn)
        ext = ''
        if target == '天图':
            ext = f'-{target}.xlsx'
        elif target in ('航乐-uk', '航乐-eu'):
            region = 'uk' if target == '航乐-uk' else 'eu'
            ext = f'-{target}.xlsx'
        else:
            ext = f'-{target}.xlsx'

        base = os.path.splitext(fn)[0]
        out_name = f'{base}{ext}'
        out_path = os.path.join(output_dir, out_name)

        try:
            tr = TRInvoice(in_path)
            if target == '天图':
                ok = convert_to_tiantu(tr, out_path)
            elif target == '航乐-uk':
                ok = convert_to_hangle(tr, out_path, region='uk')
            elif target == '航乐-eu':
                ok = convert_to_hangle(tr, out_path, region='eu')
            elif target == '美琦':
                ok = convert_to_meiqi(tr, out_path)
            else:
                print(f'❌ 未知目标格式: {target}')
                return
            if ok:
                success += 1
        except Exception as e:
            print(f'❌ 转换失败 {fn}: {e}')

    print(f'\n📊 完成: {success}/{len(files)} 文件已转换')


# ═══════════════════════════════════════════════════════════════
#  图片嵌入后处理 — 将图片嵌入为单元格 IMAGE() 公式
#  (Excel 365 "Place in Cell" / "放置在单元格中")
# ═══════════════════════════════════════════════════════════════

def _get_image_emu_size(img_bytes, default_cx=914400, default_cy=914400):
    """
    从图片字节读取实际像素尺寸，转换为 EMU（English Metric Unit）。

    EMU 转换：1 英寸 = 914400 EMU，默认 96 DPI。
    所以 1 像素 ≈ 914400 / 96 = 9525 EMU。

    返回 (cx, cy) EMU 值，保证非零（否则 WPS 不渲染）。
    """
    try:
        from PIL import Image
        from io import BytesIO
        img = Image.open(BytesIO(img_bytes))
        w, h = img.size
        dpi = img.info.get('dpi', (96, 96))
        cx = int(w * 914400 / dpi[0])
        cy = int(h * 914400 / dpi[1])
        if cx > 0 and cy > 0:
            return cx, cy
    except Exception:
        pass
    # 回退到基于文件大小估算（最差情况也有个合理值）
    return default_cx, default_cy


def _embed_images_as_cell_images(xlsx_path, cell_image_map, image_url_base=None):
    """
    后处理 xlsx 文件，将图片嵌入为标准 Excel 绘图格式（xl/drawings/drawing1.xml）。

    生成：
      1. xl/media/ - 图片文件
      2. xl/drawings/drawing1.xml - 标准 Excel 绘图定义（oneCellAnchor）
      3. xl/drawings/_rels/drawing1.xml.rels - 图片关系映射
      4. worksheet 中添加 <drawing r:id="..."> 元素
      5. worksheet rels 中添加 drawing 关系
    """
    if not cell_image_map:
        return

    NS_S = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    NS_CT = 'http://schemas.openxmlformats.org/package/2006/content-types'
    NS_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
    NS_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    NS_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    # 注册标准命名空间前缀
    ET.register_namespace('xdr', NS_XDR)
    ET.register_namespace('a', NS_A)
    ET.register_namespace('r', NS_R)
    ET.register_namespace('', NS_PKG)

    tmp_dir = tempfile.mkdtemp()
    try:
        # ── 1. 解压 xlsx ──
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            z.extractall(tmp_dir)

        sorted_refs = list(cell_image_map.items())

        # ── 2. 写入图片到 xl/media/ ──
        media_dir = os.path.join(tmp_dir, 'xl', 'media')
        os.makedirs(media_dir, exist_ok=True)
        # img_info: [(cell_ref, img_filename, col_0, row_0, cx, cy, idx), ...]
        img_info = []
        for i, (cell_ref, img_bytes) in enumerate(sorted_refs):
            import re
            m = re.match(r'([A-Z]+)(\d+)', cell_ref)
            if not m:
                print(f'  ⚠️  无法解析单元格引用 {cell_ref}')
                continue
            col_letters, row_num = m.group(1), m.group(2)
            # 列字母 → 0-based 数字
            col_0 = 0
            for ch in col_letters:
                col_0 = col_0 * 26 + (ord(ch) - ord('A') + 1)
            col_0 -= 1
            row_0 = int(row_num) - 1

            idx = i + 1
            img_filename = f'image_{idx}.png'
            with open(os.path.join(media_dir, img_filename), 'wb') as f:
                f.write(img_bytes)

            cx, cy = _get_image_emu_size(img_bytes)
            # 缩放到适合单元格的缩略图尺寸（最大约 90 像素）
            THUMB_EMU = 857250  # ~90px at 96 DPI
            if cx > THUMB_EMU or cy > THUMB_EMU:
                scale = min(THUMB_EMU / cx, THUMB_EMU / cy)
                cx = int(cx * scale)
                cy = int(cy * scale)
            img_info.append((cell_ref, img_filename, col_0, row_0, cx, cy, idx))

        # ── 3. 创建 xl/drawings/drawing1.xml ──
        drawings_dir = os.path.join(tmp_dir, 'xl', 'drawings')
        os.makedirs(drawings_dir, exist_ok=True)
        # 同时创建 drawing rels 目录
        drawings_rels_dir = os.path.join(drawings_dir, '_rels')
        os.makedirs(drawings_rels_dir, exist_ok=True)

        wsDr = ET.Element(f'{{{NS_XDR}}}wsDr')

        for cell_ref, img_filename, col_0, row_0, cx, cy, idx in img_info:
            # twoCellAnchor：图片跟随单元格大小，自动适配
            anchor = ET.SubElement(wsDr, f'{{{NS_XDR}}}twoCellAnchor')

            # from — 图片左上角（单元格左上角，无边距）
            frm = ET.SubElement(anchor, f'{{{NS_XDR}}}from')
            ET.SubElement(frm, f'{{{NS_XDR}}}col').text = str(col_0)
            ET.SubElement(frm, f'{{{NS_XDR}}}colOff').text = '0'
            ET.SubElement(frm, f'{{{NS_XDR}}}row').text = str(row_0)
            ET.SubElement(frm, f'{{{NS_XDR}}}rowOff').text = '0'

            # to — 图片右下角（相邻单元格边界，即单元格大小）
            to = ET.SubElement(anchor, f'{{{NS_XDR}}}to')
            ET.SubElement(to, f'{{{NS_XDR}}}col').text = str(col_0 + 1)
            ET.SubElement(to, f'{{{NS_XDR}}}colOff').text = '0'
            ET.SubElement(to, f'{{{NS_XDR}}}row').text = str(row_0 + 1)
            ET.SubElement(to, f'{{{NS_XDR}}}rowOff').text = '0'

            # pic
            pic = ET.SubElement(anchor, f'{{{NS_XDR}}}pic')

            # nvPicPr
            nvPicPr = ET.SubElement(pic, f'{{{NS_XDR}}}nvPicPr')
            ET.SubElement(nvPicPr, f'{{{NS_XDR}}}cNvPr', id=str(idx), name=f'Picture {idx}')
            nvPicPr2 = ET.SubElement(nvPicPr, f'{{{NS_XDR}}}cNvPicPr')
            ET.SubElement(nvPicPr2, f'{{{NS_A}}}picLocks', noChangeAspect='1')

            # blipFill
            blipFill = ET.SubElement(pic, f'{{{NS_XDR}}}blipFill')
            ET.SubElement(blipFill, f'{{{NS_A}}}blip', attrib={f'{{{NS_R}}}embed': f'rId{idx}'})
            ET.SubElement(blipFill, f'{{{NS_XDR}}}srcRect')
            stretch = ET.SubElement(blipFill, f'{{{NS_XDR}}}stretch')
            ET.SubElement(stretch, f'{{{NS_XDR}}}fillRect')

            # spPr
            spPr = ET.SubElement(pic, f'{{{NS_XDR}}}spPr')
            xfrm = ET.SubElement(spPr, f'{{{NS_A}}}xfrm')
            ET.SubElement(xfrm, f'{{{NS_A}}}off', x='0', y='0')
            ET.SubElement(xfrm, f'{{{NS_A}}}ext', cx=str(cx), cy=str(cy))
            prstGeom = ET.SubElement(spPr, f'{{{NS_A}}}prstGeom', prst='rect')
            ET.SubElement(prstGeom, f'{{{NS_A}}}avLst')

            # clientData
            ET.SubElement(anchor, f'{{{NS_XDR}}}clientData')

        drawing_path = os.path.join(drawings_dir, 'drawing1.xml')
        ET.ElementTree(wsDr).write(drawing_path, xml_declaration=True, encoding='UTF-8')

        # ── 4. 创建 xl/drawings/_rels/drawing1.xml.rels ──
        dw_rels_root = ET.Element(f'{{{NS_PKG}}}Relationships')
        for cell_ref, img_filename, col_0, row_0, cx, cy, idx in img_info:
            rel_elem = ET.SubElement(dw_rels_root, f'{{{NS_PKG}}}Relationship')
            rel_elem.set('Id', f'rId{idx}')
            rel_elem.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/image')
            rel_elem.set('Target', f'../media/{img_filename}')

        dw_rels_path = os.path.join(drawings_rels_dir, 'drawing1.xml.rels')
        ET.ElementTree(dw_rels_root).write(dw_rels_path, xml_declaration=True, encoding='UTF-8')

        # ── 5. 修改 sheet1.xml：添加 <drawing r:id="..."> 引用 ──
        sheet_path = os.path.join(tmp_dir, 'xl', 'worksheets', 'sheet1.xml')
        if not os.path.exists(sheet_path):
            print(f'  ⚠️  sheet1.xml 不存在，跳过图片嵌入')
            return

        tree = ET.parse(sheet_path)
        root = tree.getroot()

        # 找到或创建 drawing 引用元素（放在 sheetData 之后）
        existing_drawing = root.find(f'{{{NS_S}}}drawing')
        if existing_drawing is not None:
            # 已有 drawing 引用，使用已有 rId
            drawing_rid = existing_drawing.get(f'{{{NS_R}}}id', '')
        else:
            # 分配新的 rId
            drawing_rid = 'rId1'
            # 放在 sheetData 和 sheetProtection 之后
            insert_after = root.find(f'{{{NS_S}}}sheetData')
            if insert_after is not None:
                # 找到 insert_after 的后续兄弟元素位置
                parent = root
                idx = list(parent).index(insert_after)
                # 找到最后一个非尾部的兄弟
                for child in list(parent)[idx + 1:]:
                    if child.tag == f'{{{NS_S}}}sheetData':
                        continue
                    insert_after = child
                # 在 insert_after 后插入
                drawing_elem = ET.SubElement(root, f'{{{NS_S}}}drawing')
            else:
                drawing_elem = ET.SubElement(root, f'{{{NS_S}}}drawing')
            drawing_elem.set(f'{{{NS_R}}}id', drawing_rid)

        tree.write(sheet_path, xml_declaration=True, encoding='UTF-8')

        # ── 6. 添加 worksheet → drawing 关系 ──
        ws_rels_dir = os.path.join(tmp_dir, 'xl', 'worksheets', '_rels')
        os.makedirs(ws_rels_dir, exist_ok=True)
        ws_rels_path = os.path.join(ws_rels_dir, 'sheet1.xml.rels')

        if os.path.exists(ws_rels_path):
            ws_rels_tree = ET.parse(ws_rels_path)
            ws_rels_root = ws_rels_tree.getroot()
        else:
            ws_rels_root = ET.Element(f'{{{NS_PKG}}}Relationships')
            ws_rels_tree = ET.ElementTree(ws_rels_root)

        # 检查是否已有 drawing 关系
        has_drawing = any(
            rel.get('Type', '') == 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing'
            for rel in ws_rels_root
        )
        if not has_drawing:
            next_rid = 1
            for rel_elem in ws_rels_root:
                rid = rel_elem.get('Id', '')
                if rid.startswith('rId'):
                    try:
                        num = int(rid[3:])
                        next_rid = max(next_rid, num + 1)
                    except:
                        pass
            rel_elem = ET.SubElement(ws_rels_root, f'{{{NS_PKG}}}Relationship')
            rel_elem.set('Id', f'rId{next_rid}')
            rel_elem.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing')
            rel_elem.set('Target', '../drawings/drawing1.xml')
            ws_rels_tree.write(ws_rels_path, xml_declaration=True, encoding='UTF-8')

        # ── 7. 补充 Content_Types ──
        ct_path = os.path.join(tmp_dir, '[Content_Types].xml')
        if os.path.exists(ct_path):
            ct_tree = ET.parse(ct_path)
            ct_root = ct_tree.getroot()
            # PNG Default
            has_png = any(
                default.get('Extension') == 'png'
                for default in ct_root.findall(f'{{{NS_CT}}}Default')
            )
            if not has_png:
                png_default = ET.SubElement(ct_root, f'{{{NS_CT}}}Default')
                png_default.set('Extension', 'png')
                png_default.set('ContentType', 'image/png')
            # drawing.xml Override
            has_dw_override = any(
                override.get('PartName') == '/xl/drawings/drawing1.xml'
                for override in ct_root.findall(f'{{{NS_CT}}}Override')
            )
            if not has_dw_override:
                dw_override = ET.SubElement(ct_root, f'{{{NS_CT}}}Override')
                dw_override.set('PartName', '/xl/drawings/drawing1.xml')
                dw_override.set('ContentType', 'application/vnd.openxmlformats-officedocument.drawing+xml')
            ct_tree.write(ct_path, xml_declaration=True, encoding='UTF-8')

        # ── 8. 重新打包 ──
        final_path = xlsx_path + '.tmp'
        with zipfile.ZipFile(final_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for dirpath, dirnames, filenames in os.walk(tmp_dir):
                for fn in filenames:
                    full_path = os.path.join(dirpath, fn)
                    arcname = os.path.relpath(full_path, tmp_dir)
                    zout.write(full_path, arcname)

        shutil.move(final_path, xlsx_path)

        print(f'  📷 已嵌入 {len(img_info)} 张图片（标准 Excel 绘图 xl/drawings/drawing1.xml）')

    except Exception as e:
        print(f'  ⚠️  图片嵌入后处理出错: {e}')
        import traceback
        traceback.print_exc()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ═══════════════════════════════════════════════════════════════
#  5. 主入口
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='TR发票 → 供应商发票 转换工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python3 convert_invoice.py TR发票.xlsx --to 天图
  python3 convert_invoice.py TR发票.xlsx --to 航乐-uk 输出文件.xlsx
  python3 convert_invoice.py TR发票.xlsx --to 航乐-eu
  python3 convert_invoice.py --batch ./发票 --to 天图 --out ./输出
        """
    )
    parser.add_argument('input', nargs='?',
                        help='TR发票 .xlsx 文件路径')
    parser.add_argument('--to', '-t', default='天图',
                        choices=['天图', '航乐-uk', '航乐-eu', '美琦'],
                        help='目标供应商格式 (默认: 天图)')
    parser.add_argument('output', nargs='?',
                        help='输出文件路径 (可选，默认自动生成)')
    parser.add_argument('--batch', '-b', action='store_true',
                        help='批量模式: 转换目录中所有TR发票')
    parser.add_argument('--in-dir', default=SCRIPT_DIR,
                        help='批量模式的输入目录')
    parser.add_argument('--out-dir', default=os.path.join(SCRIPT_DIR, 'output'),
                        help='批量模式的输出目录 (默认: ./output)')
    parser.add_argument('--order-list', default=None,
                        help='订单列表 .xlsx（含「地址库编码」「运单号」列，可选；按地址库编码回填客户订单号为运单号）')

    # 用 parse_intermixed_args 支持 `输入 --to 目标 输出` 的位置参数穿插
    args = parser.parse_intermixed_args()

    # ── 批量模式 ──
    if args.batch:
        batch_convert(args.in_dir, args.out_dir, args.to)
        return

    # ── 单文件模式 ──
    if not args.input:
        parser.print_help()
        print('\n❌ 请指定TR发票文件路径')
        sys.exit(1)

    if not os.path.exists(args.input):
        print(f'❌ 文件不存在: {args.input}')
        sys.exit(1)

    # 读取TR发票
    print(f'📂 读取: {args.input}')
    tr = TRInvoice(args.input)
    print(tr)

    # 自动生成输出文件名
    if not args.output:
        base_name = os.path.splitext(os.path.basename(args.input))[0]
        if args.to == '天图':
            args.output = os.path.join(os.path.dirname(args.input),
                                       f'{base_name}-天图.xlsx')
        elif args.to == '航乐-uk':
            args.output = os.path.join(os.path.dirname(args.input),
                                       f'{base_name}-航乐-UK.xlsx')
        elif args.to == '航乐-eu':
            args.output = os.path.join(os.path.dirname(args.input),
                                       f'{base_name}-航乐-EU.xlsx')
        elif args.to == '美琦':
            args.output = os.path.join(os.path.dirname(args.input),
                                       f'{base_name}-美琦.xlsx')

    # 执行转换
    if args.to == '天图':
        convert_to_tiantu(tr, args.output, order_list_path=args.order_list)
    elif args.to == '航乐-uk':
        convert_to_hangle(tr, args.output, region='uk', order_list_path=args.order_list)
    elif args.to == '航乐-eu':
        convert_to_hangle(tr, args.output, region='eu', order_list_path=args.order_list)
    elif args.to == '美琦':
        convert_to_meiqi(tr, args.output, order_list_path=args.order_list)


if __name__ == '__main__':
    main()
