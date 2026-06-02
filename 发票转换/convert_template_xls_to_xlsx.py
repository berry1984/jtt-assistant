#!/usr/bin/env python3
"""
将航乐 .xls 模板转换为 .xlsx 格式（保留字体、颜色、填充、边框、对齐）。

openpyxl 不支持读取旧版 .xls，所以用 xlrd 读取格式信息后，
用 openpyxl 重新创建并应用格式。

用法：
  python3 convert_template_xls_to_xlsx.py
"""
import os
import sys

THIS_DIR = os.path.dirname(os.path.abspath(__file__))

# ── 要转换的模板 ──
TEMPLATES = [
    ('航乐-客户名称 客户单号 英国发票模板9.9更新.xls', '航乐-客户名称 客户单号 英国发票模板9.9更新.xlsx'),
    ('航乐-客户单号- 欧州发票模板2.26更新.xls', '航乐-客户单号- 欧州发票模板2.26更新.xlsx'),
]

# xlrd 边框线型 → openpyxl Side style 映射
LINE_STYLE_MAP = {
    0: None,       # no line
    1: 'thin',
    2: 'medium',
    3: 'dashed',
    4: 'dotted',
    5: 'thick',
    6: 'double',
    7: 'hair',
    8: 'mediumDashDot',
    9: 'dashDot',
    10: 'mediumDashDotDot',
    11: 'dashDotDot',
    12: 'slantDashDot',
}

def xlrd_font_to_openpyxl(xlrd_font, colour_map):
    """将 xlrd Font 对象转为 openpyxl Font"""
    from openpyxl.styles import Font, Color
    rgb = None
    if xlrd_font.colour_index and xlrd_font.colour_index < len(colour_map):
        rgb = colour_map[xlrd_font.colour_index]

    color = None
    if rgb and isinstance(rgb, tuple) and len(rgb) == 3:
        color = Color(rgb=f'FF{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}')
    elif xlrd_font.colour_index == 0:
        color = Color(rgb='FF000000')

    return Font(
        name=xlrd_font.name,
        size=xlrd_font.height / 20 if xlrd_font.height else 11,
        bold=bool(xlrd_font.bold),
        italic=bool(xlrd_font.italic),
        underline='single' if xlrd_font.underline_type else None,
        color=color,
    )


def xlrd_fill_to_openpyxl(xf, colour_map):
    """将 xlrd XF 背景转为 openpyxl PatternFill"""
    from openpyxl.styles import PatternFill, Color
    bg = xf.background

    # pattern_colour_index: 9=white, 13=yellow, 64=None
    pat_idx = bg.pattern_colour_index
    bg_idx = bg.background_colour_index

    if pat_idx is None or pat_idx == 64:
        return None

    # 获取颜色
    fg_rgb = None
    if pat_idx and pat_idx < len(colour_map):
        c = colour_map[pat_idx]
        if c and isinstance(c, tuple):
            fg_rgb = f'{c[0]:02X}{c[1]:02X}{c[2]:02X}'

    if fg_rgb:
        return PatternFill(patternType='solid', fgColor=Color(rgb=f'FF{fg_rgb}'))

    return None


def xlrd_border_to_openpyxl(xlrd_border):
    """将 xlrd Border 转为 openpyxl Border"""
    from openpyxl.styles import Border, Side

    def side(style_code):
        style = LINE_STYLE_MAP.get(style_code)
        if style:
            return Side(style=style)
        return None

    return Border(
        left=side(xlrd_border.left_line_style),
        right=side(xlrd_border.right_line_style),
        top=side(xlrd_border.top_line_style),
        bottom=side(xlrd_border.bottom_line_style),
    )


def xlrd_align_to_openpyxl(xlrd_align):
    """将 xlrd Alignment 转为 openpyxl Alignment"""
    from openpyxl.styles import Alignment

    HOR_MAP = {0: 'general', 1: 'left', 2: 'center', 3: 'right', 4: 'fill',
               5: 'justify', 6: 'centerContinuous', 7: 'distributed'}
    VER_MAP = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}

    return Alignment(
        horizontal=HOR_MAP.get(xlrd_align.hor_align, 'general'),
        vertical=VER_MAP.get(xlrd_align.vert_align, 'bottom'),
        wrap_text=bool(xlrd_align.text_wrapped),
    )


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """用 xlrd 读 .xls（含格式），用 openpyxl 写 .xlsx（保留格式）"""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    xls_wb = xlrd.open_workbook(xls_path, formatting_info=True)
    xlsx_wb = Workbook()

    # 删掉默认 Sheet
    xlsx_wb.remove(xlsx_wb.active)

    colour_map = xls_wb.colour_map

    for si in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheet_by_index(si)
        ws = xlsx_wb.create_sheet(title=xls_ws.name)

        # ── 列宽 ──
        for c in range(xls_ws.ncols):
            ci = xls_ws.colinfo_map.get(c)
            if ci:
                ws.column_dimensions[get_column_letter(c+1)].width = ci.width / 256 * 1.2

        # ── 行高 + 行数据 ──
        for r in range(xls_ws.nrows):
            ri = xls_ws.rowinfo_map.get(r)
            if ri:
                ws.row_dimensions[r+1].height = ri.height / 20

            for c in range(xls_ws.ncols):
                cell_type = xls_ws.cell_type(r, c)
                cell_value = xls_ws.cell_value(r, c)

                # 跳过完全空的单元格（无值也无格式）
                if cell_type == 0:
                    # 但检查是否有格式需要保留
                    xf_idx = xls_ws.cell_xf_index(r, c)
                    xf = xls_wb.xf_list[xf_idx]
                    font = xls_wb.font_list[xf.font_index]
                    bg = xf.background
                    # 只有非默认格式才保留
                    if font.name == 'Arial' and font.height == 220 and not font.bold:
                        continue
                    # 有格式的空单元格也写入
                    cell = ws.cell(row=r+1, column=c+1)
                    cell.value = None  # 保持空
                elif cell_type == 1:  # Text
                    ws.cell(row=r+1, column=c+1, value=str(cell_value))
                elif cell_type == 2:  # Number
                    ws.cell(row=r+1, column=c+1, value=cell_value)
                elif cell_type == 3:  # Date
                    ws.cell(row=r+1, column=c+1, value=cell_value)
                elif cell_type == 4:  # Boolean
                    ws.cell(row=r+1, column=c+1, value=bool(cell_value))
                elif cell_type == 6:  # Blank (有格式但无值)
                    cell = ws.cell(row=r+1, column=c+1)
                    cell.value = None
                else:
                    continue

                # ── 应用格式 ──
                cell = ws.cell(row=r+1, column=c+1)
                xf_idx = xls_ws.cell_xf_index(r, c)
                xf = xls_wb.xf_list[xf_idx]

                font = xls_wb.font_list[xf.font_index]
                cell.font = xlrd_font_to_openpyxl(font, colour_map)

                fill = xlrd_fill_to_openpyxl(xf, colour_map)
                if fill:
                    cell.fill = fill

                border = xlrd_border_to_openpyxl(xf.border)
                if any([border.left, border.right, border.top, border.bottom]):
                    cell.border = border

                align = xlrd_align_to_openpyxl(xf.alignment)
                if align.horizontal != 'general' or align.vertical != 'bottom' or align.wrap_text:
                    cell.alignment = align

        # ── 合并单元格 ──
        for merge in xls_ws.merged_cells:
            rlo, rhi, clo, chi = merge
            try:
                ws.merge_cells(
                    start_row=rlo+1, start_column=clo+1,
                    end_row=rhi, end_column=chi
                )
            except Exception:
                pass  # 忽略重复合并

    xlsx_wb.save(xlsx_path)
    print(f'✅ 已转换（含格式）: {os.path.basename(xls_path)} → {os.path.basename(xlsx_path)}')


if __name__ == '__main__':
    for xls_name, xlsx_name in TEMPLATES:
        xls_path = os.path.join(THIS_DIR, xls_name)
        xlsx_path = os.path.join(THIS_DIR, xlsx_name)
        if not os.path.exists(xls_path):
            print(f'⚠️  源文件不存在: {xls_path}')
            continue
        convert_xls_to_xlsx(xls_path, xlsx_path)
