"""检查输出格式是否与模板一致"""
import sys
sys.path.insert(0, '.')
from convert_invoice import TRInvoice, convert_to_hangle
from openpyxl import load_workbook

tr = TRInvoice('5月第5周（赛诺吉发票）-POZ1-FP20260529005  10箱 欧洲铁路包税-普铁快递派 是6.xlsx')
ok = convert_to_hangle(tr, '/tmp/test_final_check.xlsx', region='eu')

wb = load_workbook('/tmp/test_final_check.xlsx')
ws = wb.active

print("=" * 60)
print("格式验证")
print("=" * 60)

checks = [
    ('A1', '标题', 'Arial Black', 16, 'FFFFFFFF'),
    ('A3', '标签ADD', '宋体', 11, 'FFFFFF00'),
    ('B3', '值B3', '宋体', 11, 'FFFFFFFF'),
    ('A4', '标签ZIP', '宋体', 12, 'FFFFFF00'),
    ('D4', '值D4', '宋体', 12, 'FFFFFFFF'),
    ('A5', '标签CITY', '宋体', 12, 'FFFFFF00'),
    ('B5', '值B5', '宋体', 12, 'FFFFFFFF'),
    ('A7', '标签公司', '微软雅黑', 12, 'FFFFFFFF'),
    ('I7', '渠道值', '宋体', 18, 'FFFFFFFF'),
    ('A11', '表头箱号', '宋体', 15, 'FFFFFF00'),
    ('Y11', 'PO表头', '宋体', 15, 'FFFFFF00'),
    ('C24', '合计行', '宋体', 11, 'FFFFFFFF'),
]

all_ok = True
for cell_ref, desc, exp_font, exp_size, exp_fill in checks:
    cell = ws[cell_ref]
    f = cell.font
    fill = cell.fill

    issues = []
    if f.name != exp_font:
        issues.append('font=%s' % f.name)
    if abs(f.size - exp_size) > 0.5:
        issues.append('size=%.1f' % f.size)
    actual_fill = fill.fgColor.rgb if fill.fill_type and fill.fill_type != 'none' else 'none'
    if actual_fill != exp_fill:
        issues.append('fill=%s' % actual_fill)

    if issues:
        all_ok = False
        print('  ❌ %s (%s): %s' % (cell_ref, desc, ' | '.join(issues)))
    else:
        print('  ✅ %s (%s) [%s %spt]' % (cell_ref, desc, exp_font, exp_size))

print()
if all_ok:
    print("✅ 全部格式与模板一致！")
else:
    print("❌ 有格式差异")

# 验证数据正确性
print()
print("=" * 60)
print("数据验证")
print("=" * 60)
data_checks = [
    ('B3', 'POZ1'),
    ('G3', 'DE'),
    ('B4', '02977'),
    ('D4', 'Amazon Fulfilment Center'),
    ('B5', 'Hoyerswerda'),
    ('D5', 'POZ1'),
    ('I5', '报关退税'),
    ('I7', 'FBA德国铁路DHL'),
    ('L7', '包税'),
    ('M7', '灯类'),
]
for cell_ref, expected in data_checks:
    actual = str(ws[cell_ref].value or '')
    if actual == expected:
        print('  ✅ %s = %s' % (cell_ref, expected))
    else:
        print('  ❌ %s = %s (期望 %s)' % (cell_ref, actual, expected))

# 图片验证
import zipfile
with zipfile.ZipFile('/tmp/test_final_check.xlsx', 'r') as z:
    imgs = [f for f in z.namelist() if 'media' in f]
    has_img = b'IMAGE' in z.read('xl/worksheets/sheet1.xml')
    print('\n  ✅ xl/media: %d张' % len(imgs))
    print('  ✅ IMAGE公式: 已嵌入' if has_img else '  ❌ IMAGE公式: 未找到')
