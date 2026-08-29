#!/usr/bin/env python3
"""
提单及电放保函生成模块 — 供 JTT电商AI助手 Flask 应用调用

用法（在 app.py 中）:
    from gen_bl_docs import generate_bl_docs
    zip_path = generate_bl_docs(uploaded_excel_path, output_dir)

输入要求:
    Excel 文件含 sheet "5月提单信息"，列名与 TR 退税资料明细 一致
    (JTT no. / 渠道 / 引用模板 / B/L No. / Ocean Vessel / ...)

输出:
    返回 ZIP 文件路径，内含 提单PDF + 电放保函xlsx
"""

import os, sys, io, re, zipfile, tempfile, shutil
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import fitz  # PyMuPDF

# ── 模板路径（相对于本模块） ──
MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(MODULE_DIR, 'templates_bl')
BL_SEA = os.path.join(TEMPLATES_DIR, '提单By sea.pdf')
BL_TRAIN = os.path.join(TEMPLATES_DIR, '提单By train.pdf')
BL_TRUCK = os.path.join(TEMPLATES_DIR, '提单By truck.pdf')
TELEX = os.path.join(TEMPLATES_DIR, '电放保函.xlsx')

# ── 标准发货人/收货人（上传数据缺 Shipper/Consignee 列或为空时的兜底）──
FALLBACK_SHIPPER = ("Guangzhou Tuorui Technology Co.,Ltd\n"
                    "Room 411,No.101 Dexing Road Wanggang Jiahe Street\n"
                    "Baiyun District, Guangzhou")
FALLBACK_CONSIGNEE = ("Hong Kong Lixiang Trading Company Limited\n"
                      "FLAT/RM 3B 3/F BANK TOWER NOS.351&353 KING'S\n"
                      "ROAD NORTH POINT HK")

# ── 跨平台字体查找（macOS / Linux 通用） ──
_FONT_CACHE = None  # (fontname, fontfile)

def _get_font():
    """返回 (fontname, fontfile) 用于 PyMuPDF insert_text。
    优先找 Arial → Liberation Sans → DejaVu Sans → 内置 helv。"""
    global _FONT_CACHE
    if _FONT_CACHE:
        return _FONT_CACHE
    candidates = [
        # macOS
        ('Arial', '/System/Library/Fonts/Supplemental/Arial.ttf'),
        ('Arial', '/Library/Fonts/Arial.ttf'),
        # Linux: fonts-liberation (apt install fonts-liberation)
        ('Arial', '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'),
        ('Arial', '/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf'),
        # Linux: msttcorefonts
        ('Arial', '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf'),
        # Linux: DejaVu Sans 作为兜底
        ('DejaVu Sans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
        ('DejaVu Sans', '/usr/share/fonts/dejavu/DejaVuSans.ttf'),
    ]
    for fname, fpath in candidates:
        if os.path.exists(fpath):
            _FONT_CACHE = (fname, fpath)
            return _FONT_CACHE
    # 没有任何 TTF 时使用 PyMuPDF 内置 Helvetica
    _FONT_CACHE = ('helv', None)
    return _FONT_CACHE


# ═══════════════════════════════════════════════
#  工具函数
# ═══════════════════════════════════════════════

_FONT_OBJ_CACHE = None  # fitz.Font 对象缓存（用于文本宽度测量）


def _get_font_obj():
    """返回 fitz.Font 对象用于宽度测量。

    fitz.get_text_length() 不接受 fontfile 参数（PyMuPDF 1.26），无法按自定义字体
    测量宽度；改用 fitz.Font(text_length) 即可按 Arial/替代字体精确测量。
    """
    global _FONT_OBJ_CACHE
    if _FONT_OBJ_CACHE:
        return _FONT_OBJ_CACHE
    font_name, font_file = _get_font()
    try:
        if font_file:
            _FONT_OBJ_CACHE = fitz.Font(font_name, font_file)
        else:
            _FONT_OBJ_CACHE = fitz.Font(font_name)
    except Exception:
        _FONT_OBJ_CACHE = fitz.Font('helv')
    return _FONT_OBJ_CACHE


def _fit_text(page, text, point, font_name, font_file, fontsize, max_width, min_fontsize=5):
    """写入文本，如果超宽则自动缩小字号直至能在 max_width（点）内放下"""
    if not text:
        return
    size = fontsize
    font_obj = _get_font_obj()
    while size >= min_fontsize:
        w = font_obj.text_length(text, fontsize=size)
        if w <= max_width:
            break
        size -= 0.5
    kwargs = dict(fontsize=size, fontname=font_name, color=(0, 0, 0))
    if font_file:
        kwargs['fontfile'] = font_file
    page.insert_text(fitz.Point(*point), text, **kwargs)


def _write_wrapped(page, text, point, font_name, font_file, fontsize, max_width, max_words=10):
    """写入可自动换行的多行文本（Description of goods）。

    换行规则（满足任一即换行）：
      1. 遇到逗号（, 或 ，）或数据中的换行 → 换行
      2. 每行不超过 max_words（10）个单词
      3. 行宽不超过 max_width（点）
    避免单行溢出页面。
    """
    if not text:
        return
    font_obj = _get_font_obj()
    segments = [seg for seg in re.split(r'[,，\n]+', text) if seg.strip()]
    if not segments:
        return
    lines = []
    cur = []
    for seg in segments:
        for w in seg.split():
            if not cur:
                cur = [w]
                continue
            candidate = cur + [w]
            too_wide = font_obj.text_length(' '.join(candidate), fontsize=fontsize) > max_width
            if len(cur) >= max_words or too_wide:
                lines.append(cur)
                cur = [w]
            else:
                cur = candidate
        if cur:  # 逗号/换行分隔 → 另起一行
            lines.append(cur)
            cur = []
    if cur:
        lines.append(cur)
    x, y = point
    kwargs = dict(fontsize=fontsize, fontname=font_name, color=(0, 0, 0))
    if font_file:
        kwargs['fontfile'] = font_file
    line_h = fontsize * 1.25
    for i, line in enumerate(lines):
        page.insert_text(fitz.Point(x, y + i * line_h), ' '.join(line), **kwargs)


def _safe_str(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    return str(v)


def _fmt_date(d):
    if isinstance(d, datetime):
        return d
    if isinstance(d, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=d)
    return d


def _sanitize(s):
    s = str(s).replace('/', '_').replace('\\', '_').replace(':', '_')
    s = s.replace('*', '_').replace('?', '_').replace('"', '_')
    s = s.replace('<', '_').replace('>', '_').replace('|', '_')
    s = s.replace(' ', '_')
    return s


# ═══════════════════════════════════════════════
#  数据提取
# ═══════════════════════════════════════════════

def _split_jtt_cell(raw):
    """拆分 JTT no. 单元格中的多个单号。

    单元格可能是一个单号，也可能是多个单号用逗号/顿号/斜杠/空格/换行等分隔：
        "JTT202605000328"                        → ['JTT202605000328']
        "JTT202605000328,340,330,331,334,335"    → ['JTT202605000328','JTT202605000340',...]
        "JTT202605000328/JTT202605000340"        → 两个完整单号
    裸序号（无 JTT 前缀）自动补全基础前缀。
    """
    parts = [p.strip() for p in re.split(r'[,，、;；/\s]+', str(raw)) if p.strip()]
    if not parts:
        return []
    # 只有首段是 JTT 编号才算 JTT 单元格；备注说明行（首段如 "2."/"备注"）原样返回
    if not parts[0].startswith('JTT'):
        return parts
    # 基础前缀：JTT + YYYYMM + "000"（12 位，与 _fmt_jtts 的 PREFIX_LEN 一致）
    # JTT202605000328 → 前缀 JTT202605000 + 序号 328
    base = parts[0][:12]
    return [p if p.startswith('JTT') else base + p for p in parts]


def _load_shipments(excel_path):
    """从 Excel 的含"提单信息"的 sheet 加载数据（不限制文件名和月份）"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 优先找含"提单信息"的 sheet → 其次含"提单" → 默认第一个
    sheet_name = None
    for sn in wb.sheetnames:
        if '提单信息' in sn:
            sheet_name = sn
            break
    if not sheet_name:
        for sn in wb.sheetnames:
            if '提单' in sn:
                sheet_name = sn
                break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]  # 默认第一个

    ws = wb[sheet_name]

    headers = []
    for c in list(ws.iter_rows(min_row=1, max_row=1))[0]:
        headers.append(c.value)

    # ── 向下填充：同组空白行继承上行的 B/L No / 渠道 / 模板 / 船名航次等 ──
    FILL_DOWN_COLS = {
        'B/L No.', '引用模板', '渠道',
        'Ocean Vessel', 'Voy.No',
        'Place of receipt', 'Port of loading',
        'Port of discharge', 'Place of delivery',
        'Container no.', 'collect',
        'Place and date of issue', 'on board date',
        'Shipper', 'Consignee', 'Notify party',
    }

    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    # 构建每列的 fill-down 缓存（从表头映射到列索引）
    fill_cache = {}

    shipments = []
    for row in raw_rows:
        d = dict(zip(headers, row))
        raw_jtt = _safe_str(d.get('JTT no.', '')).strip()
        if not raw_jtt:
            continue  # 完全空行直接跳过
        # 拆分单元格中的多个单号
        jtt_nos = _split_jtt_cell(raw_jtt)
        if not jtt_nos:
            continue
        # 跳过非 JTT 编号的行（如底部备注说明行）
        if not jtt_nos[0].startswith('JTT'):
            continue

        # 向下填充：属于填充列表且当前为 None 的列，从上一条缓存取值
        for col_name in FILL_DOWN_COLS:
            if col_name in d and d[col_name] is None and col_name in fill_cache:
                d[col_name] = fill_cache[col_name]
            elif col_name in d and d[col_name] is not None:
                fill_cache[col_name] = d[col_name]

        template = _safe_str(d.get('引用模板', '')).strip()
        if not template:
            continue
        # 跳过查验
        if _safe_str(d.get('Place of receipt', '')).strip() == '查验':
            continue
        bl_no = _safe_str(d.get('B/L No.', '')).strip()
        if not bl_no:
            continue

        # 单元格含多个单号 → 拆分为多票；箱数/KGS/CBM 只在首个单号保留，
        # 避免同 B/L 合并时重复累加
        for i, jtt in enumerate(jtt_nos):
            item = dict(d)
            item['JTT no.'] = jtt
            if i > 0:
                item['cartons'] = 0
                item['KGS'] = 0
                item['CBM'] = 0
            shipments.append(item)

    wb.close()
    return shipments


def _data_fns():
    """数据提取函数字典"""
    return {
        'bl_no':    lambda s: _safe_str(s.get('B/L No.', '')),
        'vessel':   lambda s: (_safe_str(s.get('Ocean Vessel', '')) + '/' +
                               _safe_str(s.get('Voy.No', ''))) if s.get('Voy.No') else _safe_str(s.get('Ocean Vessel', '')),
        'place_rcpt': lambda s: _safe_str(s.get('Place of receipt', '')),
        'port_load':  lambda s: _safe_str(s.get('Port of loading', '')),
        'port_disc':  lambda s: _safe_str(s.get('Port of discharge', '')),
        'place_delv': lambda s: _safe_str(s.get('Place of delivery', '')),
        'container':  lambda s: _safe_str(s.get('Container no.', '')),
        'marks':    lambda s: _safe_str(s.get('Marks & No.', '')) or 'N/M',
        'cartons':  lambda s: str(int(s.get('cartons', 0) or 0)) + ' CARTONS',
        'desc':     lambda s: _safe_str(s.get('Description of goods', '')),
        'kgs':      lambda s: (_safe_str(s.get('KGS', 0) or 0) + '\nKGS') if s.get('KGS') else '',
        'cbm':      lambda s: (_safe_str(s.get('CBM', 0) or 0) + ' CBM') if s.get('CBM') else '',
        'pdi_date': lambda s: (_fmt_date(s.get('Place and date of issue', '')).strftime('%Y/%m/%d')
                               if isinstance(_fmt_date(s.get('Place and date of issue', '')), datetime)
                               else _safe_str(s.get('Place and date of issue', ''))),
        'ob_date':  lambda s: (_fmt_date(s.get('on board date', '')).strftime('%Y/%m/%d')
                               if isinstance(_fmt_date(s.get('on board date', '')), datetime)
                               else _safe_str(s.get('on board date', ''))),
        'notify_party': lambda s: _safe_str(s.get('Notify party', '')).strip() or 'SAME AS CONSIGNEE',
    }


# ═══════════════════════════════════════════════
#  电放保函生成
# ═══════════════════════════════════════════════

def _gen_telex(shipment, out_dir, jtt_part=None, total_cartons=None):
    if not os.path.exists(TELEX):
        return None

    jtt_no = jtt_part or _safe_str(shipment.get('JTT no.', ''))
    channel = _safe_str(shipment.get('渠道', ''))
    cartons = total_cartons if total_cartons is not None else (shipment.get('cartons', 0) or 0)
    bl_no = _safe_str(shipment.get('B/L No.', ''))
    vessel = _safe_str(shipment.get('Ocean Vessel', ''))
    voy = _safe_str(shipment.get('Voy.No', ''))
    container = _safe_str(shipment.get('Container no.', ''))
    collect_date = _fmt_date(shipment.get('collect', ''))
    shipper_name = _safe_str(shipment.get('Shipper', '')).split('\n')[0].strip()
    consignee_name = _safe_str(shipment.get('Consignee', '')).split('\n')[0].strip()

    fname = f'{jtt_no}{_sanitize(channel)}{cartons}件电放保函.xlsx'
    out_path = os.path.join(out_dir, fname)

    wb = openpyxl.load_workbook(TELEX)
    ws = wb['sheet1']

    ws['E10'] = bl_no
    ws['E12'] = f'{vessel}/{voy}' if voy else vessel
    ws['E14'] = container

    # Shipper / Consignee — 写入数据中的完整信息（名称+地址，多行换行显示），
    # 避免只显示模板内置的第一行公司名导致"信息不全"。
    # 数据缺 Shipper/Consignee 列或为空时，回退到标准公司信息兜底，保证单元格填满
    shipper_full = _safe_str(shipment.get('Shipper', '')).strip() or FALLBACK_SHIPPER
    consignee_full = _safe_str(shipment.get('Consignee', '')).strip() or FALLBACK_CONSIGNEE
    if shipper_full:
        ws.merge_cells('A16:I16')
        ws['A16'] = f'Shipper （发货人）: {shipper_full}'
        ws['A16'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[16].height = 42
    else:
        ws['A16'] = 'Shipper （发货人）                   :'
    if consignee_full:
        ws.merge_cells('A18:I18')
        ws['A18'] = f'Consignee （收货人）: {consignee_full}'
        ws['A18'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[18].height = 42
    else:
        ws['A18'] = 'Consignee （收货人）               :   '

    if isinstance(collect_date, datetime):
        m, d, y = collect_date.month, collect_date.day, collect_date.year
    else:
        m, d, y = '5', '30', '2026'
    ws['C31'] = m
    ws['D31'] = d
    ws['E31'] = y

    wb.save(out_path)
    wb.close()
    return out_path


# ═══════════════════════════════════════════════
#  提单 PDF 生成
# ═══════════════════════════════════════════════

def _get_bl_template(template_type):
    mapping = {
        'by sea':  BL_SEA,
        'by train': BL_TRAIN,
        'by truck': BL_TRUCK,
    }
    return mapping.get(template_type.lower().strip())


def _sea_train_fields(is_train=False):
    """By sea / By train 模板的清除矩形 + 文本插入点"""
    clear_rects = [
        # B/L No
        (421, 55, 509, 66),
        # Notify party data（在蓝标签 y=233-243 之下，Pre-carriage by 蓝标签 y=272 之上）
        (52, 244, 250, 270),
        # Place of receipt（单元格 y=286-300）
        (156, 288, 249, 299),
        # Vessel + Port loading（单元格 y=313-327）
        (52, 315, 154, 325),
        (156, 315, 249, 325),
        # Port discharge + Place delivery（单元格 y=341-354）
        (52, 343, 158, 352),
        (156, 343, 268, 352),
        # Container No
        (52, 535, 118, 545),
        # 货物表逐列（y=382-533，距蓝线 2px）
        (52, 384, 117, 531),
        (121, 384, 208, 531),
        (212, 384, 382, 531),
        (386, 384, 455, 531),
        (459, 384, 511, 531),
        # 底部日期
        (426, 618, 473, 629),
        (395, 645, 455, 654),
        (164, 697, 213, 708),
    ]
    if is_train:
        clear_rects.append((52, 314, 249, 326))

    # 字体大小参照模板（baseline 取在单元格中上部，避免 descender 穿过下边框）:
    #   place_rcpt / port_load: 8.5（模板 Calibri 10.6 → 缩至 8.5，留足蓝框边距）
    #   vessel:                  7.5（模板 SimSun 7.5；超宽自动缩小）
    #   port_disc / place_delv:  9.0（10.5 → 9.0，避免文字贴近/超出下蓝线）
    #   bl_no / container / notify_party: 10.5
    #   ArialMT 9 → marks/cartons/cbm/dates 9；desc/kgs 10.5
    #   第5项 max_width（点）— 超出则自动缩小字号
    text_inserts = [
        ((428, 67),   'bl_no',        10.5, 85),
        ((56, 267),   'notify_party', 10.5, 190),
        ((161, 296),  'place_rcpt',    8.5, 85),
        ((52, 320.5), 'vessel',        7.5, 104),
        ((161, 322),  'port_load',     8.5, 85),
        ((52, 350.5), 'port_disc',     9.0, 100),
        ((161, 350.5), 'place_delv',   9.0, 90),
        ((56, 544),   'container',    10.5, 60),
        ((77, 395),   'marks',        9,    58),
        ((130, 395),  'cartons',      9,    78),
        ((218, 395),  'desc',        10.5, 158),
        ((390, 395),  'kgs',         10.5,  58),
        ((459, 395),  'cbm',          9,    48),
        ((428, 628),  'ob_date',      9,    42),
        ((397, 654),  'pdi_date',     9,    55),
        ((166, 706),  'ob_date',      9,    45),
    ]
    return clear_rects, text_inserts


def _truck_fields():
    """By truck 模板的清除矩形 + 文本插入点"""
    clear_rects = [
        (480, 14, 560, 32),
        (35, 288, 310, 310),
        (35, 310, 310, 340),
        (35, 355, 590, 455),
        (375, 665, 460, 695),
        (180, 718, 240, 750),
    ]
    text_inserts = [
        ((488, 18),   'bl_no',        10.5, 75),
        ((37, 292),   'vessel',       9,   155),
        ((202, 292),  'port_load',    9,   100),
        ((37, 316),   'port_disc',    9,   155),
        ((202, 316),  'place_delv',   9,   100),
        ((42, 360),   'container',    8,   250),
        ((46, 362),   'marks',        9,   245),
        ((298, 362),  'desc',        10.5, 130),
        ((435, 362),  'kgs',          8,   145),
        ((435, 372),  'cbm',          8,   145),
        ((380, 683),  'pdi_date',     9,    75),
        ((182, 739),  'ob_date',      9,    52),
    ]
    return clear_rects, text_inserts


def _gen_bl(shipment, out_dir, jtt_part=None, total_cartons=None):
    template_type = _safe_str(shipment.get('引用模板', ''))
    template_path = _get_bl_template(template_type)
    if not template_path or not os.path.exists(template_path):
        return None

    jtt_no = jtt_part or _safe_str(shipment.get('JTT no.', ''))
    channel = _safe_str(shipment.get('渠道', ''))
    cartons = total_cartons if total_cartons is not None else (shipment.get('cartons', 0) or 0)
    fname = f'{jtt_no}{_sanitize(channel)}{cartons}件提单.pdf'
    out_path = os.path.join(out_dir, fname)

    F = _data_fns()
    tt = template_type.lower().strip()

    if tt in ('by sea', 'by train'):
        clear_rects, inserts = _sea_train_fields(tt == 'by train')
    elif tt == 'by truck':
        clear_rects, inserts = _truck_fields()
    else:
        return None

    try:
        doc = fitz.open(template_path)
        page = doc[0]

        # 阶段1：清除旧数据
        for rect in clear_rects:
            page.add_redact_annot(fitz.Rect(*rect), fill=(1, 1, 1))
        page.apply_redactions()

        # 阶段2：写入新数据（使用 Arial / 替代字体，超宽自动缩小）
        font_name, font_file = _get_font()
        for pt, field_key, fontsize, max_width in inserts:
            text = F[field_key](shipment)
            if text:
                if field_key == 'desc':
                    # Description of goods 多行自动换行（每行最多 10 词，超宽自动换行）
                    _write_wrapped(page, text, pt, font_name, font_file, fontsize, max_width)
                else:
                    _fit_text(page, text, pt, font_name, font_file, fontsize, max_width)

        doc.save(out_path, garbage=4, deflate=True)
        doc.close()
        return out_path
    except Exception:
        return None


# ═══════════════════════════════════════════════
#  合并逻辑 — 同一 B/L No 的多票合并为一票
# ═══════════════════════════════════════════════

def _fmt_jtts(jtt_list):
    """格式化 JTT 号列表用于文件名

    单票: JTT202605000307
    多票: JTT202605000364,353 （共享前缀 + 逗号分隔的序号）
    """
    if len(jtt_list) == 1:
        return jtt_list[0]
    # 所有 JTT 号格式: "JTT" + 年月日 + 序号，前12字符为公共前缀
    # 格式化为: JTT202605000364,353
    PREFIX_LEN = 12  # "JTT202605000"
    prefix = jtt_list[0][:PREFIX_LEN]
    if all(j.startswith(prefix) for j in jtt_list):
        suffixes = [j[PREFIX_LEN:] for j in jtt_list]
        return prefix + ','.join(suffixes)
    # fallback: 用 os.path.commonprefix
    prefix = os.path.commonprefix(jtt_list)
    if len(prefix) >= 6:
        suffixes = [j[len(prefix):] for j in jtt_list]
        return prefix + ','.join(suffixes)
    return ','.join(jtt_list)


def _merge_shipments(group):
    """合并同一 B/L 的多票货为一个 shipment dict

    多行数据用 \\n 拼接，cartons/KGS/CBM 累加
    """
    merged = dict(group[0])

    marks_list = []
    desc_list = []
    total_cartons = 0
    total_kgs = 0.0
    total_cbm = 0.0

    for s in group:
        m = str(s.get('Marks & No.', '') or '').strip()
        if m and m not in marks_list:
            marks_list.append(m)
        d = str(s.get('Description of goods', '') or '').strip()
        if d and d not in desc_list:
            desc_list.append(d)
        total_cartons += int(s.get('cartons', 0) or 0)
        try:
            total_kgs += float(s.get('KGS', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            total_cbm += float(s.get('CBM', 0) or 0)
        except (ValueError, TypeError):
            pass

    merged['Marks & No.'] = '\n'.join(marks_list) if marks_list else ''
    merged['Description of goods'] = '\n'.join(desc_list) if desc_list else ''
    merged['cartons'] = total_cartons
    merged['KGS'] = round(total_kgs, 2)
    merged['CBM'] = round(total_cbm, 3)

    merged['_jtt_list'] = [str(s.get('JTT no.', '') or '').strip() for s in group]
    return merged


# ═══════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════

def generate_bl_docs(excel_path, output_dir=None):
    """
    生成提单 + 电放保函 ZIP

    参数:
        excel_path: 上传的 Excel 文件路径（含 '5月提单信息' sheet）
        output_dir: 输出目录，None 则自动创建临时目录

    返回:
        zip_path: 生成的 ZIP 文件路径（内含所有 提单PDF + 电放保函xlsx）
    """
    if output_dir is None:
        output_dir = tempfile.mkdtemp(prefix='bl_docs_')
    else:
        os.makedirs(output_dir, exist_ok=True)

    # 加载数据
    shipments = _load_shipments(excel_path)

    if not shipments:
        raise ValueError("Excel 中未找到有效的提单数据（需要 '5月提单信息' sheet，且含有效的 B/L No.）")

    # 按 B/L No 分组（同一提单的多票合并输出）
    bl_groups = defaultdict(list)
    for s in shipments:
        bl_no = _safe_str(s.get('B/L No.', '')).strip()
        bl_groups[bl_no].append(s)

    # 生成所有文件到临时目录
    telex_ok = bl_ok = 0
    for bl_no, group in bl_groups.items():
        if len(group) == 1:
            # 单票 — 直接生成
            s = group[0]
            jtt_no = _safe_str(s.get('JTT no.', ''))
            cartons = int(s.get('cartons', 0) or 0)
            if _gen_telex(s, output_dir, jtt_part=jtt_no, total_cartons=cartons):
                telex_ok += 1
            if _gen_bl(s, output_dir, jtt_part=jtt_no, total_cartons=cartons):
                bl_ok += 1
        else:
            # 多票合并 — 汇总数据
            merged = _merge_shipments(group)
            jtt_part = _fmt_jtts(merged['_jtt_list'])
            total_cartons = merged['cartons']
            if _gen_telex(merged, output_dir, jtt_part=jtt_part, total_cartons=total_cartons):
                telex_ok += 1
            if _gen_bl(merged, output_dir, jtt_part=jtt_part, total_cartons=total_cartons):
                bl_ok += 1

    if telex_ok == 0 and bl_ok == 0:
        raise ValueError("未能生成任何文件，请检查数据格式")

    # 打包为 ZIP
    zip_name = f'提单电放保函_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
    zip_path = os.path.join(output_dir, zip_name)

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname in sorted(os.listdir(output_dir)):
            fpath = os.path.join(output_dir, fname)
            if fname.endswith('.pdf') or fname.endswith('.xlsx'):
                if fname != zip_name:
                    zf.write(fpath, fname)

    return zip_path, telex_ok, bl_ok
