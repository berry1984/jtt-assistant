#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
英美专版成本报价匹配工具（ym_cost_match）
==========================================
按 `TR成本报价分析运行指令文档-英美专版-ClaudeCode.md`（v2.0）实现。

核心流程：
    读取 JTT+英美渠道匹配表（渠道对应关系）
        ↓
    读取输出模板（已含 JTT 每周报价数据：时间段/国家/渠道/仓库/价格列）
        ↓
    解析 4 份英美供应商文件（美线 / 英欧线 / 加拿大 / 空派）
        ↓
    逐行匹配：JTT渠道 → 英美渠道 → 供应商渠道块 → 仓点/邮编 → 成本单价
        ↓
    计算预估利润（美国 5%、其余 8%）
        ↓
    填写模板供应商列（美国/加拿大 H–M，欧洲/英国 I–N），输出 Excel

用法（CLI）:
    python ym_cost_match.py --supplier-dir 英美
    python ym_cost_match.py --mapping ... --template ... --supplier-dir ... --output ...

规则要点（对应指令文档 §三）:
    - 渠道不在匹配表 / 匹配表为 "/" → 供应商列全部填 "/"
    - 成本单价 = 华南地区最高重量段 KG 单价，逐行独立匹配仓点
    - 多渠道（换行分隔）→ 各渠道都查，取最高价，渠道名列填多个
    - 报价单价 = 该区段最高重量段报价列的值；利润 = 报价 - 报价×扣点 - 成本
"""

import os
import re
import sys
import glob
import argparse
from datetime import datetime

from openpyxl import load_workbook

THIS_DIR = os.path.dirname(os.path.abspath(__file__))

# ─────────────────────────────────────────────────────────────
#  文件/Sheet 常量
# ─────────────────────────────────────────────────────────────
US_HAIKA_SHEETS = ['美森海卡', 'EXX海卡', '合德以星海卡', '美东纽约快线',
                   '美西普船海卡渠道', '美中普船海卡渠道', '美东普船海卡渠道']
US_HAIPAI_SHEETS = ['美国海派']
EU_SHEETS = ['欧洲海运KG', '苏新号中欧卡航', '欧洲铁路']
UK_SHEETS = ['英国海运', '苏新号中英卡航', '英国铁路']
CA_SHEETS = ['直航快船', '直航统配特惠', '合德美转加', 'COSCO美转加', '美森美转加']
AIR_EU_UK_SHEETS = ['欧洲空派包税', '英国空派普货经济线', '英国空派普货5日提', '英国空派带电']
AIR_US_SHEET = '美国空派普货'

SUPPLIER_PATTERNS = {
    'us':  ['英美跨境-美线*'],
    'eu':  ['英美跨境物流英欧线*'],
    'ca':  ['英美跨境-加拿大*'],
    'air': ['英美跨境空派*'],
}

DATE_RE = re.compile(r'(\d{4})年(\d{1,2})月(\d{1,2})日')

COUNTRY_WORDS = ['德国', '法国', '意大利', '西班牙', '英国', '荷兰', '比利时', '波兰',
                 '捷克', '斯洛伐克', '匈牙利', '罗马尼亚', '希腊', '葡萄牙', '奥地利']


# ─────────────────────────────────────────────────────────────
#  基础工具
# ─────────────────────────────────────────────────────────────

def normalize_channel(name):
    """渠道名归一化：取首行、去括号备注、去空白。"""
    if not name:
        return ''
    s = str(name).split('\n')[0].strip()
    s = re.sub(r'[（(][^）)]*[）)]', '', s).strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def extract_date(filename):
    """从供应商文件名提取生效日期，格式 YYYY/M/D。"""
    m = DATE_RE.search(os.path.basename(filename))
    if not m:
        return ''
    return f'{int(m.group(1))}/{int(m.group(2))}/{int(m.group(3))}'


def _num(row, col):
    """取行中指定列的数值（>0），否则 None。"""
    if col is None or col >= len(row):
        return None
    v = row[col]
    if v is None:
        return None
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _split_codes(wh):
    """拆分供应商仓点代码串：去括号备注，按 ,，、/ \\ 空白 拆分。"""
    w = re.sub(r'[（(][^）)]*[）)]', '', str(wh))
    return [c.strip() for c in re.split(r'[，,、/\\\s]+', w) if c.strip()]


def _leading_digit(s):
    """字符串中第一个数字。"""
    m = re.search(r'\d', s)
    return int(m.group(0)) if m else None


def _zone_digits(zone_str):
    """分区字符串中的数字集合，如 '美西（8.9）' → {8,9}。"""
    return set(int(c) for c in re.findall(r'\d', str(zone_str)))


# ─────────────────────────────────────────────────────────────
#  渠道匹配表
# ─────────────────────────────────────────────────────────────

def parse_channel_mapping(path):
    """读取 供应商渠道匹配 Sheet → {JTT渠道名: [英美渠道名, ...]}。"""
    mapping = {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f'  ⚠️ 渠道匹配表读取失败: {e}')
        return mapping
    if '供应商渠道匹配' not in wb.sheetnames:
        wb.close()
        return mapping
    ws = wb['供应商渠道匹配']
    for r in ws.iter_rows(values_only=True):
        if not r or not r[0]:
            continue
        jtt = str(r[0]).strip().split('\n')[0].strip()
        if not jtt or jtt in ('JTT', '渠道名称'):
            continue
        val = str(r[1] or '').strip()
        chans = [c.strip() for c in re.split(r'[\n/]', val)
                 if c.strip() and c.strip() not in ('/', '')]
        mapping[jtt] = chans
    wb.close()
    return mapping


# ─────────────────────────────────────────────────────────────
#  供应商文件解析
# ─────────────────────────────────────────────────────────────

def discover_supplier_files(supplier_dir):
    """在目录中自动发现 4 份供应商文件。"""
    found = {}
    for ftype, pats in SUPPLIER_PATTERNS.items():
        for p in pats:
            g = [x for x in glob.glob(os.path.join(supplier_dir, p))
                 if x.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
            if g:
                found[ftype] = sorted(g)[-1]
                break
    return found


REGION_KEYS = ['义乌', '东莞', '华南', '深圳', '福州', '厦门', '泉州',
               '华中', '华东', '华北', '合肥', '青岛', '临沂', '国家']


def _detect_layout(rows):
    """探测华南地区起始列 + 华南各重量段列 + 最高段列。
    表头行 = A列含「渠道名称/产品渠道名称」，重量段表头在其下一行。
    返回 {'south_start': int|None, 'best_tier': int|None, 'header_row': int, 'tier_row': int}
    """
    # 1. 找表头行（A 列含 渠道名称）
    header_row = None
    for i, r in enumerate(rows[:6]):
        if r and r[0] and '渠道名称' in str(r[0]):
            header_row = i
            break
    if header_row is None:
        return {'south_start': None, 'best_tier': None, 'header_row': -1, 'tier_row': -1}
    hdr = rows[header_row]
    region_cols = [(ci, str(v)) for ci, v in enumerate(hdr)
                   if v is not None and any(k in str(v) for k in REGION_KEYS)]
    south = None
    for ci, s in region_cols:
        if '华南' in s or '东莞' in s:
            south = ci
            break
    end = next((ci for ci, _ in region_cols if ci > (south if south is not None else -1)),
               len(hdr))
    # 2. 重量段表头行（紧接表头行下方）
    tier_row = header_row + 1
    best = None
    if tier_row < len(rows) and south is not None:
        thr = rows[tier_row]
        if any(re.search(r'\d+\s*KG', str(v or '')) for v in thr):
            kg = [ci for ci in range(south, min(end, len(thr)))
                  if re.search(r'\d+\s*KG', str(thr[ci] or ''))]
            kg_clean = [ci for ci in kg
                        if not any(k in str(thr[ci] or '') for k in ['CBM', '方'])]
            best = (kg_clean or kg)[-1] if (kg_clean or kg) else None
    return {'south_start': south, 'best_tier': best,
            'header_row': header_row, 'tier_row': tier_row}


def _is_channel_line(a):
    """渠道块起始行判断：首行以 字母+数字 开头（如 B3、N30、C6）。"""
    return bool(a) and re.match(r'^[A-Z]\d', a)


def _maybe_add(cur, r, wh_col, tier_col):
    wh = str(r[wh_col] or '').strip() if wh_col < len(r) else ''
    price = _num(r, tier_col)
    if wh and price is not None:
        cur['rows'].append({'wh': wh, 'price': price})


def _parse_us(path):
    """美线文件：海卡类 sheets（华南 H–K 取 350KG+）+ 美国海派（华南 E–F 取 101KG起）。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    blocks = []
    for sn in wb.sheetnames:
        if sn not in US_HAIKA_SHEETS and sn not in US_HAIPAI_SHEETS:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        layout = _detect_layout(rows)
        if layout['best_tier'] is None:
            print(f'  ⚠️ {sn}: 未探测到华南重量段列，跳过')
            continue
        kind = 'haiqia' if sn in US_HAIKA_SHEETS else 'haipai'
        wh_col = 1  # B 列：海卡=仓库代码，海派=分区
        cur = None
        for r in rows[layout['tier_row'] + 1:]:
            if not r:
                continue
            a = str(r[0] or '').strip().split('\n')[0].strip() if r[0] else ''
            if _is_channel_line(a):
                if '按方' in a:
                    cur = None
                    continue
                cur = {'name': a, 'sheet': sn, 'kind': kind, 'rows': []}
                blocks.append(cur)
                _maybe_add(cur, r, wh_col, layout['best_tier'])
                continue
            if cur is None:
                continue
            _maybe_add(cur, r, wh_col, layout['best_tier'])
    wb.close()
    return blocks


def _parse_eu(path):
    """英欧线文件：欧洲海运KG/卡航/铁路 + 英国海运/卡航/铁路。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    blocks = []
    for sn in wb.sheetnames:
        if sn not in EU_SHEETS and sn not in UK_SHEETS:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        layout = _detect_layout(rows)
        if layout['best_tier'] is None:
            print(f'  ⚠️ {sn}: 未探测到华南重量段列，跳过')
            continue
        kind = 'uk' if sn in UK_SHEETS else 'eu'
        wh_col = 1  # B 列：国家/仓点
        cur = None
        for r in rows[layout['tier_row'] + 1:]:
            if not r:
                continue
            a = str(r[0] or '').strip().split('\n')[0].strip() if r[0] else ''
            if _is_channel_line(a):
                if '按方' in a:
                    cur = None
                    continue
                cur = {'name': a, 'sheet': sn, 'kind': kind, 'rows': []}
                blocks.append(cur)
                _maybe_add(cur, r, wh_col, layout['best_tier'])
                continue
            if cur is None:
                continue
            _maybe_add(cur, r, wh_col, layout['best_tier'])
    wb.close()
    return blocks


def _parse_ca(path):
    """加拿大文件：C 列仓点代码（/ 分隔），华南 D–F 取 100KG。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    blocks = []
    for sn in wb.sheetnames:
        if sn not in CA_SHEETS:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        layout = _detect_layout(rows)
        if layout['best_tier'] is None:
            print(f'  ⚠️ {sn}: 未探测到华南重量段列，跳过')
            continue
        wh_col = 2  # C 列：仓库代码
        cur = None
        for r in rows[layout['tier_row'] + 1:]:
            if not r:
                continue
            a = str(r[0] or '').strip().split('\n')[0].strip() if r[0] else ''
            if _is_channel_line(a):
                if '按方' in a:
                    cur = None
                    continue
                cur = {'name': a, 'sheet': sn, 'kind': 'ca', 'rows': []}
                blocks.append(cur)
                _maybe_add(cur, r, wh_col, layout['best_tier'])
                continue
            if cur is None:
                continue
            _maybe_add(cur, r, wh_col, layout['best_tier'])
    wb.close()
    return blocks


def _parse_air(path):
    """空派文件：欧洲空派包税(C块) + 英国空派(L块) + 美国空派普货(M块)。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    blocks = []
    for sn in wb.sheetnames:
        if sn in AIR_EU_UK_SHEETS:
            blocks.extend(_parse_air_country_sheet(wb[sn], sn))
        elif sn == AIR_US_SHEET:
            blocks.extend(_parse_air_us_sheet(wb[sn]))
    wb.close()
    return blocks


def _find_tier_col(rows, start, span=4):
    """从 start 行起向下找含 KG 的表头行，返回最后一个重量档列（含大小写）。"""
    for h in rows[start:start + span]:
        if h and any(re.search(r'\d+\s*KG', str(v or ''), re.IGNORECASE) for v in h):
            tiers = [ci for ci, v in enumerate(h)
                     if re.search(r'\d+\s*KG', str(v or ''), re.IGNORECASE)]
            if tiers:
                return tiers[-1]
    return None


def _parse_air_country_sheet(ws, sn):
    """欧洲/英国空派：块首行 A=渠道名，数据行 A=国家（空派仓点在国家列），取最高重量档。"""
    rows = list(ws.iter_rows(values_only=True))
    blocks = []
    cur = None
    for i, r in enumerate(rows[1:], start=1):
        if not r:
            continue
        a = str(r[0] or '').strip().split('\n')[0].strip() if r[0] else ''
        if _is_channel_line(a):
            if '按方' in a:
                cur = None
                continue
            cur = {'name': a, 'sheet': sn, 'kind': 'air_country',
                   'rows': [], 'tier_col': None}
            blocks.append(cur)
            cur['tier_col'] = _find_tier_col(rows, i + 1)
            _maybe_add(cur, r, 0, cur['tier_col'])
            continue
        if cur is None:
            continue
        _maybe_add(cur, r, 0, cur['tier_col'])
    return blocks


def _parse_air_us_sheet(ws):
    """美国空派普货：块首行 A=渠道名，数据行 A=分区，取最高重量档。"""
    rows = list(ws.iter_rows(values_only=True))
    blocks = []
    cur = None
    for i, r in enumerate(rows[1:], start=1):
        if not r:
            continue
        a = str(r[0] or '').strip().split('\n')[0].strip() if r[0] else ''
        if _is_channel_line(a):
            cur = {'name': a, 'sheet': AIR_US_SHEET, 'kind': 'air_zone',
                   'rows': [], 'tier_col': None}
            blocks.append(cur)
            cur['tier_col'] = _find_tier_col(rows, i + 1)
            _maybe_add(cur, r, 0, cur['tier_col'])
            continue
        if cur is None:
            continue
        _maybe_add(cur, r, 0, cur['tier_col'])
    return blocks


# ─────────────────────────────────────────────────────────────
#  渠道/仓点 匹配
# ─────────────────────────────────────────────────────────────

def _block_tokens(block):
    """渠道块名拆成多个渠道 token（按 换行/斜杠 拆分，用于 F1/F2 同块场景）。"""
    toks = []
    for part in re.split(r'[\n/]', block['name']):
        p = normalize_channel(part)
        if p:
            toks.append(p)
    return toks


def match_ym_channel(ym_name, blocks):
    """在已解析的供应商渠道块中定位英美渠道块。"""
    target = normalize_channel(ym_name)
    if not target:
        return None
    for b in blocks:
        for t in _block_tokens(b):
            if t == target:
                return b
    # 子串回退（需足够长度避免误配）
    for b in blocks:
        for t in _block_tokens(b):
            if len(target) >= 4 and (target in t or t in target):
                return b
    return None


def _match_code_membership(block, jtt_d):
    """仓点代码成员匹配（美卡派/加拿大）。同一仓点多行时取最高价。"""
    code = jtt_d.strip().upper()
    if not code:
        return None, None
    best, best_wh = None, None
    for row in block['rows']:
        wh = row['wh']
        for c in _split_codes(wh):
            if c == code or (len(code) >= 4 and c.startswith(code)):
                if best is None or row['price'] > best:
                    best, best_wh = row['price'], wh
                break
    return (best, best_wh) if best is not None else (None, None)


def _match_zip_zone(block, jtt_d):
    """邮编区间 → 分区匹配（美快递派/美国空派）。"""
    ld = _leading_digit(jtt_d)
    if ld is None:
        return None, None
    for row in block['rows']:
        if ld in _zone_digits(row['wh']):
            return row['price'], row['wh']
    return None, None


def _match_eu(block, jtt_d):
    """欧洲：仓点-邮编 提取仓点代码；国家名精确匹配。"""
    d = jtt_d.strip()
    if not d:
        return None, None
    if not re.search(r'[A-Z]{3}\d', d):
        # 国家名
        for row in block['rows']:
            wh = row['wh'].strip()
            if wh and (wh == d or wh.startswith(d)):
                return row['price'], wh
        return None, None
    m = re.search(r'([A-Z]{3}\d+)', d)
    code = m.group(1) if m else d
    for row in block['rows']:
        wh = row['wh'].strip()
        if code and (code in _split_codes(wh) or code in wh):
            return row['price'], wh
    return None, None


def _match_uk(block, jtt_d):
    """英国：仓点代码逐码匹配；FBA仓 → 无仓点代码的国家行。"""
    d = jtt_d.strip()
    if not d:
        return None, None
    if 'FBA' in d.upper():
        for row in block['rows']:
            wh = row['wh'].strip()
            if wh and not re.search(r'[A-Z]{3}\d', wh):
                return row['price'], wh
        return None, None
    codes = [c for c in re.split(r'[、，,/\\\s]+', d) if c.strip()]
    for row in block['rows']:
        wh = row['wh'].strip()
        for c in codes:
            if c and (c in _split_codes(wh) or c in wh):
                return row['price'], wh
    return None, None


def _extract_country_from_channel(ch):
    for w in COUNTRY_WORDS:
        if ch.startswith(w):
            return w
    return ''


def _match_country(block, country):
    for row in block['rows']:
        wh = row['wh'].strip()
        if wh and (wh == country or wh.startswith(country)):
            return row['price'], wh
    return None, None


def match_warehouse(block, jtt_channel, jtt_d):
    """按渠道块类型分发仓点匹配。"""
    kind = block['kind']
    if kind in ('haiqia', 'ca'):
        return _match_code_membership(block, jtt_d)
    if kind in ('haipai', 'air_zone'):
        return _match_zip_zone(block, jtt_d)
    if kind == 'eu':
        return _match_eu(block, jtt_d)
    if kind == 'uk':
        return _match_uk(block, jtt_d)
    if kind == 'air_country':
        country = _extract_country_from_channel(jtt_channel) or '英国'
        return _match_country(block, country)
    return None, None


# ─────────────────────────────────────────────────────────────
#  模板解析与填充
# ─────────────────────────────────────────────────────────────

def _parse_template_sheet(ws, default_sup):
    """解析模板一个 Sheet，返回分区列表（含表头行、报价档位、供应商列起点、数据行）。"""
    rows = list(ws.iter_rows(values_only=True))
    sections = []
    cur = None
    data = []
    last_tiers = {}
    last_sup = None
    for i, r in enumerate(rows):
        joined = ''.join(str(c or '') for c in r[:14])
        if '系统下单渠道名称' in joined:
            if cur:
                cur['data'] = data
                sections.append(cur)
            tiers = {}
            sup = None
            for ci, v in enumerate(r):
                s = str(v or '').strip()
                m = re.search(r'(\d+)\s*KG', s)
                if m:
                    tiers[int(m.group(1))] = ci
                if '供应商1' in s and sup is None:
                    sup = ci
            if tiers:
                last_tiers = tiers
            if sup is not None:
                last_sup = sup
            cur = {'header_row': i, 'tiers': last_tiers,
                   'supplier_start': last_sup if last_sup is not None else default_sup}
            data = []
            continue
        if cur is None:
            continue
        data.append((i, r))
    if cur:
        cur['data'] = data
        sections.append(cur)
    return sections


def _quote_price(r, tiers):
    """取该行报价单价：最高重量段列，若空则回退下一档。"""
    if not tiers:
        return None
    for tier in sorted(tiers.keys(), reverse=True):
        col = tiers[tier]
        if col < len(r):
            v = r[col]
            if isinstance(v, (int, float)) and v > 0:
                return float(v)
    return None


def _unmerge_target(ws, row, col):
    """若目标单元格处于合并区域，先解除合并（避免 MergedCell 只读报错）。"""
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.unmerge_cells(str(mr))


def _fill_row(ws, excel_row, sup, vals):
    for off, v in enumerate(vals):
        col = sup + off + 1  # sup 为 0-based，转为 openpyxl 1-based 列
        _unmerge_target(ws, excel_row, col)
        ws.cell(row=excel_row, column=col).value = v


def _fill_slash(ws, excel_row, sup):
    _fill_row(ws, excel_row, sup, ['/'] * 6)


# ─────────────────────────────────────────────────────────────
#  主流程
# ─────────────────────────────────────────────────────────────

def generate(template_path, mapping_path, supplier_dir, output_path):
    """执行完整匹配并输出 Excel。返回 (输出路径, 统计)。"""
    mapping = parse_channel_mapping(mapping_path)
    print(f'渠道匹配表: {len(mapping)} 条映射')

    supplier_files = discover_supplier_files(supplier_dir)
    if len(supplier_files) < 4:
        print(f'  ⚠️ 仅发现 {len(supplier_files)}/4 份供应商文件: {list(supplier_files.keys())}')
    blocks = []
    dates = {}
    for ftype in ('us', 'eu', 'ca', 'air'):
        fpath = supplier_files.get(ftype)
        if not fpath:
            print(f'  ⚠️ 缺少 {ftype} 供应商文件')
            continue
        if ftype == 'us':
            b = _parse_us(fpath)
        elif ftype == 'eu':
            b = _parse_eu(fpath)
        elif ftype == 'ca':
            b = _parse_ca(fpath)
        else:
            b = _parse_air(fpath)
        dates[ftype] = extract_date(fpath)
        for blk in b:
            blk['file'] = ftype
        blocks.extend(b)
        print(f'  [{ftype}] {os.path.basename(fpath)}: {len(b)} 渠道块, 更新日期 {dates[ftype] or "?"}')
    print(f'共解析 {len(blocks)} 个供应商渠道块')

    wb = load_workbook(template_path)  # 保留格式/公式
    stats = {}
    for sheet in ['美国', '加拿大', '欧洲', '英国']:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        default_sup = 7 if sheet in ('美国', '加拿大') else 8
        sections = _parse_template_sheet(ws, default_sup)
        deduction = 0.05 if sheet == '美国' else 0.08
        st = {'rows': 0, 'matched': 0, 'unmatched': 0, 'unmatched_channels': []}
        cur_channel = ''
        for sec in sections:
            sup = sec['supplier_start']
            for (i, r) in sec['data']:
                ch_raw = r[2] if len(r) > 2 else None
                cs = str(ch_raw or '').strip()
                if cs and cs not in ('None', '系统下单渠道名称'):
                    cur_channel = cs.split('\n')[0].strip()
                ch = cur_channel
                if not ch:
                    continue
                d = str(r[3] or '').strip() if len(r) > 3 else ''
                if d in ('仓库/邮编', ''):
                    # 完全空行跳过；有报价但无仓点的行也跳过（如空行）
                    if d == '':
                        continue
                    continue
                quote = _quote_price(r, sec['tiers'])
                ym = mapping.get(ch, [])
                st['rows'] += 1
                if not ym:
                    _fill_slash(ws, i + 1, sup)
                    st['unmatched'] += 1
                    if ch not in st['unmatched_channels']:
                        st['unmatched_channels'].append(ch)
                    continue
                # 逐英美渠道匹配，取最高价
                best = None
                best_wh = None
                best_file = None
                block_found_file = None
                for ymc in ym:
                    block = match_ym_channel(ymc, blocks)
                    if block is None:
                        continue
                    if block_found_file is None:
                        block_found_file = block['file']
                    price, wh = match_warehouse(block, ch, d)
                    if price is not None and (best is None or price > best):
                        best, best_wh, best_file = price, wh, block['file']
                if best is None:
                    date = dates.get(block_found_file or best_file, '')
                    _fill_row(ws, i + 1, sup,
                              ['英美', '\n'.join(ym), '/', date, '/', '/'])
                    st['unmatched'] += 1
                    continue
                date = dates.get(best_file, '')
                cost = round(best, 2)
                profit = round(quote - quote * deduction - best, 2) if quote is not None else '/'
                _fill_row(ws, i + 1, sup,
                          ['英美', '\n'.join(ym), best_wh, date, cost, profit])
                st['matched'] += 1
        stats[sheet] = st
        print(f'  {sheet}: {st["rows"]} 行, 匹配 {st["matched"]}, 无 {st["unmatched"]}')
        if st['unmatched_channels']:
            print(f'      未匹配渠道: {st["unmatched_channels"]}')

    wb.save(output_path)
    return output_path, stats


def _find(base, pattern):
    g = [x for x in glob.glob(os.path.join(base, pattern))
         if x.lower().endswith(('.xlsx', '.xlsm'))]
    return sorted(g)[-1] if g else None


def main():
    parser = argparse.ArgumentParser(
        description='英美专版成本报价匹配工具（v2.0）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    parser.add_argument('--mapping', help='渠道匹配表路径 (默认自动发现)')
    parser.add_argument('--template', help='输出模板路径 (默认自动发现)')
    parser.add_argument('--supplier-dir', help='供应商文件目录 (默认 报价工具/英美)')
    parser.add_argument('--output', help='输出文件路径')
    args = parser.parse_args()

    mapping = args.mapping or _find(THIS_DIR, 'JTT+英美渠道匹配表*.xlsx')
    template = args.template or _find(THIS_DIR, 'JTT物流每周成本报价分析表输出模板-英美*.xlsx')
    supplier_dir = args.supplier_dir
    if not supplier_dir or not os.path.isdir(supplier_dir):
        cand = os.path.join(THIS_DIR, '英美')
        supplier_dir = cand if os.path.isdir(cand) else THIS_DIR
    if not mapping or not template:
        print('错误: 未找到渠道匹配表或输出模板')
        sys.exit(1)
    output = args.output or os.path.join(
        THIS_DIR, f'JTT物流每周成本报价分析表-英美_{datetime.now().strftime("%Y%m%d")}.xlsx')

    print('=' * 60)
    print(f'匹配表: {mapping}')
    print(f'模板  : {template}')
    print(f'供应商: {supplier_dir}')
    print('=' * 60)
    path, stats = generate(template, mapping, supplier_dir, output)
    print('=' * 60)
    print(f'✅ 输出: {path}')
    total = sum(s['rows'] for s in stats.values())
    matched = sum(s['matched'] for s in stats.values())
    print(f'统计: {total} 行, 匹配 {matched}, 未匹配 {total - matched}')


if __name__ == '__main__':
    main()
