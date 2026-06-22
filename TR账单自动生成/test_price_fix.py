#!/usr/bin/env python3
import sys
sys.path.insert(0, '/Users/admin/bb plan1/TR账单自动生成')
from gen_bill import load_data, build_rows, sort_rows, generate_bill
from datetime import datetime, timedelta

orders, picks, prices, price_rows_raw, decl_groups = load_data(
    '/Users/admin/bb plan1/TR账单自动生成/导出订单列表 2026-06-14.xlsx',
    '/Users/admin/bb plan1/TR账单自动生成/导出拣货数据 2026-06-14.xlsx',
    '/Users/admin/bb plan1/TR账单自动生成/内部-TR每周走货确认价格6.12.xlsx'
)

print(f'Prices: {len(prices)} 个仓库')
for k, v in prices.items():
    print(f'  {k}: price={v["price"]}')

rows = build_rows(orders, picks, prices)
rows = sort_rows(rows, declaration_groups=decl_groups)
print(f'\nRows: {len(rows)} 行')
for r in rows:
    print(f'  SO={r["so"]:20s}  wh={r["wh"]:6s}  weight={r["weight"]:4}  unit_price={r["unit_price"]}')

# Compute date range from actual orders
date_serials = []
for o in orders.values():
    d = o.get('发货日期', o.get('工作日期'))
    if d:
        date_serials.append(d)

base = datetime(1899, 12, 30)
min_dt = base + timedelta(days=min(date_serials))
max_dt = base + timedelta(days=max(date_serials))
mon = min_dt - timedelta(days=min_dt.weekday())
sun = mon + timedelta(days=6)
date_range_str = f"{mon.month}.{mon.day}-{sun.month}.{sun.day}"
title_str = f"至：广州拓锐科技有限公司（{mon.month}.{mon.day}-{sun.month}.{sun.day}）"

success = generate_bill(rows, '/tmp/test_bill_output.xlsx',
    title_str=title_str, date_range_str=date_range_str,
    price_rows_raw=price_rows_raw, year=mon.year,
    declaration_groups=decl_groups)

if success:
    import openpyxl
    wb = openpyxl.load_workbook('/tmp/test_bill_output.xlsx')

    # 检查人民币账单 sheet
    month_nums = [int(x) for x in __import__('re').findall(r'\d+', date_range_str)]
    bill_month = month_nums[0] if month_nums else 6
    sheet_name = f'{bill_month}月人民币账单（已调格式）'
    ws = wb[sheet_name]

    print(f'\n=== {sheet_name} S列(报关费) 和 K列(单价) ===')
    for r in range(4, 4 + len(rows)):
        wh = ws.cell(row=r, column=5).value
        k = ws.cell(row=r, column=11).value  # K=单价
        s = ws.cell(row=r, column=19).value  # S=报关费
        t = ws.cell(row=r, column=20).value  # T=报关费税
        print(f'  Row {r}: wh={str(wh):6s}  K(单价)={str(k):8s}  S(报关费)={str(s):12s}  T(报关税)={str(t):12s}')

    # 合计行
    sr = 4 + len(rows) + 1
    print(f'\n  合计行 (Row {sr}):')
    print(f'  S(报关费)={ws.cell(row=sr, column=19).value}')
    print(f'  T(报关税)={ws.cell(row=sr, column=20).value}')

    # 报价表A
    ws_q = wb['报价表A']
    print(f'\n=== 报价表A ===')
    for r in range(4, 12):
        wh = ws_q.cell(row=r, column=4).value
        price = ws_q.cell(row=r, column=5).value
        if wh:
            print(f'  Row {r}: 仓库={str(wh):6s}  单价={price}')
