"""
供应商报价文件解析模块
========================
支持多 sheet、多文件的复杂报价格式，从各供应商 Excel 中提取渠道与价格。
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 文件名关键词 → 标准供应商名
SUPPLIER_KEYWORDS = {
    "天图/心一": ["天图", "心一"],
    "英美": ["英美"],
    "美琦": ["美琦"],
    "凯鑫": ["凯鑫"],
    "航乐": ["航乐"],
    "皓鹏": ["皓鹏"],
}

# 各国适用的供应商（用于过滤）
COUNTRY_SUPPLIERS = {
    "美国": ["天图/心一", "英美", "美琦", "皓鹏"],
    "加拿大": ["天图/心一", "英美", "美琦", "皓鹏"],
    "欧洲": ["天图/心一", "英美", "凯鑫", "航乐", "皓鹏"],
    "英国": ["天图/心一", "英美", "凯鑫", "航乐", "皓鹏"],
}

# 合理 KG 单价范围（RMB）
MIN_KG_PRICE = 3.0
MAX_KG_PRICE = 85.0

# 每个 sheet 只读取前 N 行 × 前 M 列（报价表头与数据都在前几百行、前几十列）。
# 皓鹏「美国大货中转二级补仓」sheet 声明 1M+ 行 × 16K 列，全量读取会 OOM。
MAX_SHEET_ROWS = 3000
MAX_SHEET_COLS = 60


def detect_supplier_from_path(filepath):
    """从文件路径检测供应商名称。"""
    name = Path(filepath).stem
    for supplier, keywords in SUPPLIER_KEYWORDS.items():
        for kw in keywords:
            if kw in name:
                return supplier
    return None


def discover_supplier_files(supplier_dir):
    """
    扫描目录（含子目录），按供应商合并文件列表。

    返回 { 供应商名: [文件路径, ...] }
    """
    supplier_files = {s: [] for s in SUPPLIER_KEYWORDS}

    for root, _, files in os.walk(supplier_dir):
        for fname in files:
            # 扩展名大小写不敏感（.XLSX / .Xls 都接受）
            if not fname.lower().endswith((".xlsx", ".xls")):
                continue
            if fname.startswith("~$"):
                continue
            fpath = os.path.join(root, fname)
            supplier = detect_supplier_from_path(fpath)
            if supplier:
                supplier_files[supplier].append(fpath)

    return {k: v for k, v in supplier_files.items() if v}


def parse_supplier_files(filepaths, region_filter=True):
    """
    解析一个或多个供应商文件并合并渠道价格。

    返回 {
        "supplier_info": str,
        "update_date": str,
        "channels": { 渠道名: 价格 }
    }
    """
    all_channels = {}
    update_dates = []
    info_parts = []
    ch_sheets_all = {}

    for fp in filepaths:
        parsed = _parse_single_workbook(fp)
        if not parsed:
            continue
        info_parts.append(parsed.get("supplier_info", ""))
        if parsed.get("update_date"):
            update_dates.append(parsed["update_date"])
        for ch, price in parsed.get("channels", {}).items():
            if region_filter and not _is_huanan_channel(ch):
                continue
            # 同一渠道多文件：取较高价（规则5）
            if ch in all_channels:
                try:
                    all_channels[ch] = max(float(all_channels[ch]), float(price))
                except (ValueError, TypeError):
                    pass
            else:
                all_channels[ch] = price
        # 记录渠道首次出现的 sheet（用于运输类别判定）
        for ch, sname in (parsed.get("channel_sheets") or {}).items():
            if ch not in ch_sheets_all:
                ch_sheets_all[ch] = sname

    if not all_channels:
        return None

    return {
        "supplier_info": " | ".join(p for p in info_parts if p),
        "update_date": max(update_dates) if update_dates else "",
        "channels": all_channels,
        "channel_sheets": ch_sheets_all,
    }


def parse_all_suppliers(supplier_dir, region_filter=True):
    """
    自动发现并解析目录下所有供应商报价。

    返回 { 供应商名: parsed_data }
    """
    file_map = discover_supplier_files(supplier_dir)
    result = {}

    for supplier, files in file_map.items():
        logger.info("解析供应商 %s: %d 个文件", supplier, len(files))
        for f in files:
            logger.info("  - %s", os.path.basename(f))
        parsed = parse_supplier_files(files, region_filter=region_filter)
        if parsed:
            result[supplier] = parsed
            logger.info("  → %d 个渠道", len(parsed["channels"]))
        else:
            logger.warning("  → 未解析到渠道")

    return result


def _parse_single_workbook(filepath):
    """解析单个 Excel 工作簿的所有 sheet。"""
    logger.debug("解析工作簿: %s", filepath)
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        logger.warning("无法打开文件 %s: %s", filepath, e)
        return None

    channels = {}
    ch_sheets = {}
    update_date = _extract_date_from_filename(filepath)

    for sheet_name in wb.sheetnames:
        if _skip_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        # 截断读取：报价表头与数据都在 sheet 前几百行、前几十列。
        # 某些文件（如皓鹏「美国大货中转二级补仓」）存在病态行列尺寸
        # （1M+ 行 × 16K 列），全量物化会占用数 GB 内存甚至 OOM。
        rows = list(ws.iter_rows(values_only=True, max_row=MAX_SHEET_ROWS, max_col=MAX_SHEET_COLS))
        if not rows:
            continue

        sheet_channels = _parse_sheet_rows(rows, sheet_name)
        for ch, price in sheet_channels.items():
            if ch not in ch_sheets:
                ch_sheets[ch] = sheet_name  # 记录渠道首次出现的 sheet（运输类别判定用）
            if ch in channels:
                try:
                    channels[ch] = max(float(channels[ch]), float(price))
                except (ValueError, TypeError):
                    pass
            else:
                channels[ch] = price

        sheet_date = _extract_date_from_sheet(rows)
        if sheet_date and (not update_date or sheet_date > update_date):
            update_date = sheet_date

    wb.close()

    if not channels:
        return None

    supplier = detect_supplier_from_path(filepath) or Path(filepath).stem
    return {
        "supplier_info": f"{supplier} {update_date}".strip(),
        "update_date": update_date,
        "channels": channels,
        "channel_sheets": ch_sheets,
    }


def _skip_sheet(name):
    """跳过非报价 sheet。"""
    skip_kw = [
        "目录", "关于", "须知", "必读", "赔偿", "附加", "分区", "仓库",
        "合作", "定位", "偏远", "计划", "船期", "认证", "反倾销", "海外仓",
        "辅助", "指引", "查询表-间接", "快捷查价", "发货须知", "时效参考",
        "装柜计划", "出运计划", "超大件", "罗马尼亚", "产品附加",
    ]
    return any(kw in name for kw in skip_kw)


def _parse_sheet_rows(rows, sheet_name):
    """
    从一个 sheet 的所有行中提取 {渠道名: 价格}。
    策略：找到含渠道名的行，从该行提取 KG 单价（优先最高重量段）。
    """
    channels = {}
    header_tiers = _find_weight_tier_headers(rows)

    # 提取「下单渠道」行的渠道名映射（凯鑫/航乐等格式）
    order_channel_map = _extract_order_channel_rows(rows)
    # 航乐式分段标题解析
    section_channels = _parse_section_title_channels(rows)

    for row_idx, row in enumerate(rows):
        if not row:
            continue

        # 找行内可能的渠道名（前 5 列的字符串）
        channel_candidates = []
        for ci in range(min(5, len(row))):
            val = row[ci]
            if val and isinstance(val, str) and len(val.strip()) >= 4:
                text = val.strip()
                if _looks_like_channel(text):
                    channel_candidates.append(text)

        if not channel_candidates:
            continue

        # 提取价格
        price = _extract_row_price(row, header_tiers, row_idx, rows)
        if price is None:
            continue

        for ch in channel_candidates:
            # 清理渠道名（取第一行、去多余空白）
            clean = _normalize_channel(ch)
            if clean:
                if clean in channels:
                    channels[clean] = max(channels[clean], price)
                else:
                    channels[clean] = price
            # 也存储原始名（含换行）的各行
            for part in ch.split("\n"):
                part = part.strip()
                if part and len(part) >= 4:
                    if part in channels:
                        channels[part] = max(channels[part], price)
                    else:
                        channels[part] = price

    # 合并「下单渠道」映射（如 欧洲铁路B → 价格）
    for order_name, price in order_channel_map.items():
        if order_name in channels:
            channels[order_name] = max(channels[order_name], price)
        else:
            channels[order_name] = price

    for sec_name, price in section_channels.items():
        if sec_name in channels:
            channels[sec_name] = max(channels[sec_name], price)
        else:
            channels[sec_name] = price

    return channels


def _parse_section_title_channels(rows):
    """
    航乐式报价：分段标题即渠道名，下方数据行含各重量段价格。
    如「欧洲铁路包税-专仓卡派专线」→ 101KG+ 价格。
    """
    channels = {}
    current_section = None
    tier_headers = {}

    for row_idx, row in enumerate(rows):
        if not row:
            continue

        r0 = row[0]
        r0_str = str(r0).strip() if r0 else ""

        # 重量段表头行
        row_tiers = _find_weight_tier_headers([row])
        if row_tiers:
            tier_headers = row_tiers
            continue

        # 分段标题行（首列是渠道/产品名）
        if r0_str and len(r0_str) >= 6 and _looks_like_section_title(r0_str):
            current_section = _normalize_channel(r0_str)
            continue

        # 数据行：有价格 + 当前分段
        if current_section and tier_headers:
            price = _extract_row_price(row, tier_headers, row_idx, rows)
            if price is not None:
                if current_section in channels:
                    channels[current_section] = max(channels[current_section], price)
                else:
                    channels[current_section] = price

    return channels


def _looks_like_section_title(text):
    """航乐/凯鑫分段标题特征。"""
    if not _looks_like_channel(text):
        return False
    title_kw = ["包税", "专线", "派送", "不包税", "递延", "卡派", "快递"]
    return any(kw in text for kw in title_kw)


def _extract_order_channel_rows(rows):
    """
    解析「下单渠道」行（凯鑫等格式）。
    将下单渠道名映射到同列组 101KG+/100KG+ 价格。
    """
    result = {}

    for row_idx, row in enumerate(rows):
        if not row:
            continue
        if not any(str(c).strip() == "下单渠道" for c in row if c):
            continue

        # 向上找重量段表头行和数据行
        tier_row = None
        data_row = None
        for hi in range(row_idx - 1, max(0, row_idx - 12), -1):
            prev = rows[hi]
            if not prev:
                continue
            if tier_row is None and any(
                c and ("101KG" in str(c) or "100KG" in str(c) or "51KG" in str(c))
                for c in prev
            ):
                tier_row = prev
            if data_row is None:
                for ci, c in enumerate(prev[:5]):
                    if c and str(c).strip() in ("德国", "欧洲", "法国"):
                        data_row = prev
                        break
                    if c and isinstance(c, str) and ("四大仓" in c or "DTM2" in c or "FBA" in c):
                        data_row = prev
                        break

        if not data_row:
            continue

        for ci, val in enumerate(row):
            if not val or not isinstance(val, str):
                continue
            text = val.strip()
            if text == "下单渠道" or len(text) < 3:
                continue
            if not _looks_like_channel(text):
                continue

            price = _price_at_column_group(data_row, tier_row, ci)
            if price is not None:
                result[text] = price
                for part in re.split(r"[\n①②③④]", text):
                    part = part.strip()
                    if part and len(part) >= 3 and _looks_like_channel(part):
                        result[part] = price

    return result


def _price_at_column_group(data_row, tier_row, col_idx):
    """从数据行中取指定列组最高重量段价格。"""
    if not data_row:
        return None

    target_col = col_idx
    if tier_row:
        # 在 col_idx 附近找 101/100KG+ 列
        best_tier = -1
        for offset in range(4):
            tci = col_idx + offset
            if tci < len(tier_row) and tier_row[tci]:
                t_str = str(tier_row[tci])
                m = re.search(r"(\d+)\s*KG", t_str, re.I)
                tier = int(m.group(1)) if m else 0
                if tier >= 100 or "100" in t_str or "101" in t_str:
                    target_col = tci
                    break
                if tier > best_tier:
                    best_tier = tier
                    target_col = tci

    if target_col < len(data_row):
        return _to_kg_price(data_row[target_col])

    return _extract_row_price(data_row, _find_weight_tier_headers([tier_row] if tier_row else []), 0, [data_row])


def _find_weight_tier_headers(rows):
    """
    扫描所有行，找出含 KG+/CBM+ 的重量段列索引。
    返回 list of {row_idx, col_idx: min_kg}
    """
    tier_map = {}  # col_idx -> min_kg weight

    for row_idx, row in enumerate(rows[:30]):
        if not row:
            continue
        for ci, val in enumerate(row):
            if not val:
                continue
            text = str(val).strip()
            m = re.search(r"(\d+)\s*KG\s*\+", text, re.I)
            if m:
                tier_map[ci] = int(m.group(1))
            elif re.search(r"100\s*KG", text, re.I):
                tier_map[ci] = 100
            elif re.search(r"50\s*KG", text, re.I):
                tier_map[ci] = 50

    return tier_map


def _extract_row_price(row, tier_headers, row_idx, all_rows):
    """
    从数据行提取 KG 单价。
    优先取最高重量段列的价格；否则取合理范围内的最大值。
    """
    # 尝试从表头重量段选最高 tier 列
    if tier_headers:
        best_tier = -1
        best_price = None
        for ci, min_kg in tier_headers.items():
            if ci < len(row):
                p = _to_kg_price(row[ci])
                if p is not None and min_kg >= best_tier:
                    if min_kg > best_tier or (min_kg == best_tier and (best_price is None or p > best_price)):
                        best_tier = min_kg
                        best_price = p
        if best_price is not None:
            return best_price

    # 向上搜索附近表头（10 行内）
    for hi in range(max(0, row_idx - 10), row_idx):
        nearby_tiers = _find_weight_tier_headers([all_rows[hi]] if hi < len(all_rows) else [])
        if nearby_tiers:
            best_tier = -1
            best_price = None
            for ci, min_kg in nearby_tiers.items():
                if ci < len(row):
                    p = _to_kg_price(row[ci])
                    if p is not None and min_kg >= best_tier:
                        best_tier = min_kg
                        best_price = p
            if best_price is not None:
                return best_price

    #  fallback：取行内合理 KG 价的最大值
    prices = []
    for ci, val in enumerate(row):
        if ci == 0:
            continue
        p = _to_kg_price(val)
        if p is not None:
            prices.append(p)

    return max(prices) if prices else None


def _to_kg_price(val):
    """将单元格值转为 KG 单价，过滤 CBM 价等非 KG 价。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if MIN_KG_PRICE <= val <= MAX_KG_PRICE:
            return float(val)
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ("/", "-", "", "None"):
            return None
        try:
            p = float(val)
            if MIN_KG_PRICE <= p <= MAX_KG_PRICE:
                return p
        except ValueError:
            pass
    return None


def _looks_like_channel(text):
    """判断文本是否像渠道名称。"""
    if len(text) < 4:
        return False
    skip = ["备注", "说明", "合计", "小计", "亚马逊", "热门仓", "分区",
            "有效日期", "注意事项", "发货须知", "返回目录", "主营业务",
            "不包税/递延", "包税/不包税", "下单渠道", "清关费", "单证报关",
            "派送方式", "提取时效", "装柜时间", "国家/渠道", "FBA仓点"]
    if any(kw in text for kw in skip):
        return False
    if text.startswith("（") or text.startswith("("):
        return False
    channel_kw = [
        "海派", "海卡", "空派", "铁路", "铁卡", "卡派", "卡航", "海运", "空运",
        "EXX", "Match", "OA", "洛杉矶", "纽约", "欧洲", "英国", "加拿大",
        "包税", "不包税", "递延", "限时", "极速", "统配", "美转加", "铁路",
        "中英", "中欧", "苏新号", "合德", "以星", "COSCO", "ZIM", "CLX",
        "DTM", "WRO", "HAJ", "BHX", "LBA", "德国", "法国", "波兰",
    ]
    return any(kw in text for kw in channel_kw)


def _normalize_channel(name):
    """标准化渠道名：取首行、去空白。"""
    if not name:
        return ""
    first_line = name.split("\n")[0].strip()
    return re.sub(r"\s+", " ", first_line)


# ── 渠道运输类别（海运/空运/铁路/卡航/快递） ──

# 主运输词优先级判定（命中即停）：快递派/卡派 等是派送方式，不影响主运输类别。
# 不用单字「铁/船」以免多模式 sheet（如「空卡铁海」）误判。
_TRANSPORT_MODES = [
    ("海运", ["海运", "海派", "海卡", "普船", "直航", "快航", "美森", "以星",
              "合德", "统配", "EXX", "CLX", "OA", "MAX", "中远", "COSCO", "ZIM"]),
    ("空运", ["空运", "空派", "航空", "五日提", "十日提"]),
    ("铁路", ["铁路", "铁卡", "快铁", "中欧", "中英铁"]),
    ("卡航", ["卡航", "卡派", "专车", "卡车", "陆运", "苏新号"]),
    ("快递", ["快递", "DHL", "FEDEX", "UPS", "DPD", "闪送", "快递派", "派送",
              "快线", "超快线"]),
]

# 单字运输动词（sheet 回退用：仅当 sheet 只有一个明确运输动词才判定）
_SHEET_VERBS = {'空': '空运', '海': '海运', '铁': '铁路', '卡': '卡航',
                '陆': '卡航', '车': '卡航', '船': '海运'}

TRANSPORT_MODES = [m for m, _ in _TRANSPORT_MODES] + ["未知"]


def classify_transport(name, sheet=''):
    """判定渠道运输类别：海运/空运/铁路/卡航/快递/未知。

    渠道名优先（主运输词按优先级，命中即返回）；渠道名无主运输词时回退 sheet 名：
    先按明确主运输词（海运>空运>铁路>卡航>快递），再按单字运输动词
    （仅当 sheet 只有一种动词，避免「空卡铁海」等多模式 sheet 误判）。
    """
    text = (name or "").upper()
    for mode, kws in _TRANSPORT_MODES:
        for kw in kws:
            if kw in text:
                return mode

    sheet_text = (sheet or "").upper()
    if sheet_text:
        for mode, kws in _TRANSPORT_MODES:
            if any(kw in sheet_text for kw in kws):
                return mode
        found = {_SHEET_VERBS[c] for c in sheet_text if c in _SHEET_VERBS}
        if len(found) == 1:
            return found.pop()
    return "未知"


def _is_huanan_channel(channel_name):
    """区域过滤：保留华南/深圳/广州/东莞，或无区域标识的渠道。"""
    region_kw = ["华南", "深圳", "广州", "东莞", "清溪", "宝安", "盐田", "南沙"]
    other_region = ["华东", "上海", "宁波", "义乌", "青岛", "天津", "武汉", "重庆", "厦门"]
    if any(kw in channel_name for kw in region_kw):
        return True
    if any(kw in channel_name for kw in other_region):
        return False
    return True  # 无区域标识默认保留


def _extract_date_from_filename(filepath):
    """从文件名提取日期字符串。"""
    name = Path(filepath).stem
    patterns = [
        r"(\d{4})[年./](\d{1,2})[月./](\d{1,2})",
        r"(\d{4})\.(\d{2})\.(\d{2})",
        r"(\d{4})(\d{2})(\d{2})",
        r"(\d{1,2})[月./](\d{1,2})[日]?",
        r"(\d{1,2})\.(\d{1,2})",
    ]
    for pat in patterns:
        m = re.search(pat, name)
        if m:
            groups = m.groups()
            if len(groups) == 3:
                y, mo, d = groups
                if len(y) == 4:
                    return f"{y}/{int(mo)}/{int(d)}"
            elif len(groups) == 2:
                return f"2026/{int(groups[0])}/{int(groups[1])}"
    return ""


def _extract_date_from_sheet(rows):
    """从 sheet 内容提取更新日期。"""
    for row in rows[:10]:
        if not row:
            continue
        for val in row:
            if isinstance(val, datetime):
                return val.strftime("%Y/%m/%d")
            if val and isinstance(val, str):
                m = re.search(r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})", val)
                if m:
                    return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
                m = re.search(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", val)
                if m:
                    return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return ""


def find_channel_price(channel_name, channels_dict):
    """
    在渠道字典中查找价格：精确 → 标准化 → 模糊。
    返回 (price, matched_channel_name) 或 (None, None)
    """
    if not channel_name or not channels_dict:
        return None, None

    if channel_name in channels_dict:
        return channels_dict[channel_name], channel_name

    norm = _normalize_channel(channel_name)
    if norm in channels_dict:
        return channels_dict[norm], norm

    # 去空格精确匹配
    norm_compact = re.sub(r"\s+", "", norm)
    for name, price in channels_dict.items():
        if re.sub(r"\s+", "", name) == norm_compact:
            return price, name

    # 模糊：渠道名包含关系
    best_score = 0
    best_price = None
    best_name = None
    keywords = _extract_match_keywords(norm)

    for name, price in channels_dict.items():
        score = _match_score(keywords, name)
        if score > best_score:
            best_score = score
            best_price = price
            best_name = name
        elif score == best_score and score > 0 and best_price is not None:
            try:
                if float(price) > float(best_price):
                    best_price = price
                    best_name = name
            except (ValueError, TypeError):
                pass

    if best_score >= 2:
        return best_price, best_name

    return None, None


def _extract_match_keywords(name):
    """提取用于模糊匹配的关键词。"""
    parts = re.split(r"[-\s/（）()\\、，,+]+", name)
    return [p for p in parts if len(p) >= 2]


def _match_score(keywords, target):
    """计算关键词匹配得分。"""
    score = 0
    for kw in keywords:
        if len(kw) >= 2 and kw in target:
            score += 1
    return score
