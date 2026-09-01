"""
FBA 仓点库 + 覆盖渠道价查询模块
============================
- WAREHOUSE_POINTS：各国常用 FBA 仓点静态库（依据天图/每周报价/航乐文件分析整理），
  部署 Railway 无外部依赖也能用。
- extract_warehouse_coverage(fp)：从供应商报价文件提取 仓点码 → {渠道: 单价} 与 邮编关联。
- query_warehouses(data, countries=None, keyword="")：聚合所有已存报价的覆盖渠道价，
  支持多国同时筛选 + 仓点码/邮编关键词搜索。

覆盖渠道数据随供应商文件上传后缓存进 price_supplier.load_prices() 的
`data[sup][country]['warehouses']`；JTT 每周报价的仓点覆盖用
weekly_quotation.match_warehouse 反向匹配。
"""

import os
import re

from openpyxl import load_workbook

from supplier_parser import (
    _skip_sheet, _looks_like_channel, _normalize_channel, _to_kg_price,
    _parse_sheet_rows, MIN_KG_PRICE, MAX_KG_PRICE, MAX_SHEET_ROWS, MAX_SHEET_COLS,
)

# JTT 每周报价伪供应商名（price_supplier 复用，避免重复定义）
WEEKLY_SUPPLIER_NAME = "JTT每周报价"
WEEKLY_COUNTRY = "每周"

# 仓点码：字母开头 + 数字（含 WM-LAX1 带连字符形式）。
# 不用 \b —— Python 把中文当单词字符，\b 在「国DTM2」间不成立，用字母/数字前后瞻替代。
FBA_CODE_RE = re.compile(r"(?<![A-Z0-9])((?:[A-Z]{2,4}-)?[A-Z]{2,4}\d{1,2})(?![A-Z0-9])")


# ── 各国常用仓点库（区域分组） ──

WAREHOUSE_POINTS = {
    "美国": {
        "西部": ["ONT8", "LGB8", "LAX9", "SBD1", "POC1", "POC2", "POC3",
                "IUSP", "IUSJ", "IUSQ", "IUTI", "WM-LAX1", "WM-LAX2",
                "GYR1", "PHX3", "PHX7", "LAS1", "SCK1", "SMF1"],
        "中部": ["FTW1", "DFW2", "HOU2", "MEM1", "STL4"],
        "东部": ["IND9", "CVG1", "CMH1", "BNA1", "MDW2", "ORD2", "RDU1",
                "CLT2", "MCO2", "TPA1", "JAX2", "PHL1", "ACY1", "BWI1",
                "ATL1", "MQJ1", "RIC2", "CHA1", "CHO1"],
    },
    "加拿大": {
        "多伦多": ["YYZ1", "YYZ2", "YYZ3", "YYZ4", "YYZ5", "YYZ6", "YYZ7", "YYZ8", "YYZ9"],
        "渥太华": ["YOW1", "YOW3"],
        "汉密尔顿": ["YHM1"],
        "温哥华": ["YVR1", "YVR2", "YVR3", "YVR4", "YVR5", "YVR6", "YVR7", "YVR8", "YVR9",
                  "YXX1", "YXX2"],
        "卡尔加里": ["YYC1", "YYC4", "YYC6"],
        "埃德蒙顿": ["YEG1", "YEG2"],
        "温尼伯": ["YWG1"],
    },
    "英国": {
        "英国": ["BHX4", "LBA4", "LBA5", "MAN1", "DSA1", "EMA1", "NCL1",
                "BRS1", "PIK1", "CWL1", "STN1"],
    },
    "欧洲": {
        "德国": ["DTM2", "HAJ1", "HAJ2", "KSF1", "LEJ1", "MUC3", "BER3",
                "DUS2", "XDU1", "XDP1", "XDX1"],
        "波兰": ["WRO5", "POZ1", "WAW1", "WAW2"],
        "法国": ["LYS1", "CDG1", "ORY1"],
    },
}


# ── 每周报价表模块（懒加载，拣货数据目录不在 sys.path 默认里） ──

def _load_weekly_module():
    import sys
    try:
        import weekly_quotation
        return weekly_quotation
    except ImportError:
        d = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '拣货数据')
        if d not in sys.path:
            sys.path.insert(0, d)
        import weekly_quotation
        return weekly_quotation


def _weekly_covers(pattern, code):
    """JTT 每周报价仓点模式是否覆盖给定仓点码（严格版，不做 FBA 兜底）。

    只认显式仓点码/范围（ONT8、YYZ1-9、DTM2-44145）与国家模式（德国/波兰），
    邮编前缀（8,9 / 97.98.99开头）与 FBA 泛兜底无法映射到具体仓点码 → 不算覆盖。
    """
    p = (pattern or '').upper()
    code = (code or '').upper()
    if not p or not code:
        return False
    wq = _load_weekly_module()
    if any(name in p for name in wq.COUNTRY_NAMES):
        return wq.match_warehouse(pattern, code)
    expanded = set()
    for tok in wq._split_tokens(p):
        expanded |= wq._expand_token(tok)
    return code in expanded


def _pattern_postals(pattern):
    """从仓点模式提取 5 位邮编：'DTM2-44145' → {'44145'}、'CA ONTARIO 91761' → {'91761'}。

    邮编只取独立的 5 位数字段（代码/范围/前缀模式如 YYZ1-9、97.98.99开头 不产生邮编）。
    """
    p = (pattern or '').upper()
    return set(re.findall(r'(?<!\d)(\d{5})(?!\d)', p))


def _digits_of(pattern):
    """提取模式中的数字片段（邮编前缀）：'97.98.99开头' → ['97','98','99']；'8, 9' → ['8','9']。"""
    return re.findall(r'\d{1,3}', str(pattern or '').upper())


def _is_zone_pattern(pattern):
    """是否为「邮编/区域段」模式（无具体仓点码，如 美国东岸(邮编0-3)、FBA（8/9邮编）、
    '97.98.99开头'、'CA ONTARIO 91761'、'3, 2, 1, 0'）。纯仓点码模式（LAX9/DTM2-44145）不算。"""
    p = (pattern or '').upper()
    if '邮编' in p or '开头' in p or 'FBA' in p:
        return True
    if re.fullmatch(r'[\d,\s]+', p):   # 纯邮编前缀列表："8, 9" / "3, 2, 1, 0"
        return True
    if _pattern_postals(p):            # 地名-邮编："NEW BRUNSWICK-08901" / "CA ONTARIO 91761"
        return True
    return False


def _kw_hits_zone(kw, pattern):
    """邮编/区域段模式是否命中关键词（邮编段按前缀匹配，避免 '91761' 误命中 '邮编4-7'）。"""
    p = (pattern or '').upper()
    if kw in p:
        return True
    for pz in _pattern_postals(p):
        if kw in pz or pz in kw:
            return True
    for pr in _digits_of(p):
        if kw.startswith(pr) or pr.startswith(kw):
            return True
    return False


def _cell_code_postals(v):
    """从「CODE-邮编-国家」单元格提取 (code, postal) 对，支持多行（\n 分隔）。

    每行独立匹配，避免多仓点单元格（如 'WRO5-06126-DE\\nWRO5-59225-PL'）串号。
    """
    out = []
    for line in str(v or '').split('\n'):
        line = line.strip().upper()
        m = re.match(r'([A-Z]{2,4}\d{1,2})-(\d{5})', line)
        if m:
            out.append((m.group(1), m.group(2)))
    return out


# ── 覆盖渠道提取 ──

def _fba_code_cell(v):
    """单元格是否为纯仓点码（如 YYZ1 / WM-LAX1）。"""
    if isinstance(v, str):
        s = v.strip().upper()
        if FBA_CODE_RE.fullmatch(s):
            return s
    return None


def _set_cov(cov, code, ch, price):
    m = cov.setdefault(code, {})
    if ch not in m or (price is not None and (m[ch] is None or price > m[ch])):
        m[ch] = price


def _clean_channel_label(v):
    """「仓库代码」表头旁的渠道名：取首行、去尾部标点（如「洛杉矶普船28日达直送-」）。"""
    s = str(v or '').split('\n')[0].strip()
    return re.sub(r'[-（(、，,：: ·]+$', '', s)


def _parse_price_cell(v):
    """解析 KG 单价，兼容「6.5/KG」「￥6.5」等格式。"""
    p = _to_kg_price(v)
    if p is not None:
        return p
    if isinstance(v, str):
        m = re.search(r'(\d+(?:\.\d+)?)', v.replace(',', ''))
        if m:
            p = float(m.group(1))
            if MIN_KG_PRICE <= p <= MAX_KG_PRICE:
                return p
    return None


def _sheet_base_name(sn):
    """sheet 名去仓点码/标点 → 语境名（如「德国DTM2.WRO5.HAJ1前置仓」→「德国前置仓」）。"""
    s = FBA_CODE_RE.sub('', sn)
    return re.sub(r'[\s\-_./·]+', '', s) or sn


# 目的仓行渠道：短运输词（海运/卡航/铁路…）
_TRANSPORT_LABELS = {"海运", "空运", "卡航", "铁路", "快递", "专车", "陆运", "海派", "空派", "快递派"}


def extract_warehouse_coverage(fp):
    """解析单个供应商报价文件 → (覆盖渠道, 邮编关联)。

    cov = {仓点码: {渠道名: 单价}}；postals = {仓点码: {邮编, ...}}。
    只用两类可靠结构（避免查询参考表跨国家仓点码污染）：
      A. 「仓库代码」列（天图加拿大/美西/美转加式）：某行某列标题为「仓库代码」，
         其下行为仓点码（YYZ1/ONT8/LGB8/…），渠道名在该行右侧，价格为行内首个
         KG 单价（华南列）。
      B. sheet 名含仓点码（航乐「德国DTM2.WRO5.HAJ1前置仓」式）：该 sheet 解析出的
         渠道覆盖这些仓点（用渠道解析价）；目的仓行（col0=「CODE-邮编-国家」）另
         提取 邮编→仓点 关联。
    """
    cov = {}
    postals = {}
    try:
        wb = load_workbook(fp, data_only=True, read_only=True)
    except Exception:
        return cov, postals
    try:
        for sn in wb.sheetnames:
            if _skip_sheet(sn):
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True, max_row=MAX_SHEET_ROWS, max_col=MAX_SHEET_COLS))
            if not rows:
                continue

            # A. 「仓库代码」列 → 渠道行 × 仓点行取价
            hdr_idx = hdr_col = None
            for i, row in enumerate(rows[:40]):
                for ci, v in enumerate(row[:8]):
                    if isinstance(v, str) and '仓库代码' in v:
                        hdr_idx, hdr_col = i, ci
                        break
                if hdr_idx is not None:
                    break
            if hdr_idx is not None:
                ch_label = ''
                if hdr_col + 1 < len(rows[hdr_idx]):
                    ch_label = _clean_channel_label(rows[hdr_idx][hdr_col + 1])
                if not ch_label and hdr_idx + 1 < len(rows) and hdr_col + 1 < len(rows[hdr_idx + 1]):
                    ch_label = _clean_channel_label(rows[hdr_idx + 1][hdr_col + 1])
                for row in rows[hdr_idx + 1:]:
                    if not row or hdr_col >= len(row):
                        continue
                    code = _fba_code_cell(row[hdr_col])
                    if not code:
                        continue
                    price = None
                    for pv in row[hdr_col + 1:]:
                        p = _to_kg_price(pv)
                        if p is not None:
                            price = p
                            break
                    if ch_label and price is not None:
                        _set_cov(cov, code, ch_label, price)

            # B. sheet 名含仓点码 → 该 sheet 渠道覆盖（不覆盖 A 轮专属单价）
            name_codes = set(FBA_CODE_RE.findall(sn.upper()))
            if name_codes:
                sheet_channels = _parse_sheet_rows(rows, sn) or {}
                if sheet_channels:
                    for code in name_codes:
                        for ch, price in sheet_channels.items():
                            if ch in cov.get(code, {}):
                                continue
                            if price is not None:
                                _set_cov(cov, code, ch, price)
                else:
                    # B2. 目的仓行（航乐前置仓式）：col0=「CODE-邮编-国家」、col1=运输词、col2=KG价
                    base = _sheet_base_name(sn)
                    for row in rows:
                        if not row or not row[0] or not row[1]:
                            continue
                        cell = str(row[0]).upper()
                        m = FBA_CODE_RE.search(cell)
                        if not m:
                            continue
                        label = str(row[1]).strip()
                        if label not in _TRANSPORT_LABELS:
                            continue
                        price = _parse_price_cell(row[2] if len(row) > 2 else None)
                        if price is not None:
                            _set_cov(cov, m.group(1), f"{base}-{label}", price)
                        # 邮编关联：每行「CODE-邮编」对（跨行独立，避免串号）
                        for ccode, cpostal in _cell_code_postals(cell):
                            postals.setdefault(ccode, set()).add(cpostal)
    finally:
        wb.close()
    return cov, postals


# ── 仓点查询 ──

def query_warehouses(data, countries=None, keyword=""):
    """聚合所有已存报价 → {grid: {...}, zones: [...], weekly_present: bool}。

    data 来自 price_supplier.load_prices()：
      {supplier: {country: {update_date, channels, warehouses?, wh_postals?, wh_entries?}}}
    warehouses 为 extract_warehouse_coverage 的产物（覆盖渠道），wh_postals 为其邮编关联；
    wh_entries 为 JTT 每周报价原始条目（其 wh_pattern 再反解邮编）。
    countries：可选国家列表（空/None = 全部），多国同时筛选；
    keyword：仓点码或邮编关键词（子串/前缀匹配，跨国家同时过滤）。
    grid 只展示 WAREHOUSE_POINTS 内置仓点，覆盖渠道跨供应商合并去重；
    zones 为命中关键词的「邮编/区域段」渠道（美国邮编前缀段等无具体仓点码的模式）。
    """
    if isinstance(countries, str):
        countries = [c for c in countries.split(',') if c.strip()]
    sel_countries = set(c for c in (countries or []) if c)
    kw = (keyword or "").strip().upper()

    coverage = {}       # code -> {(supplier, channel): price}
    point_postals = {}  # code -> {postal, ...}
    weekly_entries = []
    for sup, sup_countries in (data or {}).items():
        for cn, p in sup_countries.items():
            for code, chmap in (p.get("warehouses") or {}).items():
                for ch, price in chmap.items():
                    d = coverage.setdefault(code, {})
                    key = (sup, ch)
                    if key not in d or (price is not None and (d[key] is None or price > d[key])):
                        d[key] = price
            for code, ps in (p.get("wh_postals") or {}).items():
                point_postals.setdefault(code, set()).update(ps)
            if p.get("wh_entries"):
                weekly_entries.extend(p["wh_entries"])

    pick_price = None
    if weekly_entries:
        wq = _load_weekly_module()
        pick_price = wq.pick_tier_price

    def _wh_match(code, pl):
        """仓点码或邮编子串匹配（'DTM2-44145' 式整体输入也命中）。"""
        if kw in code.upper():
            return True
        return any(kw in pz or pz in kw for pz in pl)

    out = {}
    for cn, regions in WAREHOUSE_POINTS.items():
        if sel_countries and cn not in sel_countries:
            continue
        out[cn] = {}
        for region, codes in regions.items():
            pts = []
            for code in codes:
                channels = []
                for (sup, ch), price in coverage.get(code, {}).items():
                    channels.append({"supplier": sup, "name": ch, "price": price})
                postals = set(point_postals.get(code, ()))
                if pick_price:
                    for e in weekly_entries:
                        if not _weekly_covers(e.get("wh_pattern"), code):
                            continue
                        p = pick_price(e.get("tiers") or {}, 10 ** 6)
                        pv = round(p, 2) if p is not None else None
                        channels.append({
                            "supplier": WEEKLY_SUPPLIER_NAME,
                            "name": e.get("channel_raw") or e.get("channel"),
                            "price": pv,
                        })
                        postals |= _pattern_postals(e.get("wh_pattern"))
                ded = {}
                for c in channels:
                    k = (c["supplier"], c["name"])
                    if k not in ded or (c["price"] is not None
                                        and (ded[k]["price"] is None or c["price"] > ded[k]["price"])):
                        ded[k] = c
                channels = sorted(ded.values(),
                                  key=lambda c: -(c["price"] if c["price"] is not None else 0))
                pl = sorted(postals)
                if kw and not _wh_match(code, pl):
                    continue
                pts.append({"code": code, "region": region, "postals": pl, "channels": channels})
            if pts:
                out[cn][region] = pts

    # ── 邮编/区域直查（美国邮编前缀段等无具体仓点码的模式） ──
    # 仅当关键词未命中任何内置仓点（避免 44145 这类全球不唯一邮编同时带出美国中岸渠道）
    # 且命中关键词时展示；能映射到内置仓点的模式由网格展示，不重复。
    postal_zones = []
    if kw and pick_price and not any(out[cn] for cn in out):
        static_codes = {code for regions in WAREHOUSE_POINTS.values()
                        for codes in regions.values() for code in codes}
        seen = set()
        for e in weekly_entries:
            pat = str(e.get("wh_pattern") or '').upper()
            if not pat or not _is_zone_pattern(pat):
                continue
            if any(_weekly_covers(pat, code) for code in static_codes):
                continue  # 模式有具体仓点码 → 网格已展示，避免重复
            if not _kw_hits_zone(kw, pat):
                continue
            price = pick_price(e.get("tiers") or {}, 10 ** 6)
            pv = round(price, 2) if price is not None else None
            key = (pat, e.get("section", ""), e.get("channel_raw") or e.get("channel"), pv)
            if key in seen:
                continue
            seen.add(key)
            postal_zones.append({
                "pattern": pat,
                "section": e.get("section", ""),
                "channel": e.get("channel_raw") or e.get("channel"),
                "price": pv,
            })
            if len(postal_zones) >= 400:
                break

    return {"grid": out, "zones": postal_zones, "weekly_present": bool(weekly_entries)}
