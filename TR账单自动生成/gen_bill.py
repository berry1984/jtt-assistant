#!/usr/bin/env python3
"""
番茄钟账单生成器
用途：根据 内部拣货数据参考值（拣货数据模块生成） Excel文件，生成标准格式账单（不再依赖订单列表）。
发货日期直接抓取参考值「下单时间」列（A列），日期格式保持账单模板不变（如 8月19日）。

用法：
  python3 gen_bill.py <内部拣货数据参考值.xlsx> [输出文件名.xlsx]

规则：
  - 单价(K列) = 参考值「应收单价」列（按 SO+FBA 行对应，回退仓库代码）
  - J列(计费重) = 复刻参考值模版「计费重」公式 ROUND(MAX(参考实重×箱数, 参考材积重×箱数))
  - 特殊服务(美国快递-*包税 等)：按参考值计费重直接取
  - S列(报关费) = 350/1.06, 报关费与税额按每一行填写（不区分报关组）
  - 保留完整浮点精度（使用公式）
  - 排序：按渠道分组 → FBA列(箱号)含FBA字眼的运单在前 → SO → FBA
"""

import sys, re, os, shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

TEMPLATE = None  # Will look for 账单模板.xlsx next to script or in cwd

# ── Column mapping ──
COL = {chr(65+i): i+1 for i in range(26)}  # A=1, B=2, ...
COL.update({'AA': 27, 'AB': 28, 'AC': 29})

# 特殊计费重服务：按拣货数据 单箱 max(材积重, 实际重量) 向上取整后按 FBA 合计
VOLUMETRIC_SERVICES = {
    '美国快递-DHL包税',
    '美国快递-Fedex包税-HKIP',
    '美国快递-Fedex包税-HKIE',
    '美国快递-UPS包税-HK红单',
}

def parse_order_date(v):
    """把订单日期解析为 datetime，兼容 datetime / Excel 数字序列号 / 常见字符串格式（含尾部空白）。
    不强调日期格式：如 '2026-08-21 10:23:45\t'、'2026/08/21'、'2026.08.21'、'8月21日' 均可。
    解析不了返回 None。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(v))
        except (ValueError, TypeError):
            return None
    s = str(v).strip()
    if not s:
        return None
    # 完整日期（可能带时间部分），如 2026-08-13 17:12:17 / 2026/08/13 / 2026.08.13
    m = re.match(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # 无年份，如 8月21日（年份取当前年）
    m = re.match(r'^(\d{1,2})月(\d{1,2})日', s)
    if m:
        try:
            return datetime(datetime.now().year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def find_template():
    """Find template file"""
    candidates = [
        os.path.join(os.path.dirname(__file__), '账单模板.xlsx'),
        os.path.join(os.getcwd(), '账单模板.xlsx'),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

def _norm_header(h):
    """表头归一化：去空白/换行，转大写。"""
    return re.sub(r'\s+', '', str(h or '').upper())


def _find_col(headers, needles, exclude=()):
    """按归一化表头找列：needles 子串命中即返回，exclude 中的列号跳过。
    调用方需保证更精确的列（如「参考实重」）先匹配并纳入 exclude。"""
    for c, h in headers.items():
        if c in exclude:
            continue
        nh = _norm_header(h)
        for n in needles:
            if _norm_header(n) in nh:
                return c
    return None


def _parse_reference(ref_path):
    """解析「内部拣货数据参考值」模版（拣货数据模块生成）。

    列布局（行1为表头，A 为新增「下单时间」列）：
      A=下单时间  B=系统SO号  C=客户渠道  D=国家  E=仓库代码  F=应收单价  G=应付单价
      K=FBA ID  N=总箱数  O=实重  P=长  Q=宽  R=高
      W=参考实重  X=参考长  Y=参考宽  Z=参考高  H=供应商渠道
    下单时间/SO/渠道/国家/仓库代码 只在分组首行填写（跨行合并），需向下填充。

    返回:
      ref_rows: {SO: [{order_time, fba, wh, channel, country, supplier_ch, e_price,
                       f_price, boxes, weight(N), length, width, height,
                       ref_w, ref_l, ref_wid, ref_h}, ...]}
      warehouse_prices: {仓库代码: 应收单价}（首见为准，供回退与报价表A）
      price_rows_raw:   [(客户渠道, 仓库代码, 应收单价)]（按仓库去重，供报价表A）
    """
    wb = load_workbook(ref_path, data_only=True)
    ws = wb.active
    headers = {c.value: c.column for c in ws[1]}
    # 翻转成 {列号: 表头} 便于顺序匹配
    headers_by_col = {}
    for h, c in headers.items():
        headers_by_col[c] = h

    # 先匹配最精确/带「参考」前缀的列，并加入 exclude 防「实重/长/宽/高」误配到参考列
    refw_c    = _find_col(headers_by_col, ['参考实重'])
    refl_c    = _find_col(headers_by_col, ['参考长'])
    refwid_c  = _find_col(headers_by_col, ['参考宽'])
    refh_c    = _find_col(headers_by_col, ['参考高'])
    _excl = set(x for x in [refw_c, refl_c, refwid_c, refh_c] if x)
    totw_c    = _find_col(headers_by_col, ['总实重'], _excl)   # 防「实重」命中「总实重」
    if totw_c:
        _excl.add(totw_c)

    so_c        = _find_col(headers_by_col, ['系统SO号', 'SO号', 'SO'])
    order_time_c = _find_col(headers_by_col, ['下单时间', '下单日期'])
    ch_c        = _find_col(headers_by_col, ['客户渠道'])
    country_c = _find_col(headers_by_col, ['国家'])
    wh_c      = _find_col(headers_by_col, ['仓库代码'])
    e_c       = _find_col(headers_by_col, ['应收单价', '应收'])
    f_c       = _find_col(headers_by_col, ['应付单价', '应付'])
    sup_c     = _find_col(headers_by_col, ['供应商渠道'])
    fba_c     = _find_col(headers_by_col, ['FBA'])
    boxes_c   = _find_col(headers_by_col, ['总箱数', '箱数', 'CTN'])
    w_c       = _find_col(headers_by_col, ['实重'], _excl)
    len_c     = _find_col(headers_by_col, ['长'], _excl)
    wid_c     = _find_col(headers_by_col, ['宽'], _excl)
    hei_c     = _find_col(headers_by_col, ['高'], _excl)

    if so_c is None or e_c is None or fba_c is None:
        raise ValueError('参考值模版缺少必需列（系统SO号 / 应收单价 / FBA ID）')

    def _s(v):
        return '' if v is None else str(v).strip()

    ref_rows = {}
    warehouse_prices = {}
    price_rows_raw = []
    cur_so = cur_ch = cur_country = cur_wh = ''
    cur_order_time = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if _s(r[so_c - 1]):
            cur_so = _s(r[so_c - 1])
            cur_ch = _s(r[ch_c - 1]) if ch_c else ''
            cur_country = _s(r[country_c - 1]) if country_c else ''
            cur_wh = _s(r[wh_c - 1]) if wh_c else ''
            cur_order_time = r[order_time_c - 1] if order_time_c else None
        if not cur_so:
            continue
        fba = r[fba_c - 1]
        if fba in (None, ''):
            continue
        e_price = r[e_c - 1]
        if cur_wh and cur_wh not in warehouse_prices:
            warehouse_prices[cur_wh] = e_price
            price_rows_raw.append((cur_ch, cur_wh, e_price))
        ref_rows.setdefault(cur_so, []).append({
            'so': cur_so,
            'order_time': cur_order_time,
            'fba': _s(fba),
            'wh': cur_wh,
            'channel': cur_ch,
            'country': cur_country,
            'supplier_ch': _s(r[sup_c - 1]) if sup_c else '',
            'e_price': e_price,
            'f_price': r[f_c - 1] if f_c else '',
            'boxes': r[boxes_c - 1] if boxes_c else None,
            'weight': r[w_c - 1] if w_c else None,
            'length': r[len_c - 1] if len_c else None,
            'width': r[wid_c - 1] if wid_c else None,
            'height': r[hei_c - 1] if hei_c else None,
            'ref_w': r[refw_c - 1] if refw_c else None,
            'ref_l': r[refl_c - 1] if refl_c else None,
            'ref_wid': r[refwid_c - 1] if refwid_c else None,
            'ref_h': r[refh_c - 1] if refh_c else None,
        })
    return ref_rows, warehouse_prices, price_rows_raw


def load_data(ref_path):
    """Load 内部拣货数据参考值（TR账单已不依赖订单列表）, return
    (ref_rows, warehouse_prices, price_rows_raw, declaration_groups)"""
    ref_rows, warehouse_prices, price_rows_raw = _parse_reference(ref_path)
    # ── 报关分组规则：在生成账单时按账单内每行的 (走货渠道, SO) 直接判定（见 generate_bill） ──
    declaration_groups = []
    return ref_rows, warehouse_prices, price_rows_raw, declaration_groups

def lookup_price(prices, wh_code):
    """Look up price by warehouse code, with suffix matching"""
    if not wh_code:
        return {}
    p = prices.get(wh_code, {})
    if not p:
        prefix = wh_code.split('-')[0]
        p = prices.get(prefix, {})
    if not p:
        for key in prices:
            if wh_code.startswith(key):
                p = prices[key]
                break
    return p

def compute_ref_weight(row, channel):
    """复刻参考值模版「计费重」列公式：ROUND(MAX(参考实重×箱数, 参考材积重×箱数), 0)。

    参考实重/参考材积重（U/V/W/X 列，来自箱规历史）缺失时回退实际尺寸（N/O/P/Q）。
    「快递派」渠道按模版公式单箱最低计费 12kg：MAX(总参考材积重, 箱数×12)。
    """
    def num(v):
        try:
            return float(v) if v not in (None, '') else 0
        except (ValueError, TypeError):
            return 0
    M = num(row.get('boxes'))
    if M <= 0:
        return 0
    ref_actual = num(row.get('ref_w'))
    ref_l, ref_w, ref_h = num(row.get('ref_l')), num(row.get('ref_wid')), num(row.get('ref_h'))
    ref_vol = ref_l * ref_w * ref_h / 6000.0 if (ref_l and ref_w and ref_h) else 0
    if ref_actual or ref_vol:
        total_actual = ref_actual * M
        total_vol = ref_vol * M
        if '快递派' in (channel or ''):
            return round(max(total_vol, M * 12))
        return round(max(total_actual, total_vol))
    # 无历史匹配：回退实际尺寸
    actual = num(row.get('weight'))
    vol = num(row.get('length')) * num(row.get('width')) * num(row.get('height')) / 6000.0
    return round(max(actual * M, vol * M))


def build_rows(ref_rows, warehouse_prices):
    """Build bill rows from 内部拣货数据参考值（不再依赖订单列表）。

    单价取参考值「应收单价」；重量复刻「计费重」公式（参考值 A 列「下单时间」为发货日期来源，
    账单日期格式不变）；每行全部来自参考值，无订单回退分支。
    """

    def to_num(v, default=0):
        """Safely convert Excel value to float"""
        if v is None:
            return float(default)
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).strip())
        except (ValueError, TypeError):
            return float(default)

    def lookup_wh_price(wh_code):
        """仓库代码 → 应收单价（支持截断前缀/前缀匹配回退）。"""
        wh_code = wh_code or ''
        if wh_code in warehouse_prices and warehouse_prices[wh_code] is not None:
            return warehouse_prices[wh_code]
        prefix = wh_code.split('-')[0]
        if prefix in warehouse_prices and warehouse_prices[prefix] is not None:
            return warehouse_prices[prefix]
        for key in warehouse_prices:
            if wh_code.startswith(key):
                return warehouse_prices[key]
        return 0

    all_rows = []

    for so in sorted(ref_rows.keys()):
        ref_list = ref_rows[so]
        # 渠道取参考值「客户渠道」列（该 SO 分组内一致）
        channel = (ref_list[0].get('channel') or '').strip()
        # 发货日期 = 参考值「下单时间」列（8月19日格式不变，解析不了留空）
        date_val = parse_order_date(ref_list[0].get('order_time')) or ''
        for ref in ref_list:
            wh_code_ref = ref['wh'] or ''
            # 应收单价缺失时回退仓库匹配单价（参考值中 E 空 = 报价单无该仓库）
            e_price = ref['e_price']
            if e_price in (None, ''):
                e_price = lookup_wh_price(wh_code_ref)
            ref_weight = compute_ref_weight(ref, channel)
            all_rows.append({
                'so': so,
                'fba': ref['fba'],
                'date': date_val,
                'service': channel,
                'wh': wh_code_ref,
                'boxes': int(to_num(ref['boxes'])),
                'length': to_num(ref['length']),
                'width': to_num(ref['width']),
                'height': to_num(ref['height']),
                'weight': round(ref_weight),
                'weight_raw': ref_weight,
                'unit_price': to_num(e_price),
            })

    # Filter out rows with zero weight (no actual cargo)
    all_rows = [r for r in all_rows if r['weight'] > 0]

    return all_rows

def sort_rows(rows, declaration_groups=None):
    """Sort rows: by channel group; within a channel, rows whose FBA column (箱号) contains
    'FBA' come first, then rows without; each group ordered by SO then FBA"""
    # Determine channel group order (走货渠道 appearance order)
    seen = []
    for r in rows:
        ch = r['service']
        if ch not in seen:
            seen.append(ch)

    def sort_key(r):
        ch_idx = seen.index(r['service']) if r['service'] in seen else 999
        # FBA列（箱号）含 "FBA" 字眼的运单靠前，无则排后
        has_fba = 0 if 'FBA' in str(r['fba']) else 1
        return (ch_idx, has_fba, r['so'], r['fba'])

    return sorted(rows, key=sort_key)

def generate_bill(rows, output_path, template_path=None, title_str=None, date_range_str=None, price_rows_raw=None, year=None, declaration_groups=None):
    """Generate bill Excel file"""
    
    if template_path is None:
        template_path = find_template()
    if not template_path:
        print("ERROR: 账单模板.xlsx not found!")
        sys.exit(1)
    
    # Copy template
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)

    # Determine month from date_range_str (dynamic, not hardcoded "5月")
    month_nums = [int(x) for x in re.findall(r'\d+', date_range_str)] if date_range_str else []
    bill_month = month_nums[0] if month_nums else datetime.now().month
    sheet_name_old = '5月人民币账单（已调格式）'
    sheet_name_new = f'{bill_month}月人民币账单（已调格式）'

    # Rename sheet if month differs from template
    if sheet_name_old in wb.sheetnames:
        if sheet_name_old != sheet_name_new:
            wb[sheet_name_old].title = sheet_name_new
    ws = wb[sheet_name_new]

    # Capture template column fills from row 4 (first data row) for style preservation
    from openpyxl.styles import PatternFill
    template_fills = {}
    for c in range(1, 29):
        src = ws.cell(row=4, column=c)
        f = src.fill
        if f.patternType and f.patternType != 'none':
            try:
                template_fills[c] = PatternFill(patternType=f.patternType,
                    fgColor=src.fill.fgColor.rgb if f.fgColor else None)
            except:
                template_fills[c] = PatternFill(patternType='solid', fgColor='FFFFFFFF')

    # ── Unmerge old cells in data area ──
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 4:
            ws.unmerge_cells(str(mr))
    for row in range(4, 30):
        for col in range(1, 29):
            try:
                ws.cell(row=row, column=col).value = None
            except AttributeError:
                pass

    # ── Styles ──
    thin_border = Border(
        left=Side(style='hair'), right=Side(style='hair'),
        top=Side(style='hair'), bottom=Side(style='hair'))
    data_font = Font(name='微软雅黑', size=9)
    bold_font = Font(name='微软雅黑', size=9, bold=True)
    center = Alignment(horizontal='center', vertical='center')
    
    n = len(rows)
    
    # ── Write data rows ──
    for i, r in enumerate(rows):
        row_num = 4 + i
        ws.row_dimensions[row_num].height = 20
        
        # Format date: datetime / serial → "8月19日"
        if isinstance(r['date'], datetime):
            date_str = f"{r['date'].month}月{r['date'].day}日"
        elif isinstance(r['date'], (int, float)):
            dt = datetime(1899, 12, 30) + timedelta(days=r['date'])
            date_str = f"{dt.month}月{dt.day}日"
        else:
            date_str = str(r['date'])
        vals = [
            ('A', date_str, '@'),
            ('B', r['so'], '@'),
            ('C', r['fba'], '@'),
            ('D', r['service'], '@'),
            ('E', r['wh'], '@'),
            ('F', r['boxes'], '0'),
            ('G', r['length'] if r['length'] else '', '0'),
            ('H', r['width'] if r['width'] else '', '0'),
            ('I', r['height'] if r['height'] else '', '0'),
            ('J', r['weight'], '0'),
            ('K', r['unit_price'], '0.00'),
        ]
        
        for col_l, val, nf in vals:
            cell = ws[f'{col_l}{row_num}']
            cell.value = val
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            if nf != '@':
                cell.number_format = nf
            # Fill applied in post-processing below
        
        # L = K*0.07/1.06, M = K*0.35, N = K*0.58 (formulas)
        for col_l, formula in [('L', f'=K{row_num}*0.07/1.06'),
                                ('M', f'=K{row_num}*0.35'),
                                ('N', f'=K{row_num}*0.58')]:
            cell = ws[f'{col_l}{row_num}']
            cell.value = formula
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            cell.number_format = '0.00'
        
        # O = L*J, P = O*0.06, Q = M*J, R = N*J (formulas)
        for col_l, formula in [('O', f'=L{row_num}*J{row_num}'),
                                ('P', f'=O{row_num}*0.06'),
                                ('Q', f'=M{row_num}*J{row_num}'),
                                ('R', f'=N{row_num}*J{row_num}')]:
            cell = ws[f'{col_l}{row_num}']
            cell.value = formula
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            cell.number_format = '#,##0.00'
        
        # S, T (customs fee - filled per channel group later)
        # U-Y = 0
        for col_l in ['U', 'V', 'W', 'X', 'Y']:
            cell = ws[f'{col_l}{row_num}']
            cell.value = 0
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            cell.number_format = '#,##0.00'
        
        # Z = RMB
        ws[f'Z{row_num}'].value = 'RMB'
        ws[f'Z{row_num}'].font = data_font
        ws[f'Z{row_num}'].alignment = center
        ws[f'Z{row_num}'].border = thin_border
        
        # AA = SUM(O:Y) (formula)
        ws[f'AA{row_num}'].value = f'=SUM(O{row_num}:Y{row_num})'
        ws[f'AA{row_num}'].font = data_font
        ws[f'AA{row_num}'].alignment = center
        ws[f'AA{row_num}'].border = thin_border
        ws[f'AA{row_num}'].number_format = '#,##0.00'
        
        # AB blank
        ws[f'AB{row_num}'].border = thin_border

    # Apply template column fills to all data rows
    for rn in range(4, 4 + n):
        # A column border
        ws[f'A{row_num}'].border = thin_border
    
    # ── Customs fee: fill every data row ──
    # 规则：不再区分报关组，报关费(S)=350/1.06 与 报关税费(T)=S*0.06 直接填写到每一行。
    for i, r in enumerate(rows):
        rn = 4 + i
        for cl in ['S', 'T']:
            cell = ws[f'{cl}{rn}']
            cell.border = thin_border

        ws[f'S{rn}'].value = '=350/1.06'
        ws[f'S{rn}'].font = data_font
        ws[f'S{rn}'].alignment = center
        ws[f'S{rn}'].number_format = '#,##0.00'
        if 19 in template_fills:
            ws[f'S{rn}'].fill = template_fills[19]

        ws[f'T{rn}'].value = f'=S{rn}*0.06'
        ws[f'T{rn}'].font = data_font
        ws[f'T{rn}'].alignment = center
        ws[f'T{rn}'].number_format = '#,##0.00'
        if 20 in template_fills:
            ws[f'T{rn}'].fill = template_fills[20]

    # Apply template column fills to ALL data rows (after customs section)
    # ── Summary row (with blank separator before it, like correct bill) ──
    sr = 4 + n + 1  # blank row at 4+n, 合计 at sr
    info_row = sr + 1

    ws.merge_cells(f'B{sr}:E{sr}')
    ws[f'B{sr}'].value = '合计'
    ws[f'B{sr}'].font = bold_font
    ws[f'B{sr}'].alignment = center

    sum_cols = ['F', 'J', 'O', 'P', 'Q', 'R', 'S', 'T', 'W', 'AA']
    for cl in sum_cols:
        ws[f'{cl}{sr}'].value = f'=SUM({cl}4:{cl}{sr-2})'  # data up to last data row (sr-2), blank sr-1 excluded
        ws[f'{cl}{sr}'].font = bold_font
        ws[f'{cl}{sr}'].alignment = center
        ws[f'{cl}{sr}'].border = thin_border
        ws[f'{cl}{sr}'].number_format = '#,##0.00'

    for cl in ['A','B','C','D','E','G','H','I','K','L','M','N','U','V','X','Y','Z','AB']:
        ws[f'{cl}{sr}'].border = thin_border
        ws[f'{cl}{sr}'].font = bold_font
        ws[f'{cl}{sr}'].alignment = center
    
    # ── Bank info ──
    ws.merge_cells(f'A{info_row}:AA{info_row}')
    ws[f'A{info_row}'].value = '以下账户信息为我司唯一合法收款账户，如付款至其他账户我司概不负责!'
    ws[f'A{info_row}'].font = Font(name='微软雅黑', size=9, color='FF0000')
    ws[f'A{info_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    acct_row = info_row + 1
    bank_text = ('对公账户：\n'
        '账户名：赛诺吉(深圳)国际货运代理有限公司\n'
        '公司地址：深圳市福田区沙头街道天安社区泰然四路29号天安创新科技广场一期A座1001B\n'
        '联系电话：0755-82720817\n'
        '账  号：41009000040015688\n'
        '开户行：中国农业银行深圳福田保税区支行')
    ws.merge_cells(f'A{acct_row}:AA{acct_row}')
    ws[f'A{acct_row}'].value = bank_text
    ws[f'A{acct_row}'].font = Font(name='微软雅黑', size=9)
    ws[f'A{acct_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[acct_row].height = 110
    
    note_row = acct_row + 1
    ws.merge_cells(f'A{note_row}:AA{note_row}')
    ws[f'A{note_row}'].value = '请勿用私人账户付到我司公账，谢谢配合。'
    ws[f'A{note_row}'].font = Font(name='微软雅黑', size=9, color='FF0000')
    ws[f'A{note_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    # ── Update title ──
    if title_str:
        ws['A2'].value = title_str
    
    # ── 开票金额 sheet ──
    ws_inv = wb['开票金额']
    for r in range(2, 6):
        for c in range(1, 6):
            ws_inv.cell(row=r, column=c).value = None
    
    # Write formulas referencing the bill sheet
    sheet_name = sheet_name_new
    ref_row = sr  # subtotal row
    
    inv_data = [
        ('国际货物运输代理服务', f"='{sheet_name}'!Q{ref_row}+'{sheet_name}'!R{ref_row}+'{sheet_name}'!U{ref_row}+'{sheet_name}'!W{ref_row}", '免税', 0),
        ('代理入仓费', f"='{sheet_name}'!O{ref_row}", 0.06, f"='{sheet_name}'!P{ref_row}"),
        ('经纪代理服务-报关费', f"='{sheet_name}'!S{ref_row}", 0.06, f"='{sheet_name}'!T{ref_row}"),
    ]
    
    ib = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i, (name, amt_formula, rate, tax_formula) in enumerate(inv_data):
        r = 2 + i
        ws_inv.cell(row=r, column=1, value=name).font = data_font
        ws_inv.cell(row=r, column=1).alignment = center
        ws_inv.cell(row=r, column=1).border = ib
        
        ws_inv.cell(row=r, column=2, value=amt_formula if isinstance(amt_formula, str) else amt_formula).font = data_font
        ws_inv.cell(row=r, column=2).alignment = center
        ws_inv.cell(row=r, column=2).number_format = '#,##0.00'
        ws_inv.cell(row=r, column=2).border = ib
        
        ws_inv.cell(row=r, column=3, value=rate if rate == '免税' else rate).font = data_font
        ws_inv.cell(row=r, column=3).alignment = center
        ws_inv.cell(row=r, column=3).border = ib
        if rate != '免税':
            ws_inv.cell(row=r, column=3).number_format = '0%'
        
        ws_inv.cell(row=r, column=4, value=tax_formula if isinstance(tax_formula, str) else tax_formula).font = data_font
        ws_inv.cell(row=r, column=4).alignment = center
        ws_inv.cell(row=r, column=4).number_format = '#,##0.00'
        ws_inv.cell(row=r, column=4).border = ib
        
        # 合计 = B + D
        ws_inv.cell(row=r, column=5, value=f'=B{r}+D{r}').font = data_font
        ws_inv.cell(row=r, column=5).alignment = center
        ws_inv.cell(row=r, column=5).number_format = '#,##0.00'
        ws_inv.cell(row=r, column=5).border = ib
    
    # Total row
    ws_inv.cell(row=5, column=1, value='总额').font = bold_font
    ws_inv.cell(row=5, column=1).alignment = center
    ws_inv.cell(row=5, column=1).border = ib
    ws_inv.cell(row=5, column=2, value='=B2+B3+B4').font = bold_font
    ws_inv.cell(row=5, column=2).alignment = center
    ws_inv.cell(row=5, column=2).number_format = '#,##0.00'
    ws_inv.cell(row=5, column=2).border = ib
    ws_inv.cell(row=5, column=4, value='=D2+D3+D4').font = bold_font
    ws_inv.cell(row=5, column=4).alignment = center
    ws_inv.cell(row=5, column=4).number_format = '#,##0.00'
    ws_inv.cell(row=5, column=4).border = ib
    ws_inv.cell(row=5, column=5, value='=E2+E3+E4').font = bold_font
    ws_inv.cell(row=5, column=5).alignment = center
    ws_inv.cell(row=5, column=5).number_format = '#,##0.00'
    ws_inv.cell(row=5, column=5).border = ib

    # ── 报价表A: update with actual price data ──
    ws_quote = wb['报价表A']
    # Clear old price rows (keep headers R1-R3)
    for r in range(4, 30):
        for c in range(1, 9):
            ws_quote.cell(row=r, column=c).value = None

    # Determine week info from date range
    week_str = ''
    month = datetime.now().month  # fallback if no date_range_str
    if date_range_str:
        yr = year or 2026
        # Extract month and start_day (handle "6月" or "6.1-6.7")
        nums = [int(x) for x in re.findall(r'\d+', date_range_str)]
        month = nums[0] if len(nums) > 0 else month
        start_day = nums[1] if len(nums) > 1 else 1
        from datetime import date as dt_date
        monday_dt = dt_date(yr, month, start_day)
        first_of_month = dt_date(yr, month, 1)
        days_to_monday = (7 - first_of_month.weekday()) % 7
        first_monday = first_of_month + timedelta(days=days_to_monday) if days_to_monday else first_of_month
        week_num = 1 if monday_dt < first_monday else 2 + (monday_dt - first_monday).days // 7
        week_map = {1:'第一周', 2:'第二周', 3:'第三周', 4:'第四周', 5:'第五周', 6:'第六周'}
        week_str = week_map.get(week_num, f'第{week_num}周')

    # Write price data rows
    if price_rows_raw:
        for i, vals in enumerate(price_rows_raw):
            r = 4 + i
            ch, wh, price, customs = vals[0], vals[1], vals[2], vals[3] if len(vals) > 3 else None
            if not wh:
                continue
            ws_quote.cell(row=r, column=1, value=month).font = data_font
            ws_quote.cell(row=r, column=1).alignment = center
            ws_quote.cell(row=r, column=2, value=week_str).font = data_font
            ws_quote.cell(row=r, column=2).alignment = center
            ws_quote.cell(row=r, column=3, value=ch).font = data_font
            ws_quote.cell(row=r, column=3).alignment = center
            ws_quote.cell(row=r, column=4, value=wh).font = data_font
            ws_quote.cell(row=r, column=4).alignment = center
            ws_quote.cell(row=r, column=5, value=price).font = data_font
            ws_quote.cell(row=r, column=5).alignment = center
            ws_quote.cell(row=r, column=5).number_format = '0.0'
            # Calculated columns as formulas
            for c_idx, formula_tmpl in [(6, f'=E{r}*0.07/1.06'), (7, f'=E{r}*0.35'), (8, f'=E{r}*0.58')]:
                cell = ws_quote.cell(row=r, column=c_idx)
                cell.value = formula_tmpl
                cell.font = data_font
                cell.alignment = center
                cell.number_format = '0.0000'

    # ── Save ──
    # Apply template column fills to all data rows (after all writes/merges)
    from openpyxl.cell.cell import MergedCell
    for rn in range(4, 4 + n):
        for col_idx in template_fills:
            cell = ws.cell(row=rn, column=col_idx)
            if not isinstance(cell, MergedCell):
                try:
                    cell.fill = template_fills[col_idx]
                except:
                    pass

    # 导出默认普通视图（模板的账单/报价表 sheet 为分页预览模式）
    for _ws in wb.worksheets:
        _ws.sheet_view.view = 'normal'

    wb.save(output_path)
    return True


def main():
    if len(sys.argv) < 2:
        print("用法: python3 gen_bill.py <内部拣货数据参考值.xlsx> [输出文件名]")
        sys.exit(1)

    ref_path = sys.argv[1]

    print(f"📂 内部拣货数据参考值: {ref_path}")

    # Load（仅参考值，已不依赖订单列表）
    ref_rows, warehouse_prices, price_rows_raw, declaration_groups = load_data(ref_path)
    n_so = len(ref_rows)
    n_ref = sum(len(v) for v in ref_rows.values())
    print(f"✅ 参考值SO: {n_so} 条, 行: {n_ref} 行, 仓库单价: {len(warehouse_prices)} 个")
    print(f"📋 报关组: 在账单内按 (走货渠道, SO) 判定")

    # 日期范围从参考值「下单时间」列计算（发货日期来源）
    date_serials = []
    for so, ref_list in ref_rows.items():
        d = parse_order_date(ref_list[0].get('order_time'))
        if d:
            date_serials.append((d - datetime(1899, 12, 30)).days)

    if date_serials:
        base = datetime(1899, 12, 30)
        min_dt = base + timedelta(days=min(date_serials))
        max_dt = base + timedelta(days=max(date_serials))
        # Find Monday of the week containing min_dt
        mon = min_dt - timedelta(days=min_dt.weekday())
        sun = mon + timedelta(days=6)
        year = mon.year
        date_range_str = f"{mon.month}.{mon.day}-{sun.month}.{sun.day}"
        title_str = f"至：广州拓锐科技有限公司（{mon.month}.{mon.day}-{sun.month}.{sun.day}）"
    else:
        year = datetime.now().year
        date_range_str = f"{datetime.now().month}.1-{datetime.now().month}.7"
        title_str = "至：广州拓锐科技有限公司"

    # Build rows
    rows = build_rows(ref_rows, warehouse_prices)

    # Sort
    rows = sort_rows(rows, declaration_groups=declaration_groups)
    print(f"✅ 账单行: {len(rows)} 行")

    # Compute total accurately (same formula path as Excel)
    sum_O = sum(r['weight'] * r['unit_price'] * 0.07/1.06 for r in rows)
    sum_P = sum_O * 0.06
    sum_Q = sum(r['weight'] * r['unit_price'] * 0.35 for r in rows)
    sum_R = sum(r['weight'] * r['unit_price'] * 0.58 for r in rows)
    customs_count = len(rows)  # 报关费按每一行收取
    customs_S = customs_count * 350 / 1.06
    customs_T = customs_S * 0.06
    total = sum_O + sum_P + sum_Q + sum_R + customs_S + customs_T
    total_rounded = round(total, 1)

    # Auto-generate filename (dynamic year/month)
    file_month = date_range_str.split('.')[0] if date_range_str and '.' in date_range_str else f'{datetime.now().month}'
    output_path = sys.argv[2] if len(sys.argv) > 2 else \
        f'{year}年{file_month}月拓锐FBA仓-分段开票账单-JTT({date_range_str}) RMB {total_rounded}.xlsx'

    print(f"📄 输出: {output_path}")

    # Generate bill
    success = generate_bill(rows, output_path, title_str=title_str, date_range_str=date_range_str, price_rows_raw=price_rows_raw, year=year, declaration_groups=declaration_groups)

    if success:
        print(f"\n📊 费用汇总:")
        print(f"   报关组: {customs_count} (报关费 {customs_count}×330.19 = {customs_S:.2f})")
        print(f"   国内运费: {sum_O:.2f} → 含税 {sum_O+sum_P:.2f}")
        print(f"   国际运费: {sum_Q+sum_R:.2f}")
        print(f"   报关费:   {customs_S:.2f} → 含税 {customs_S+customs_T:.2f}")
        print(f"   ─────────────────────────────")
        print(f"   ✅ 总计应收: {total_rounded} RMB")
        print(f"\n✅ 账单已生成: {output_path}")


if __name__ == '__main__':
    main()
