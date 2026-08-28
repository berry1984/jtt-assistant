#!/usr/bin/env python3
"""
JTT电商AI助手 - Web应用

功能：
  1. TR账单自动生成 - 上传订单列表+拣货数据+应收价格，生成标准格式账单
  2. TR发票转换 - 上传TR发票，转换为天图/航乐等供应商模板

用法：
  python3 app.py
  # 浏览器访问 http://localhost:5000
"""

import os
import sys
import json
import tempfile
import shutil
import atexit
import uuid
from datetime import datetime, timedelta
from flask import Flask, request, render_template, send_file, send_from_directory, flash, redirect, make_response

# ── 模块路径 ──
THIS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(THIS_DIR)  # bb plan1
INVOICE_DIR = os.path.join(PROJECT_DIR, '发票转换')

sys.path.insert(0, THIS_DIR)
sys.path.insert(0, INVOICE_DIR)

from gen_bill import load_data, build_rows, sort_rows, generate_bill, parse_order_date
from convert_invoice import (TRInvoice, convert_to_tiantu, convert_to_hangle,
                             convert_to_meiqi, _match_waybill)

# ── 提单及电放保函生成模块 ──
from gen_bl_docs import generate_bl_docs

# ── 思锐(SR)账单生成模块（延迟导入，避免启动时依赖缺失） ──
SR_DIR = os.path.join(PROJECT_DIR, 'SR账单自动生成')
sys.path.insert(0, PROJECT_DIR)

# ── 拣货数据导出模块 ──
PICKING_DIR = os.path.join(PROJECT_DIR, '拣货数据')

# ── 投保区间拆分模块 ──
INSURANCE_DIR = os.path.join(PROJECT_DIR, '投保区间拆分发票')
sys.path.insert(0, INSURANCE_DIR)

# ── 报价查询模块 ──
# 原指向 PROJECT_DIR/报价工具（bb plan1 下不存在）导致 500: spec not found for module 'price_query'
# 改为真实项目目录 /Users/admin/报价工具（内含已修复的 supplier_pricing / price_query / cost_analyzer）
# 该路径仅保留给 英美成本匹配(ym_cost_match) 本地回退；纯供应商报价查询已由 price_supplier.py 自包含
PRICE_QUERY_DIR = '/Users/admin/报价工具'
sys.path.insert(0, PRICE_QUERY_DIR)

# ── 报价数据持久化目录（Railway Volume /data，本地开发兜底 price_data/） ──
STORAGE_DIR = os.environ.get('STORAGE_DIR', '').strip()
if not STORAGE_DIR:
    _cand = '/data'
    if not (os.path.exists(_cand) and os.access(_cand, os.W_OK)):
        _cand = os.path.join(THIS_DIR, 'price_data')
    STORAGE_DIR = _cand
os.environ['STORAGE_DIR'] = STORAGE_DIR
os.makedirs(os.path.join(STORAGE_DIR, 'suppliers'), exist_ok=True)

def _get_sr_module():
    """延迟导入 gen_sr_bill，避免铁路部署时依赖问题导致整个app崩溃"""
    import importlib
    try:
        return importlib.import_module('gen_sr_bill')
    except ImportError as e:
        print(f"[SR] 导入gen_sr_bill失败: {e}")
        return None

app = Flask(__name__)
app.secret_key = 'jtt-ai-assistant-secret'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB（8家供应商文件合计约39MB，含模板留足余量）
app.config['UPLOAD_FOLDER'] = os.path.join(THIS_DIR, 'uploads')

# 启动时清理旧上传
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
for f in os.listdir(app.config['UPLOAD_FOLDER']):
    p = os.path.join(app.config['UPLOAD_FOLDER'], f)
    try:
        if os.path.isdir(p):
            shutil.rmtree(p)
        else:
            os.remove(p)
    except:
        pass

# ── 发票转换目标格式 ──
TARGET_OPTIONS = {
    '天图': '天图下单发票',
    '航乐-uk': '航乐-英国发票',
    '航乐-eu': '航乐-欧洲发票',
    '美琦': '美琦美线发票',
}


# ═══════════════════════════════════════════════════════════
#  首页
# ═══════════════════════════════════════════════════════════

@app.route('/', methods=['GET'])
def index():
    # 默认落地到「发票转换」（页面第一栏）
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='invoice')


@app.route('/bill', methods=['GET'])
def bill_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='bill')


# ═══════════════════════════════════════════════════════════
#  页脚
# ═══════════════════════════════════════════════════════════

@app.route('/debug_template', methods=['GET'])
def debug_template():
    """调试：查看模板文件信息和环境"""
    import os
    from openpyxl import load_workbook
    tpl_path = os.path.join(INVOICE_DIR, '天图单票专用模板20260601.xlsx')
    info = {'status': 'unknown', 'path': tpl_path, 'exists': os.path.exists(tpl_path)}
    # 检查Pillow
    try:
        import PIL
        info['pillow'] = PIL.__version__
    except ImportError:
        info['pillow'] = 'NOT INSTALLED'
    # git commit
    try:
        import subprocess
        sha = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'], capture_output=True, text=True, cwd=os.path.dirname(THIS_DIR))
        info['commit'] = sha.stdout.strip() if sha.returncode == 0 else 'unknown'
    except:
        info['commit'] = 'error'
    if info['exists']:
        info['size'] = os.path.getsize(tpl_path)
        try:
            wb = load_workbook(tpl_path, data_only=True)
            info['sheet2_rows'] = wb['Sheet2'].max_row
            ws2 = wb['Sheet2']
            first3 = [ws2.cell(r, 1).value for r in range(1, 4)]
            last3 = [ws2.cell(r, 1).value for r in range(info['sheet2_rows']-2, info['sheet2_rows']+1)]
            info['first_3'] = first3
            info['last_3'] = last3
            wb.close()
        except Exception as e:
            info['error'] = str(e)
    return info


@app.route('/invoice', methods=['GET'])
def invoice_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='invoice')


# ═══════════════════════════════════════════════════════════
#  功能1：TR账单自动生成
# ═══════════════════════════════════════════════════════════

@app.route('/generate', methods=['POST'])
def generate():
    order_file = request.files.get('order_file')
    pick_file = request.files.get('pick_file')
    price_file = request.files.get('price_file')

    if not all([order_file, pick_file, price_file]):
        flash('请上传三个文件：订单列表、拣货数据、应收价格')
        return redirect('/bill')

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        order_path = os.path.join(tmp_dir, 'order.xlsx')
        pick_path = os.path.join(tmp_dir, 'pick.xlsx')
        price_path = os.path.join(tmp_dir, 'price.xlsx')
        order_file.save(order_path)
        pick_file.save(pick_path)
        price_file.save(price_path)

        orders, picks, prices, price_rows_raw, declaration_groups = load_data(order_path, pick_path, price_path)
        rows = build_rows(orders, picks, prices)
        rows = sort_rows(rows, declaration_groups=declaration_groups)
        date_serials = []
        for o in orders.values():
            # 下单时间/创建日期优先（字符串格式也解析），为空回退发货→工作日期
            d = (parse_order_date(o.get('创建日期'))
                 or parse_order_date(o.get('下单时间'))
                 or parse_order_date(o.get('发货日期'))
                 or parse_order_date(o.get('工作日期')))
            if d:
                date_serials.append((d - datetime(1899, 12, 30)).days)

        if date_serials:
            base = datetime(1899, 12, 30)
            min_dt = base + timedelta(days=min(date_serials))
            max_dt = base + timedelta(days=max(date_serials))
            mon = min_dt - timedelta(days=min_dt.weekday())
            sun = mon + timedelta(days=6)
            year = mon.year
            date_range_str = f"{mon.month}.{mon.day}-{sun.month}.{sun.day}"
            title_str = f"至：广州拓锐科技有限公司（{mon.month}.{mon.day}-{sun.month}.{sun.day}）"
        else:
            year = datetime.now().year
            date_range_str = f"{datetime.now().month}.1-{datetime.now().month}.7"
            title_str = "至：广州拓锐科技有限公司"

        def safe_float(v):
            try: return float(v) if v is not None and v != '' else 0
            except: return 0
        sum_O = sum(safe_float(r.get('weight', 0)) * safe_float(r.get('unit_price', 0)) * 0.07/1.06 for r in rows)
        sum_P = sum_O * 0.06
        sum_Q = sum(safe_float(r.get('weight', 0)) * safe_float(r.get('unit_price', 0)) * 0.35 for r in rows)
        sum_R = sum(safe_float(r.get('weight', 0)) * safe_float(r.get('unit_price', 0)) * 0.58 for r in rows)
        customs_count = len(rows)  # 报关费按每一行收取
        customs_S = customs_count * 350 / 1.06
        customs_T = customs_S * 0.06
        total = round(sum_O + sum_P + sum_Q + sum_R + customs_S + customs_T, 1)

        file_month = date_range_str.split('.')[0] if '.' in date_range_str else f'{datetime.now().month}'
        output_name = f'{year}年{file_month}月拓锐FBA仓-分段开票账单-JTT({date_range_str}) RMB {total}.xlsx'
        output_path = os.path.join(tmp_dir, output_name)

        success = generate_bill(rows, output_path, title_str=title_str,
                                date_range_str=date_range_str, price_rows_raw=price_rows_raw, year=year,
                                declaration_groups=declaration_groups)

        if not success or not os.path.exists(output_path):
            flash('生成账单失败，请检查文件内容')
            return redirect('/bill')

        return send_file(output_path,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=output_name)

    except Exception as e:
        flash(f'处理出错: {str(e)}')
        return redirect('/bill')
    finally:
        pass


# ═══════════════════════════════════════════════════════════
#  功能2：TR发票 → 供应商模板 转换
# ═══════════════════════════════════════════════════════════

# 临时图片存储路径，用于 IMAGE() 公式引用
TEMP_IMAGE_DIR = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_images')
os.makedirs(TEMP_IMAGE_DIR, exist_ok=True)

@app.route('/temp_images/<session_id>/<filename>')
def serve_temp_image(session_id, filename):
    """提供转换时提取的临时图片（IMAGE 公式引用）"""
    return send_from_directory(os.path.join(TEMP_IMAGE_DIR, session_id), filename)

@app.route('/invoice_convert', methods=['POST'])
def invoice_convert():
    invoice_file = request.files.get('invoice_file')
    order_list_file = request.files.get('order_list')
    target = request.form.get('target', '天图')

    if not invoice_file or invoice_file.filename == '':
        flash('请上传 TR 发票文件')
        return redirect('/invoice')

    if target not in TARGET_OPTIONS:
        flash('请选择有效的目标格式')
        return redirect('/invoice')

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        invoice_path = os.path.join(tmp_dir, 'invoice.xlsx')
        invoice_file.save(invoice_path)

        tr = TRInvoice(invoice_path)

        # ── 提取图片到临时目录（供 IMAGE() 公式以 HTTP URL 引用；美琦新版无图片列，跳过） ──
        image_url_base = None
        if target != '美琦':
            image_session_id = uuid.uuid4().hex[:12]
            session_img_dir = os.path.join(TEMP_IMAGE_DIR, image_session_id)
            os.makedirs(session_img_dir, exist_ok=True)

            img_count = 0
            for src_row, img_bytes in tr.images.items():
                img_filename = f'image_tiantu_{img_count + 1}.png'
                with open(os.path.join(session_img_dir, img_filename), 'wb') as f:
                    f.write(img_bytes)
                img_count += 1

            if img_count > 0:
                # 构建服务器 URL 前缀
                host_url = request.host_url.rstrip('/')
                image_url_base = f'{host_url}/temp_images/{image_session_id}'

        base_name = os.path.splitext(invoice_file.filename)[0]

        # 订单列表（可选）：所有目标均支持按地址库编码回填客户订单号为运单号
        order_list_path = None
        if order_list_file and order_list_file.filename:
            order_list_path = os.path.join(tmp_dir, 'order_list.xlsx')
            order_list_file.save(order_list_path)

        # 文件名：航乐按 "客户名称 订单号 欧洲/英国发票.xlsx" 格式
        if target.startswith('航乐'):
            import re
            customer = ''
            m = re.search(r'（(.+?)发票）', base_name)
            if m:
                customer = m.group(1)
            order_no = _match_waybill(tr, order_list_path) or tr.get('客户订单号', '') or ''
            region_label = '欧洲' if target == '航乐-eu' else '英国'
            output_name = f'{customer} {order_no} {region_label}发票.xlsx'.strip()
            if output_name.startswith(' '):
                output_name = output_name.lstrip()
        else:
            ext_map = {'天图': '天图', '美琦': '美琦'}
            output_name = f"{base_name}-{ext_map.get(target, '天图')}.xlsx"

        output_path = os.path.join(tmp_dir, output_name)

        if target == '天图':
            ok = convert_to_tiantu(tr, output_path, image_url_base=image_url_base,
                                   order_list_path=order_list_path)
        elif target == '航乐-uk':
            ok = convert_to_hangle(tr, output_path, region='uk', image_url_base=image_url_base,
                                   order_list_path=order_list_path)
        elif target == '航乐-eu':
            ok = convert_to_hangle(tr, output_path, region='eu', image_url_base=image_url_base,
                                   order_list_path=order_list_path)
        elif target == '美琦':
            ok = convert_to_meiqi(tr, output_path, order_list_path=order_list_path)

        if not ok:
            flash('转换失败，请检查源文件格式')
            return redirect('/invoice')

        return send_file(output_path,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=output_name)

    except Exception as e:
        flash(f'转换出错: {str(e)}')
        return redirect('/invoice')
    finally:
        def cleanup():
            try: shutil.rmtree(tmp_dir)
            except: pass
        atexit.register(cleanup)


# ═══════════════════════════════════════════════════════════
#  功能3：思锐(SR)账单自动生成  （使用订单列表渠道+可调汇率）
# ═══════════════════════════════════════════════════════════

@app.route('/generate_sr', methods=['POST'])
def generate_sr():
    sr = _get_sr_module()
    if not sr:
        flash('思锐账单模块加载失败，请联系管理员')
        return redirect('/sr')

    bill_file = request.files.get('sr_bill_file')
    order_file = request.files.get('sr_order_file')

    if not bill_file:
        flash('请上传系统账单文件')
        return redirect('/sr')

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        # 保存上传文件
        bill_path = os.path.join(tmp_dir, 'system_bill.xls')
        bill_file.save(bill_path)

        order_list = {}
        order_path = None
        if order_file and order_file.filename:
            order_path = os.path.join(tmp_dir, 'order_list.xlsx')
            order_file.save(order_path)
            order_list = sr.read_order_list(order_path)

        # AB2 汇率（用户手动输入，默认0.1282）
        try:
            ab2_rate = float(request.form.get('ab2_rate', '0.1282'))
        except ValueError:
            ab2_rate = 0.1282

        # 读取系统账单
        waybills = sr.read_system_bill(bill_path)
        if not waybills:
            flash('系统账单中没有找到运单数据')
            return redirect('/sr')

        # 自动查找模板
        template_path = os.path.join(SR_DIR, '思锐账单模板模板 思锐开票账单-JTT（5.1-5.31）.xlsx')
        if not os.path.exists(template_path):
            for f in os.listdir(SR_DIR):
                if '思锐' in f and f.endswith('.xlsx') and '模板' in f:
                    template_path = os.path.join(SR_DIR, f)
                    break

        if not os.path.exists(template_path):
            flash('找不到思锐账单模板文件')
            return redirect('/sr')

        # 生成输出文件名
        from datetime import date
        today = date.today()
        output_name = f'思锐开票账单-JTT（{today.year}.{today.month:02d}.01-{today.month:02d}.{today.day:02d}）.xlsx'
        output_path = os.path.join(tmp_dir, output_name)

        # 生成账单（传入自定义AB2汇率）
        success = sr.generate_bill(waybills, template_path, output_path, order_list, ab2_rate=ab2_rate)

        if not success or not os.path.exists(output_path):
            flash('生成思锐账单失败')
            return redirect('/sr')

        return send_file(output_path,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=output_name)

    except Exception as e:
        flash(f'处理出错: {str(e)}')
        return redirect('/sr')
    finally:
        pass


@app.route('/sr', methods=['GET'])
def sr_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='sr')


# ═══════════════════════════════════════════════════════════
#  功能4：拣货数据导出
# ═══════════════════════════════════════════════════════════

@app.route('/picking', methods=['GET'])
def picking_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='picking')


@app.route('/picking_export', methods=['POST'])
def picking_export():
    """
    新规则（2026-06-10）：
    上传 发票 + 系统导出拣货数据 → 匹配箱规历史数据库 → 输出内部拣货数据参考值
    """
    sys.path.insert(0, PICKING_DIR)

    invoice_files = request.files.getlist('picking_invoice')
    system_file = request.files.get('picking_system')
    history_file = request.files.get('picking_history')
    quotation_file = request.files.get('picking_quotation')

    if not invoice_files or all(f.filename == '' for f in invoice_files):
        flash('请上传至少一份发票文件')
        return redirect('/picking')
    if not system_file or system_file.filename == '':
        flash('请上传系统导出拣货数据文件')
        return redirect('/picking')

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        from export_picking_data import generate_picking_output_multi, HISTORY_FILE, TEMPLATE_FILE, QUOTATION_FILE

        # 检查服务器端文件
        if not os.path.exists(TEMPLATE_FILE):
            flash(f'服务器缺少输出模板: {TEMPLATE_FILE}')
            return redirect('/picking')

        # 保存上传的发票文件（可能多份）
        invoice_paths = []
        for f in invoice_files:
            if f.filename:
                path = os.path.join(tmp_dir, f'invoice_{len(invoice_paths)}.xlsx')
                f.save(path)
                invoice_paths.append(path)

        # 保存系统导出文件
        system_path = os.path.join(tmp_dir, 'system.xlsx')
        system_file.save(system_path)

        # 箱规历史数据库：上传了就使用上传的，否则用服务器默认
        if history_file and history_file.filename:
            history_path = os.path.join(tmp_dir, 'history.xlsx')
            history_file.save(history_path)
        else:
            history_path = HISTORY_FILE
            if not os.path.exists(history_path):
                flash(f'服务器缺少箱规历史数据库: {history_path}')
                return redirect('/picking')

        # 报价单：上传了就使用上传的，否则用服务器默认
        if quotation_file and quotation_file.filename:
            quotation_path = os.path.join(tmp_dir, 'quotation.xlsx')
            quotation_file.save(quotation_path)
        else:
            quotation_path = QUOTATION_FILE
            if not os.path.exists(quotation_path):
                flash(f'服务器缺少报价单: {quotation_path}')
                return redirect('/picking')

        output_path = os.path.join(tmp_dir, 'temp_output.xlsx')

        result, total_boxes = generate_picking_output_multi(invoice_paths, system_path, output_path,
                                                             history_file=history_path,
                                                             quotation_file=quotation_path)

        # 重命名为带日期+箱数的文件名
        from datetime import date
        today_str = date.today().strftime('%Y-%m-%d')
        output_name = f'内部拣货数据参考值_{today_str}_{total_boxes}箱.xlsx'
        final_path = os.path.join(tmp_dir, output_name)
        if os.path.exists(result):
            os.rename(result, final_path)

        return send_file(final_path,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=output_name)

    except Exception as e:
        flash(f'处理出错: {str(e)}')
        return redirect('/picking')
    finally:
        pass


# ═══════════════════════════════════════════════════════════
#  功能5：投保区间拆分
# ═══════════════════════════════════════════════════════════

@app.route('/insurance', methods=['GET'])
def insurance_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='insurance')


@app.route('/insurance_split', methods=['POST'])
def insurance_split():
    """
    上传下单发票 → 按每箱RMB拆分为 5 个区间文件 → 打包 ZIP 下载
    """
    invoice_file = request.files.get('insurance_file')
    if not invoice_file or invoice_file.filename == '':
        flash('请上传下单发票文件')
        return redirect('/insurance')

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        invoice_path = os.path.join(tmp_dir, 'invoice.xlsx')
        invoice_file.save(invoice_path)

        # 调用拆分模块
        from split_insurance_v2 import split_invoice_to_ranges
        out_dir = os.path.join(tmp_dir, 'ranges')
        out_files = split_invoice_to_ranges(invoice_path, output_dir=out_dir)

        if not out_files:
            flash('拆分失败，未生成任何文件')
            return redirect('/insurance')

        # 打包为 ZIP
        import zipfile
        base_name = os.path.splitext(invoice_file.filename)[0]
        zip_name = f'{base_name}-投保拆分.zip'
        zip_path = os.path.join(tmp_dir, zip_name)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for range_name, file_path in out_files.items():
                arcname = os.path.basename(file_path)
                zf.write(file_path, arcname)

        return send_file(zip_path,
                         mimetype='application/zip',
                         as_attachment=True,
                         download_name=zip_name)

    except Exception as e:
        flash(f'拆分出错: {str(e)}')
        return redirect('/insurance')
    finally:
        def cleanup():
            try:
                shutil.rmtree(tmp_dir)
            except Exception:
                pass
        import atexit
        atexit.register(cleanup)


# ═══════════════════════════════════════════════
#  功能6：提单及电放保函生成
# ═══════════════════════════════════════════════

@app.route('/bl_docs', methods=['GET'])
def bl_docs_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='bl_docs')


@app.route('/generate_bl_docs', methods=['POST'])
def generate_bl_docs_route():
    excel_file = request.files.get('excel_file')
    if not excel_file or excel_file.filename == '':
        flash('请上传含提单信息的 Excel 文件')
        return redirect('/bl_docs')

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        excel_path = os.path.join(tmp_dir, 'data.xlsx')
        excel_file.save(excel_path)

        zip_path, telex_ok, bl_ok = generate_bl_docs(excel_path)

        fname = os.path.basename(zip_path)
        response = make_response(send_file(zip_path,
                         mimetype='application/zip',
                         as_attachment=True,
                         download_name=fname))
        response.headers['X-Telex-Count'] = str(telex_ok)
        response.headers['X-Bl-Count'] = str(bl_ok)
        return response

    except Exception as e:
        flash(f'生成出错: {str(e)}')
        return redirect('/bl_docs')
    finally:
        def cleanup():
            try:
                shutil.rmtree(tmp_dir)
            except Exception:
                pass
        import atexit
        atexit.register(cleanup)


# ═══════════════════════════════════════════
#  功能7：报价查询（固定供应商 + 自定义 + 云端持久化）
# ═══════════════════════════════════════════
from price_supplier import save_upload, add_custom_supplier, delete_slot, list_slots, query


@app.route('/price_query', methods=['GET'])
def price_query_page():
    return render_template('index.html', targets=TARGET_OPTIONS, active_tab='price_query')


@app.route('/api/price_query', methods=['GET'])
def price_query_api():
    """供应商报价查询：按 供应商 → 国家 → 渠道 组织，支持按国家/供应商/渠道关键词过滤。"""
    keyword = request.args.get('keyword', '').strip()
    country = request.args.get('country', '').strip()
    supplier = request.args.get('supplier', '').strip()
    try:
        data = query(keyword=keyword, country=country, supplier=supplier)
        return app.response_class(
            response=json.dumps(data, ensure_ascii=False),
            mimetype='application/json; charset=utf-8'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}


@app.route('/api/price_query/slots', methods=['GET'])
def price_query_slots():
    """已存报价槽位清单（供应商|国家|源文件名|更新时间|渠道数|自定义标记）。"""
    try:
        return {"slots": list_slots()}
    except Exception as e:
        return {"error": str(e)}


@app.route('/api/price_query/upload', methods=['POST'])
def price_query_upload():
    """上传报价文件到 (供应商 × 多国) 槽位，再次上传同槽位即覆盖更新。"""
    supplier = request.form.get('supplier', '').strip()
    countries = request.form.getlist('countries')
    file = request.files.get('file')
    if not file or file.filename == '':
        return {"error": "请选择报价文件"}
    return save_upload(supplier, countries, file)


@app.route('/api/price_query/supplier/add', methods=['POST'])
def price_query_supplier_add():
    """新增自定义供应商并上传其首个报价文件。"""
    name = request.form.get('name', '').strip()
    countries = request.form.getlist('countries')
    file = request.files.get('file')
    if not file or file.filename == '':
        return {"error": "请选择报价文件"}
    return add_custom_supplier(name, countries, file)


@app.route('/api/price_query/delete', methods=['POST'])
def price_query_delete():
    """删除单个 (供应商 × 国家) 槽位。"""
    supplier = request.form.get('supplier', '').strip()
    country = request.form.get('country', '').strip()
    if not supplier or not country:
        return {"error": "缺少 supplier/country 参数"}
    return delete_slot(supplier, country)


# ═══════════════════════════════════════════
#  功能8：英美成本报价匹配（英美跨境专版 v2.0）
# ═══════════════════════════════════════════

def _get_ym_module():
    """延迟导入/重载 ym_cost_match 模块。
    桌面优先用 /Users/admin/报价工具 的本地活版；云端(Railway)无该路径时用仓库内置版
    (TR账单自动生成/ym_cost_match.py)。"""
    import importlib
    mod_name = 'ym_cost_match'
    if mod_name in sys.modules:
        return importlib.reload(sys.modules[mod_name])
    for base in (PRICE_QUERY_DIR, THIS_DIR):
        if os.path.isfile(os.path.join(base, 'ym_cost_match.py')):
            sys.path.insert(0, base)
            break
    return importlib.import_module(mod_name)


@app.route('/ym_cost', methods=['GET'])
def ym_cost_page():
    # ?embed=1 时隐藏顶栏，供首页「报价模块 → 英美成本匹配」子 Tab 以 iframe 嵌入
    embed = request.args.get('embed', '0') == '1'
    return render_template('ym_cost.html', active_tab='ym_cost', embed=embed)


@app.route('/api/ym_cost/generate', methods=['POST'])
def ym_cost_generate():
    """
    上传 输出模板 + 渠道匹配表 + 4份供应商文件 → 生成英美成本报价匹配 Excel。

    参数:
        template_file:   JTT物流每周成本报价分析表输出模板-英美（必填）
        mapping_file:    JTT+英美渠道匹配表（必填）
        supplier_files:  供应商报价文件：美线/英欧线/加拿大/空派 各1份（可多选）
    """
    ym = _get_ym_module()
    if not ym:
        return {"error": "英美成本匹配模块加载失败，请确认 ym_cost_match.py 已部署"}, 500

    template_file = request.files.get('template_file')
    mapping_file = request.files.get('mapping_file')
    supplier_files = request.files.getlist('supplier_files')

    if not template_file or template_file.filename == '':
        return {"error": "请上传 JTT 输出模板文件"}, 400
    if not mapping_file or mapping_file.filename == '':
        return {"error": "请上传 JTT+英美渠道匹配表文件"}, 400
    if not supplier_files or all(f.filename == '' for f in supplier_files):
        return {"error": "请上传至少一份供应商报价文件"}, 400

    tmp_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    try:
        # 1. 保存模板与匹配表
        tmpl_path = os.path.join(tmp_dir, 'template.xlsx')
        template_file.save(tmpl_path)
        map_path = os.path.join(tmp_dir, 'mapping.xlsx')
        mapping_file.save(map_path)

        # 2. 保存供应商文件（保留原文件名，供 discover_supplier_files 按模式识别）
        sup_dir = os.path.join(tmp_dir, 'suppliers')
        os.makedirs(sup_dir, exist_ok=True)
        sup_paths = []
        for f in supplier_files:
            if f.filename:
                p = os.path.join(sup_dir, os.path.basename(f.filename))
                f.save(p)
                sup_paths.append(p)
        if not sup_paths:
            return {"error": "没有有效的供应商文件"}, 400

        # 3. 校验 4 类供应商文件是否齐全（英美跨境-美线/英欧线/加拿大/空派）
        found = ym.discover_supplier_files(sup_dir)
        missing = [ft for ft in ('us', 'eu', 'ca', 'air') if ft not in found]
        if missing:
            return {"error": "供应商文件不完整，缺少: " + ', '.join(missing) +
                            "（需上传 美线 / 英欧线 / 加拿大 / 空派 4 类文件）"}, 400

        # 4. 执行匹配生成
        output_path = os.path.join(tmp_dir, 'output.xlsx')
        output_path, stats = ym.generate(
            template_path=tmpl_path,
            mapping_path=map_path,
            supplier_dir=sup_dir,
            output_path=output_path,
        )

        if not output_path or not os.path.exists(output_path):
            return {"error": "生成失败，未产生输出文件"}, 500

        # 5. 返回生成的文件（附带匹配统计头，供前端展示）
        download_name = f"JTT物流每周成本报价分析表-英美_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = make_response(send_file(
            output_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=download_name,
        ))
        response.headers['X-YM-Rows'] = str(sum(s.get('rows', 0) for s in stats.values()))
        response.headers['X-YM-Matched'] = str(sum(s.get('matched', 0) for s in stats.values()))
        response.headers['X-YM-Unmatched'] = str(sum(s.get('unmatched', 0) for s in stats.values()))
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"生成失败: {str(e)}"}, 500
    finally:
        def cleanup():
            try: shutil.rmtree(tmp_dir)
            except: pass
        import atexit as at
        at.register(cleanup)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("=" * 50)
    print("  JTT电商AI助手  — 一站式跨境物流工具")
    print("=" * 50)
    print(f"  📄 发票转换  → http://localhost:{port}/invoice")
    print(f"  📦 拣货导出  → http://localhost:{port}/picking")
    print(f"  🛡️ 投保拆分  → http://localhost:{port}/insurance")
    print(f"  🗂️ 客户账单  → http://localhost:{port}/bill  (TR账单) / {port}/sr (思锐账单)")
    print(f"  💰 报价模块  → http://localhost:{port}/price_query (供应商报价查询) / {port}/ym_cost (英美成本)")
    print(f"  💾 报价存储  → {STORAGE_DIR}")
    print(f"  📜 提单保函  → http://localhost:{port}/bl_docs")
    print("=" * 50)
    app.run(host='0.0.0.0', port=port)
