#!/usr/bin/env python3
"""
思锐(SR)账单自动生成器
根据系统导出的账单Excel文件，填充到思锐账单模板中

用法:
  python3 gen_sr_bill.py <系统账单.xls> [输出文件名.xlsx]

处理逻辑:
  1. 读取系统账单"运单"sheet，按运单号分组
  2. 提取各费用类型（运费、保费、报关费、清关费、超品名费等）
  3. 根据规则确定受益部门、业务经理
  4. 填写到思锐账单模板的"原始账单"sheet中
  5. 保留模板格式、汇率信息和银行账户信息
"""

import sys, os, shutil, re
from datetime import datetime
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook

# xlrd 在函数内延迟导入，避免Railway部署时因依赖问题导致整个app崩溃
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 国家映射 ──
COUNTRY_CN = {'DE': '德国', 'GB': '英国'}
POD_MAP = {'DE': 'GERMANY', 'GB': 'THE UNITED KINGDOM'}

# ── 费用类型 → 模板列映射 ──
FEE_MAPPING = {
    '运费':             'O',    # 海运费（用公式 N*M 计算）
    '超品名费':         'Q',
    '超重费':           'R',    # 模板中R列是"超尺寸费"，系统超重费也放这里
    '出口(国内)报关费':  'V',
    '进口(海外)清关费':  'Z',
    '保费':             'AA',
}

# ── 样式 ──
THIN_BORDER = Border(
    left=Side(style='hair'), right=Side(style='hair'),
    top=Side(style='hair'), bottom=Side(style='hair'))
DATA_FONT = Font(name='微软雅黑', size=9)
BOLD_FONT = Font(name='微软雅黑', size=9, bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
CNY_FMT = '#,##0.00'
EUR_FMT = '#,##0.00'
INT_FMT = '0'
TEXT_FMT = '@'


def parse_unit_price(price_str):
    """解析单价字符串如 '4.40/KG' → 4.40"""
    if not price_str or price_str in ('/', '', '0'):
        return 0
    m = re.match(r'([\d.]+)', str(price_str).replace(',', ''))
    return float(m.group(1)) if m else 0


def read_system_bill(xls_path):
    """
    读取系统账单xls文件，按运单号分组解析。
    返回: OrderedDict { 运单号: waybill_info }
    """
    import xlrd  # 延迟导入，避免Railway部署时因依赖问题崩溃
    wb = xlrd.open_workbook(xls_path)
    ws = wb.sheet_by_name('运单')

    waybills = OrderedDict()

    for r in range(1, ws.nrows):
        row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]

        wb_no      = str(row_vals[2]).strip()
        date_val   = str(int(row_vals[0])) if row_vals[0] else ''
        ext_no     = str(row_vals[3]).strip() if row_vals[3] else ''
        country    = str(row_vals[10]).strip()
        weight     = row_vals[14]  # 收费重
        fee_type   = str(row_vals[15]).strip()
        unit_price = str(row_vals[16]).strip()
        amount     = row_vals[17]
        desc       = str(row_vals[19]).strip() if row_vals[19] else ''

        if not wb_no:
            continue

        if wb_no not in waybills:
            waybills[wb_no] = {
                'date': date_val,
                'country': country,
                'ext_no': ext_no,
                'weight': weight,
                'has_fba': bool(ext_no) and 'FBA' in ext_no,
                'fee_unit_prices': {},
                'fees': {},
                'fee_unit_prices': {},
            }

        wb_info = waybills[wb_no]
        # 如果有更新的日期则更新
        if date_val and (not wb_info['date'] or date_val > wb_info['date']):
            wb_info['date'] = date_val
        # 如果有扩展单号则更新
        if ext_no and not wb_info['ext_no']:
            wb_info['ext_no'] = ext_no
            wb_info['has_fba'] = 'FBA' in ext_no
        # 如果扩展单号是FBA但之前没有，更新
        if ext_no and 'FBA' in ext_no and not wb_info['has_fba']:
            wb_info['has_fba'] = True

        # 存储费用
        if fee_type:
            wb_info['fees'][fee_type] = amount
            if fee_type == '运费':
                wb_info['fee_unit_prices'][fee_type] = unit_price

        # 描述中有额外信息也记录
        if desc:
            wb_info['desc'] = desc

    return waybills


def read_order_list(xlsx_path):
    """
    读取思锐订单列表，按运单号建立映射
    返回: dict { 运单号: { channel, country, fba, weight } }
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        return {}

    wb = load_workbook(xlsx_path)
    ws = wb.active
    orders = {}
    for r in range(2, ws.max_row + 1):
        wb_no = str(ws.cell(row=r, column=1).value or '').strip()
        if not wb_no:
            continue
        orders[wb_no] = {
            'channel': str(ws.cell(row=r, column=34).value or '').strip(),  # AH=34 服务
            'country': str(ws.cell(row=r, column=35).value or '').strip(),  # AI=35 国家
            'fba':     str(ws.cell(row=r, column=3).value or '').strip(),   # C=3 扩展单号/FBA
            'weight':  ws.cell(row=r, column=10).value,                     # J=10 收费重
            'pieces':  ws.cell(row=r, column=4).value,                      # D=4 件数
        }
    return orders


def determine_channel(wb_no, order_list):
    """从订单列表中获取渠道"""
    info = order_list.get(wb_no, {})
    return info.get('channel', '')


def get_dept_manager(country, has_fba):
    """根据国家和是否有FBA确定受益部门和业务经理"""
    if country == 'DE':
        if has_fba:
            return '德国FBA', '温永强'
        else:
            return '德国子公司', '郑猛'
    elif country == 'GB':
        if has_fba:
            return '英国FBA', '温永强'
        else:
            return '英国子公司', '郑猛'
    else:
        return '', ''


def get_pod(country):
    """获取目的港英文名"""
    return POD_MAP.get(country, '')


def format_date(date_str):
    """将日期字符串 YYYYMMDD 格式化为 M月D日"""
    if not date_str or len(date_str) < 8:
        return ''
    try:
        dt = datetime(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
        return f"{dt.month}月{dt.day}日"
    except:
        return date_str


def get_date_serial(date_str):
    """将日期字符串转为Excel日期序列号"""
    if not date_str or len(date_str) < 8:
        return None
    try:
        dt = datetime(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
        base = datetime(1899, 12, 30)
        return (dt - base).days
    except:
        return None


def generate_bill(waybills, template_path, output_path, order_list=None, ab2_rate=None):
    """生成思锐账单
    ab2_rate: 可选，自定义人民币→欧元汇率，默认使用模板中的AB2值
    """
    if not os.path.exists(template_path):
        print(f"❌ 模板文件不存在: {template_path}")
        return False

    # 复制模板
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['原始账单']

    # ── 读取汇率信息（保留）──
    v2_rate = ws['V2'].value or 8.2939  # EUR→CNY
    y2_rate = ws['Y2'].value or 1.1914  # GBP→EUR
    # 如果传入了自定义AB2汇率，则使用自定义值并同步更新V2
    if ab2_rate is not None:
        ab2_rate = float(ab2_rate)
        ws['AB2'].value = ab2_rate
        ws['V2'].value = round(1.0 / ab2_rate, 4) if ab2_rate > 0 else 8.2939
    else:
        ab2_rate = ws['AB2'].value or 0.1282  # CNY→EUR

    # ── 清除旧数据（保留标题行1-3和汇率行2）──
    # 1. 取消合并数据区域的合并单元格
    for mr in list(ws.merged_cells.ranges):
        if 4 <= mr.min_row < 40:
            ws.unmerge_cells(str(mr))

    # 2. 清除行4-34的内容
    for row in range(4, 35):
        for col in range(1, 45):  # A~AR
            cell = ws.cell(row=row, column=col)
            cell.value = None
            # 重置样式（只清除边框，保留字体等默认）
            cell.border = Border()

    # 3. 清除旧合计行和模板残余公式
    for row in range(35, 40):
        for col in range(1, 45):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.border = Border()

    # ── 构建数据行列表（按运单号排序）──
    data_rows = []
    for wb_no, info in waybills.items():
        fees = info['fees']
        country = info['country']
        has_fba = info['has_fba']
        weight = info['weight']

        # 从订单列表获取渠道和其他信息
        order_info = order_list.get(wb_no, {}) if order_list else {}
        channel = order_info.get('channel', info.get('channel', ''))
        order_fba = order_info.get('fba', '')
        order_weight = order_info.get('weight', 0)

        # 如果订单列表中有FBA号且系统账单没有，则使用订单列表的
        if order_fba and (not info['ext_no'] or '/' in info['ext_no']):
            ext_no = order_fba
            has_fba = 'FBA' in order_fba or bool(order_fba)
        else:
            ext_no = info['ext_no']

        # 如果订单列表中有计费重，可以补充
        if order_weight and not weight:
            try:
                weight = float(order_weight)
            except:
                pass

        # 提取运费相关信息
        freight_amount_raw = fees.get('运费', 0)
        try:
            freight_amount = float(freight_amount_raw) if freight_amount_raw else 0
        except (ValueError, TypeError):
            freight_amount = 0
        freight_rate_str = info['fee_unit_prices'].get('运费', '')
        freight_rate = parse_unit_price(freight_rate_str) if freight_rate_str else 0

        # 如果运费费率为0但有运费金额，计算费率
        if freight_rate == 0 and freight_amount > 0 and weight > 0:
            freight_rate = round(freight_amount / weight, 2)

        # 确定渠道
        channel = determine_channel(wb_no, order_list) if order_list else ''

        # 确定受益部门和业务经理
        dept, manager = get_dept_manager(country, has_fba)

        # 国家中文名
        country_cn = COUNTRY_CN.get(country, country)

        def safe_float(v):
            """安全转浮点"""
            try:
                return float(v) if v else 0
            except (ValueError, TypeError):
                return 0

        # 提取各项费用
        fee_o = safe_float(freight_amount)  # 海运费
        fee_p = safe_float(fees.get('提货费', 0))
        fee_q = safe_float(fees.get('超品名费', 0))
        fee_r = safe_float(fees.get('超重费', 0))  # 模板中R列"超尺寸费"
        fee_s = 0  # 私人地址费(没有映射)
        fee_t = 0  # 其他可归属于海运的费用
        fee_v = safe_float(fees.get('出口(国内)报关费', 0))
        fee_w = 0  # 报关续页费
        fee_y = 0  # 欧洲申报费
        fee_z = safe_float(fees.get('进口(海外)清关费', 0))
        fee_aa = safe_float(fees.get('保费', 0))

        # 税金 → AN列（欧元）按用户要求：税金*AB2
        tax_cny_raw = fees.get('税金', 0)
        try:
            tax_cny = float(tax_cny_raw) if tax_cny_raw else 0
        except (ValueError, TypeError):
            tax_cny = 0
        has_tax = tax_cny > 0
        if has_tax:
            # 税金在系统账单中是人民币，用AB2汇率转为欧元
            if ab2_rate and ab2_rate != 0:
                fee_an_eur = round(tax_cny * ab2_rate, 2)
            elif v2_rate and v2_rate != 0:
                fee_an_eur = round(tax_cny / v2_rate, 2)
            else:
                fee_an_eur = 0
        else:
            fee_an_eur = 0

        # 计费重
        try:
            weight_f = float(weight) if weight else 0
            bill_weight = round(weight_f)
        except (ValueError, TypeError):
            bill_weight = 0

        row_data = {
            'wb_no': wb_no,
            'channel': channel,
            'country': country_cn,
            'pod': get_pod(country),
            'dept': dept,
            'manager': manager,
            'weight': bill_weight,
            'unit_price': freight_rate,
            'fee_o': fee_o,    # 海运费
            'fee_p': fee_p,    # 提货费
            'fee_q': fee_q,    # 超品名费
            'fee_r': fee_r,    # 超尺寸费(超重费)
            'fee_s': fee_s,    # 私人地址费
            'fee_t': fee_t,    # 其他海运费用
            'fee_v': fee_v,    # 报关费
            'fee_w': fee_w,    # 报关续页费
            'fee_y': fee_y,    # 欧洲申报费
            'fee_z': fee_z,    # 清关费
            'fee_aa': fee_aa,  # 保险费
            'fee_an_eur': fee_an_eur,  # VAT欧元
            'has_tax': has_tax,       # 是否有税金
            'ext_no': info['ext_no'],
            'date': info['date'],
            'pieces': order_info.get('pieces', '') if order_info else '',
        }
        data_rows.append(row_data)

    # ── 按渠道排序，相同渠道的运单排在一起 ──
    channel_order = OrderedDict()
    for r in data_rows:
        ch = r['channel']
        if ch not in channel_order:
            channel_order[ch] = len(channel_order)
    data_rows.sort(key=lambda r: (channel_order.get(r['channel'], 999), r['wb_no']))

    n = len(data_rows)
    print(f"📊 生成 {n} 行数据")

    # ── 填充数据行 ──
    for i, r in enumerate(data_rows):
        row_num = 4 + i
        ws.row_dimensions[row_num].height = 20

        # A: 采购订单号(留空)
        # B: 思锐销售订单号(留空)
        # C: FBA单号
        fba = r['ext_no'] if r['ext_no'] and r['ext_no'] != '/' else '/'
        # D: 货代运单号
        # E: 渠道
        # F: 国家
        # G: 报关单号(留空/待补充)
        # H: POL
        # I: POD
        # J: 受益部门
        # K: 业务经理
        # L: 件数(从订单列表D列取值)
        # M: 计费重
        # N: 单价
        # O: 海运费 = N*M
        # P: 提货费
        # Q: 超品名费
        # R: 超尺寸费(超重费)
        # S: 私人地址费
        # T: 其他海运费用

        # 基础信息列
        base_cells = [
            ('A', ''),                    # 采购订单号
            ('B', ''),                    # 销售订单号
            ('C', fba),                   # FBA单号
            ('D', r['wb_no']),            # 运单号
            ('E', r['channel']),          # 渠道
            ('F', r['country']),          # 国家
            ('G', '待补充'),              # 报关单号
            ('H', 'SHENZHEN'),            # POL
            ('I', r['pod']),              # POD
            ('J', r['dept']),             # 受益部门
            ('K', r['manager']),          # 业务经理
            ('L', r['pieces'] if r['pieces'] else ''),  # 件数(从订单列表D列)
            ('M', r['weight'] if r['weight'] else ''),  # 计费重
        ]

        for col_l, val in base_cells:
            cell = ws[f'{col_l}{row_num}']
            cell.value = val
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # 写入费用值（O~T 各项费用 + V报关费 + Z清关费 + AA保险费）
        fee_cells = [
            ('O', r['fee_o']), ('P', r['fee_p']), ('Q', r['fee_q']),
            ('R', r['fee_r']), ('S', r['fee_s']), ('T', r['fee_t']),
            ('V', r['fee_v']), ('Z', r['fee_z']), ('AA', r['fee_aa']),
        ]
        for col_l, val in fee_cells:
            cell = ws[f'{col_l}{row_num}']
            cell.value = val if val else 0
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            cell.number_format = CNY_FMT

        # O列特殊处理：有单价和计费重时用公式
        if r['unit_price'] > 0 and r['weight'] > 0:
            ws[f'O{row_num}'].value = f'=N{row_num}*M{row_num}'
            ws[f'O{row_num}'].number_format = CNY_FMT

        # N: 单价（有运费时才填）
        if r['unit_price'] > 0:
            cell_n = ws[f'N{row_num}']
            cell_n.value = r['unit_price']
            cell_n.font = DATA_FONT
            cell_n.alignment = CENTER_ALIGN
            cell_n.border = THIN_BORDER
            cell_n.number_format = '0.00'
        else:
            ws[f'N{row_num}'].font = DATA_FONT
            ws[f'N{row_num}'].alignment = CENTER_ALIGN
            ws[f'N{row_num}'].border = THIN_BORDER

    # ── 设置公式列（需要所有数据行都写入后才能设置公式）──
    for i, r in enumerate(data_rows):
        row_num = 4 + i

        # U = SUM(O:T) 运费合计
        ws[f'U{row_num}'].value = f'=SUM(O{row_num}:T{row_num})'
        ws[f'U{row_num}'].font = DATA_FONT
        ws[f'U{row_num}'].alignment = CENTER_ALIGN
        ws[f'U{row_num}'].border = THIN_BORDER
        ws[f'U{row_num}'].number_format = CNY_FMT

        # V: 报关费(已填入)
        # W: 报关续页费
        ws[f'W{row_num}'].value = 0
        ws[f'W{row_num}'].font = DATA_FONT
        ws[f'W{row_num}'].alignment = CENTER_ALIGN
        ws[f'W{row_num}'].border = THIN_BORDER
        ws[f'W{row_num}'].number_format = CNY_FMT

        # X = SUM(V:W) 报关费合计
        ws[f'X{row_num}'].value = f'=SUM(V{row_num}:W{row_num})'
        ws[f'X{row_num}'].font = DATA_FONT
        ws[f'X{row_num}'].alignment = CENTER_ALIGN
        ws[f'X{row_num}'].border = THIN_BORDER
        ws[f'X{row_num}'].number_format = CNY_FMT

        # Y: 欧洲申报费(已填/0)
        if '欧洲申报费' in r.get('fees', {}):
            ws[f'Y{row_num}'].value = r['fees']['欧洲申报费']
        else:
            ws[f'Y{row_num}'].value = 0
        ws[f'Y{row_num}'].font = DATA_FONT
        ws[f'Y{row_num}'].alignment = CENTER_ALIGN
        ws[f'Y{row_num}'].border = THIN_BORDER
        ws[f'Y{row_num}'].number_format = CNY_FMT

        # Z: 清关费(已填入)
        # AA: 保险费(已填入)

        # AB = SUM(Y:AA) 目的港费用合计
        ws[f'AB{row_num}'].value = f'=SUM(Y{row_num}:AA{row_num})'
        ws[f'AB{row_num}'].font = DATA_FONT
        ws[f'AB{row_num}'].alignment = CENTER_ALIGN
        ws[f'AB{row_num}'].border = THIN_BORDER
        ws[f'AB{row_num}'].number_format = CNY_FMT

        # AC: 国内运费 = U*10%
        ws[f'AC{row_num}'].value = f'=ROUND(U{row_num}*0.1,2)'
        ws[f'AC{row_num}'].font = DATA_FONT
        ws[f'AC{row_num}'].alignment = CENTER_ALIGN
        ws[f'AC{row_num}'].border = THIN_BORDER
        ws[f'AC{row_num}'].number_format = CNY_FMT

        # AD: 国内运费税费 = AC*6%
        ws[f'AD{row_num}'].value = f'=ROUND(AC{row_num}*0.06,2)'
        ws[f'AD{row_num}'].font = DATA_FONT
        ws[f'AD{row_num}'].alignment = CENTER_ALIGN
        ws[f'AD{row_num}'].border = THIN_BORDER
        ws[f'AD{row_num}'].number_format = CNY_FMT

        # AE: 国内运费最终开票金额 = AC+AD
        ws[f'AE{row_num}'].value = f'=AC{row_num}+AD{row_num}'
        ws[f'AE{row_num}'].font = DATA_FONT
        ws[f'AE{row_num}'].alignment = CENTER_ALIGN
        ws[f'AE{row_num}'].border = THIN_BORDER
        ws[f'AE{row_num}'].number_format = CNY_FMT

        # AF: 报关费合计 = X
        ws[f'AF{row_num}'].value = f'=X{row_num}'
        ws[f'AF{row_num}'].font = DATA_FONT
        ws[f'AF{row_num}'].alignment = CENTER_ALIGN
        ws[f'AF{row_num}'].border = THIN_BORDER
        ws[f'AF{row_num}'].number_format = CNY_FMT

        # AG: 报关费税费 = AF*6%
        ws[f'AG{row_num}'].value = f'=ROUND(AF{row_num}*0.06,2)'
        ws[f'AG{row_num}'].font = DATA_FONT
        ws[f'AG{row_num}'].alignment = CENTER_ALIGN
        ws[f'AG{row_num}'].border = THIN_BORDER
        ws[f'AG{row_num}'].number_format = CNY_FMT

        # AH: 报关费最终开票金额 = AF+AG
        ws[f'AH{row_num}'].value = f'=AF{row_num}+AG{row_num}'
        ws[f'AH{row_num}'].font = DATA_FONT
        ws[f'AH{row_num}'].alignment = CENTER_ALIGN
        ws[f'AH{row_num}'].border = THIN_BORDER
        ws[f'AH{row_num}'].number_format = CNY_FMT

        # AI: 国内运费+报关费最终开票金额 = AE+AH
        ws[f'AI{row_num}'].value = f'=AE{row_num}+AH{row_num}'
        ws[f'AI{row_num}'].font = DATA_FONT
        ws[f'AI{row_num}'].alignment = CENTER_ALIGN
        ws[f'AI{row_num}'].border = THIN_BORDER
        ws[f'AI{row_num}'].number_format = CNY_FMT

        # AJ: 国际运费 = U*90%*AB2 (欧元)
        ws[f'AJ{row_num}'].value = f'=ROUND(U{row_num}*0.9*$AB$2,2)'
        ws[f'AJ{row_num}'].font = DATA_FONT
        ws[f'AJ{row_num}'].alignment = CENTER_ALIGN
        ws[f'AJ{row_num}'].border = THIN_BORDER
        ws[f'AJ{row_num}'].number_format = EUR_FMT

        # AK: 欧洲申报费(欧元) = Y*AB2
        ws[f'AK{row_num}'].value = f'=ROUND(Y{row_num}*$AB$2,2)'
        ws[f'AK{row_num}'].font = DATA_FONT
        ws[f'AK{row_num}'].alignment = CENTER_ALIGN
        ws[f'AK{row_num}'].border = THIN_BORDER
        ws[f'AK{row_num}'].number_format = EUR_FMT

        # AL: 清关费(欧元) = Z*AB2
        ws[f'AL{row_num}'].value = f'=ROUND(Z{row_num}*$AB$2,2)'
        ws[f'AL{row_num}'].font = DATA_FONT
        ws[f'AL{row_num}'].alignment = CENTER_ALIGN
        ws[f'AL{row_num}'].border = THIN_BORDER
        ws[f'AL{row_num}'].number_format = EUR_FMT

        # AM: 保险费(欧元) = AA*AB2
        ws[f'AM{row_num}'].value = f'=ROUND(AA{row_num}*$AB$2,2)'
        ws[f'AM{row_num}'].font = DATA_FONT
        ws[f'AM{row_num}'].alignment = CENTER_ALIGN
        ws[f'AM{row_num}'].border = THIN_BORDER
        ws[f'AM{row_num}'].number_format = EUR_FMT

        # AN: 税费关税（欧元）= 税金*AB2（取不到则显示"后补"）
        if r['has_tax']:
            ws[f'AN{row_num}'].value = r['fee_an_eur']
            ws[f'AN{row_num}'].number_format = EUR_FMT
        else:
            ws[f'AN{row_num}'].value = '后补'
            ws[f'AN{row_num}'].number_format = TEXT_FMT
        ws[f'AN{row_num}'].font = DATA_FONT
        ws[f'AN{row_num}'].alignment = CENTER_ALIGN
        ws[f'AN{row_num}'].border = THIN_BORDER

        # AO, AP, AQ: 税费VAT/赔偿金/其他(留空)
        for col_l in ['AO', 'AP', 'AQ']:
            cell = ws[f'{col_l}{row_num}']
            cell.value = ''
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # AR: 目的港费用合计(欧元) = SUM(AJ:AP)
        ws[f'AR{row_num}'].value = f'=SUM(AJ{row_num}:AP{row_num})'
        ws[f'AR{row_num}'].font = DATA_FONT
        ws[f'AR{row_num}'].alignment = CENTER_ALIGN
        ws[f'AR{row_num}'].border = THIN_BORDER
        ws[f'AR{row_num}'].number_format = EUR_FMT

    # ── 合计行 ──
    # 合计行放在数据行之后，空一行
    sr = 4 + n + 1  # 合计行：4(数据起始行) + n(数据行数) + 1(空行)

    # 填写合计行
    ws.merge_cells(f'B{sr}:E{sr}')
    ws[f'B{sr}'].value = '合计'
    ws[f'B{sr}'].font = BOLD_FONT
    ws[f'B{sr}'].alignment = CENTER_ALIGN
    ws[f'B{sr}'].border = THIN_BORDER

    sum_cols = ['L', 'M', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X',
                'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI',
                'AJ', 'AK', 'AL', 'AM', 'AN', 'AR']
    for cl in sum_cols:
        cell = ws[f'{cl}{sr}']
        cell.value = f'=SUM({cl}4:{cl}{sr-2})'
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        if cl in ('AJ', 'AK', 'AL', 'AM', 'AN', 'AR'):
            cell.number_format = EUR_FMT
        elif cl in ('L', 'M'):
            cell.number_format = INT_FMT
        else:
            cell.number_format = CNY_FMT

    # 为合计行的其他列加边框
    for cl in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'N', 'AO', 'AP', 'AQ']:
        cell = ws[f'{cl}{sr}']
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN

    # ── 提取月份（从运单日期中取出现最多的月份）──
    from collections import Counter
    bill_year, bill_month = 2026, 5
    date_months = []
    for info in waybills.values():
        d = info.get('date', '')
        if d and len(d) >= 6:
            date_months.append(d[:6])  # YYYYMM
    if date_months:
        most_common = Counter(date_months).most_common(1)[0][0]
        bill_year = int(most_common[:4])
        bill_month = int(most_common[4:6])
    # ── 更新标题行 ──
    ws['B1'].value = f'赛诺吉（深圳）国际货运代理有限公司{bill_year}年{bill_month}月对账单'

    # ── 互换AN/AO列标题（AN→税费关税, AO→税费VAT）──
    ws['AN3'].value = '税费关税\n（欧元）'
    ws['AO3'].value = '税费VAT\n（欧元）'

    # ── 欧元折算人民币汇率(V2)导出为空 ──
    ws['V2'].value = None

    # ── 删除合计行和国内账号之间多余的空白行，保留1行空白 ──
    # 合计行在sr，国内账号在行40，sr+1保留为空行，删除sr+2~39行
    if sr + 2 <= 39:
        # 先保存银行信息区域(行40+)的合并单元格，delete_rows不会自动调整合并单元格
        deleted_count = 39 - (sr + 1)  # 删除的行数
        bank_merges = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= 40:
                bank_merges.append({
                    'min_row': mr.min_row, 'max_row': mr.max_row,
                    'min_col': mr.min_col, 'max_col': mr.max_col,
                })
                ws.unmerge_cells(str(mr))
        # 执行删除
        ws.delete_rows(sr + 2, deleted_count)
        # 重新建立合并单元格（减去偏移量）
        for m in bank_merges:
            new_min = m['min_row'] - deleted_count
            new_max = m['max_row'] - deleted_count
            new_range = f'{get_column_letter(m["min_col"])}{new_min}:' \
                        f'{get_column_letter(m["max_col"])}{new_max}'
            ws.merge_cells(new_range)

    # ── 创建"运费合计"汇总Sheet ──
    ws_summary = wb.create_sheet(title='运费合计')

    # 标题行
    summary_headers = ['货代运单号', '运费', '报关费', '运费税额', '报关费税额', '目的港费用合计', '合计']
    for ci, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=ci)
        cell.value = h
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 填充每一行（与原始账单数据行对应）
    for i, r in enumerate(data_rows):
        row_num_s = 2 + i
        orig_row = 4 + i  # 原始账单中的数据行号
        ws_summary.row_dimensions[row_num_s].height = 20

        # A: 货代运单号（文本引用原始账单D列）
        cell_a = ws_summary.cell(row=row_num_s, column=1)
        cell_a.value = r['wb_no']
        cell_a.font = DATA_FONT
        cell_a.alignment = CENTER_ALIGN
        cell_a.border = THIN_BORDER

        # B: 运费 = 原始账单U列
        cell_b = ws_summary.cell(row=row_num_s, column=2)
        cell_b.value = f"='原始账单'!U{orig_row}"
        cell_b.font = DATA_FONT
        cell_b.alignment = CENTER_ALIGN
        cell_b.border = THIN_BORDER
        cell_b.number_format = CNY_FMT

        # C: 报关费 = 原始账单X列
        cell_c = ws_summary.cell(row=row_num_s, column=3)
        cell_c.value = f"='原始账单'!X{orig_row}"
        cell_c.font = DATA_FONT
        cell_c.alignment = CENTER_ALIGN
        cell_c.border = THIN_BORDER
        cell_c.number_format = CNY_FMT

        # D: 运费税额 = 原始账单AD列
        cell_d = ws_summary.cell(row=row_num_s, column=4)
        cell_d.value = f"='原始账单'!AD{orig_row}"
        cell_d.font = DATA_FONT
        cell_d.alignment = CENTER_ALIGN
        cell_d.border = THIN_BORDER
        cell_d.number_format = CNY_FMT

        # E: 报关费税额 = 原始账单AG列
        cell_e = ws_summary.cell(row=row_num_s, column=5)
        cell_e.value = f"='原始账单'!AG{orig_row}"
        cell_e.font = DATA_FONT
        cell_e.alignment = CENTER_ALIGN
        cell_e.border = THIN_BORDER
        cell_e.number_format = CNY_FMT

        # F: 目的港费用合计 = 原始账单AB列
        cell_f = ws_summary.cell(row=row_num_s, column=6)
        cell_f.value = f"='原始账单'!AB{orig_row}"
        cell_f.font = DATA_FONT
        cell_f.alignment = CENTER_ALIGN
        cell_f.border = THIN_BORDER
        cell_f.number_format = CNY_FMT

        # G: 合计 = B+C+D+E+F
        cell_g = ws_summary.cell(row=row_num_s, column=7)
        cell_g.value = f'=B{row_num_s}+C{row_num_s}+D{row_num_s}+E{row_num_s}+F{row_num_s}'
        cell_g.font = DATA_FONT
        cell_g.alignment = CENTER_ALIGN
        cell_g.border = THIN_BORDER
        cell_g.number_format = CNY_FMT

    # 合计行
    sr_sum = 2 + n + 1  # 空一行后写合计
    ws_summary.merge_cells(f'A{sr_sum}:A{sr_sum}')
    cell_sum_label = ws_summary.cell(row=sr_sum, column=1)
    cell_sum_label.value = '合计'
    cell_sum_label.font = BOLD_FONT
    cell_sum_label.alignment = CENTER_ALIGN
    cell_sum_label.border = THIN_BORDER

    for ci in range(2, 8):
        cell = ws_summary.cell(row=sr_sum, column=ci)
        col_l = get_column_letter(ci)
        cell.value = f'=SUM({col_l}2:{col_l}{sr_sum-2})'
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.number_format = CNY_FMT

    # 调整列宽
    ws_summary.column_dimensions['A'].width = 22
    for ci in range(2, 8):
        ws_summary.column_dimensions[get_column_letter(ci)].width = 16

    # ── 保存 ──
    wb.save(output_path)
    print(f"✅ 账单已保存: {output_path}")

    return True


def print_summary(waybills, data_rows):
    """打印汇总信息"""
    print(f"\n📋 汇总:")
    print(f"   运单数: {len(waybills)}")
    print(f"   账单行: {len(data_rows)}")
    print()

    # 按渠道统计
    ch_stats = defaultdict(lambda: {'count': 0, 'freight': 0, 'premium': 0, 'customs': 0, 'clearance': 0})
    for r in data_rows:
        ch = r['channel']
        ch_stats[ch]['count'] += 1
        ch_stats[ch]['freight'] += r['fee_o']
        ch_stats[ch]['premium'] += r['fee_aa']
        ch_stats[ch]['customs'] += r['fee_v']
        ch_stats[ch]['clearance'] += r['fee_z']

    for ch, stats in ch_stats.items():
        print(f"   {ch}: {stats['count']}票")
        print(f"      海运费: {stats['freight']:.2f}")
        print(f"      保费:   {stats['premium']:.2f}")
        print(f"      报关费: {stats['customs']:.2f}")
        print(f"      清关费: {stats['clearance']:.2f}")


def main():
    if len(sys.argv) < 2:
        print("用法: python3 gen_sr_bill.py <系统账单.xls> [输出文件名.xlsx]")
        print("示例: python3 gen_sr_bill.py 'SR账单自动生成/系统账单-广东思锐光学股份有限公司-GDSR-260501-31.xls'")
        sys.exit(1)

    system_bill_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    # 自动寻找模板
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sr_dir = os.path.join(script_dir, 'SR账单自动生成')
    template_path = os.path.join(sr_dir, '思锐账单模板模板 思锐开票账单-JTT（5.1-5.31）.xlsx')
    if not os.path.exists(template_path):
        template_path = os.path.join(script_dir, '思锐账单模板模板 思锐开票账单-JTT（5.1-5.31）.xlsx')

    if not os.path.exists(template_path):
        print(f"❌ 找不到模板文件")
        for f in os.listdir(script_dir):
            if '思锐' in f and f.endswith('.xlsx'):
                template_path = os.path.join(script_dir, f)
                break
        if not os.path.exists(template_path):
            print("   请确保 思锐账单模板模板*.xlsx 在 SR账单自动生成/ 目录下")
            sys.exit(1)

    # 自动寻找订单列表
    order_list_path = None
    if os.path.exists(sr_dir):
        for f in os.listdir(sr_dir):
            if '思锐订单列表' in f and f.endswith('.xlsx'):
                order_list_path = os.path.join(sr_dir, f)
                break
    if not order_list_path:
        for f in os.listdir(script_dir):
            if '思锐订单列表' in f and f.endswith('.xlsx'):
                order_list_path = os.path.join(script_dir, f)
                break

    print(f"📂 系统账单: {system_bill_path}")
    print(f"📂 模板文件: {template_path}")
    if order_list_path:
        print(f"📂 订单列表: {order_list_path}")

    # 读取系统账单
    waybills = read_system_bill(system_bill_path)
    print(f"✅ 读取到 {len(waybills)} 个运单")

    # 读取订单列表
    order_list = read_order_list(order_list_path) if order_list_path else {}
    if order_list:
        matched = sum(1 for wb in waybills if wb in order_list)
        print(f"✅ 订单列表: {len(order_list)} 条, 匹配运单 {matched} 个")

    # 自动生成输出文件名
    if not output_path:
        base_name = os.path.basename(system_bill_path)
        date_match = re.search(r'(\d{2})(\d{2})[-.]?(\d{2})[-.]?(\d{2})', base_name)
        if date_match:
            year = '20' + date_match.group(1)
            month = date_match.group(2)
            day_start = date_match.group(3)
            day_end = date_match.group(4)
            date_str = f'{year}.{month}.{day_start}-{month}.{day_end}'
        else:
            dates = []
            for info in waybills.values():
                if info['date']:
                    dates.append(info['date'])
            if dates:
                dates.sort()
                m = dates[0][4:6]
                date_str = f'2026.{m}.1-{m}.31'
            else:
                from datetime import date
                today = date.today()
                date_str = today.strftime('2026.%m.1-%m.31')
        output_path = f'思锐开票账单-JTT（{date_str}）.xlsx'

    if '/' not in output_path and '\\' not in output_path:
        if os.path.exists(sr_dir):
            output_path = os.path.join(sr_dir, output_path)

    # 生成账单
    print(f"📄 输出: {output_path}")
    success = generate_bill(waybills, template_path, output_path, order_list)

    # 构建数据行用于打印汇总
    data_rows = []
    for wb_no, info in waybills.items():
        fees = info['fees']
        order_info = order_list.get(wb_no, {}) if order_list else {}
        channel = order_info.get('channel', '')
        def sf(v):
            try: return float(v) if v else 0
            except: return 0
        data_rows.append({
            'channel': channel,
            'fee_o': sf(fees.get('运费', 0)),
            'fee_aa': sf(fees.get('保费', 0)),
            'fee_v': sf(fees.get('出口(国内)报关费', 0)),
            'fee_z': sf(fees.get('进口(海外)清关费', 0)),
        })

    if success:
        print_summary(waybills, data_rows)
        print(f"\n✅ 思锐账单生成成功!")
        print(f"   📄 {output_path}")
        print(f"\n⚠️  请检查以下内容:")
        print(f"   1. 报关单号需要手动填写")
        print(f"   2. 销售订单号可能需要补全")
        print(f"   3. 税金转欧元的汇率是否正确")


if __name__ == '__main__':
    main()
