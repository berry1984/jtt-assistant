#!/usr/bin/env python3
"""
生成提单PDF和电放保函xlsx
根据 TR 退税资料明细.xlsx 的"5月提单信息"页面，对应生成提单和电放保函

规则：
- 引用模板：By sea → 提单By sea.pdf, By train → 提单By train.pdf, By truck → 提单By truck.pdf
- 跳过 Place of receipt = "查验" 的货件
- 文件名格式：JTT号+渠道+箱数+提单/电放保函.后缀
"""

import os, sys, re
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import fitz  # PyMuPDF

# ── 路径 ──
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, '模版', '模版')
DATA_FILE = os.path.join(BASE_DIR, 'TR 退税资料明细.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output_5月')

BL_SEA_TEMPLATE = os.path.join(TEMPLATE_DIR, '提单By sea.pdf')
BL_TRUCK_TEMPLATE = os.path.join(TEMPLATE_DIR, '提单By truck.pdf')
BL_TRAIN_TEMPLATE = os.path.join(TEMPLATE_DIR, '提单By train.pdf')
TELEX_TEMPLATE = os.path.join(TEMPLATE_DIR, '电放保函模板.xlsx')

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── 固定信息 ──
SHIPPER = ("Guangzhou Tuorui Technology Co.,Ltd\n"
           "Room 411,No.101 Dexing Road Wanggang Jiahe Street\n"
           "Baiyun District, Guangzhou")

CONSIGNEE = ("Hong Kong Lixiang Trading Company Limited\n"
             "FLAT/RM 3B 3/F BANK TOWER NOS.351&353 KING'S\n"
             "ROAD NORTH POINT HK")


# ── 工具函数 ──
def _safe_str(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    return str(v)


def fmt_date(d):
    if isinstance(d, datetime):
        return d
    if isinstance(d, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=d)
    return d


def sanitize_filename(s):
    s = str(s).replace('/', '_').replace('\\', '_').replace(':', '_')
    s = s.replace('*', '_').replace('?', '_').replace('"', '_')
    s = s.replace('<', '_').replace('>', '_').replace('|', '_')
    s = s.replace(' ', '_')
    return s


def _split_jtt_cell(raw):
    """拆分 JTT no. 单元格中的多个单号。

    单元格可能是一个单号，也可能是多个单号用逗号/顿号/斜杠/空格/换行等分隔：
        "JTT202605000328"                        → ['JTT202605000328']
        "JTT202605000328,340,330,331,334,335"    → ['JTT202605000328','JTT202605000340',...]
    裸序号（无 JTT 前缀）自动补全基础前缀。
    """
    parts = [p.strip() for p in re.split(r'[,，、;；/\s]+', str(raw)) if p.strip()]
    if not parts:
        return []
    # 只有首段是 JTT 编号才算 JTT 单元格；备注说明行（首段如 "2."/"备注"）原样返回
    if not parts[0].startswith('JTT'):
        return parts
    # 基础前缀：JTT + YYYYMM + "000"（12 位，与 _fmt_jtts 的 PREFIX_LEN 一致）
    base = parts[0][:12]
    return [p if p.startswith('JTT') else base + p for p in parts]


def load_shipments():
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb['5月提单信息']
    headers = []
    for c in list(ws.iter_rows(min_row=1, max_row=1))[0]:
        headers.append(c.value)

    # ── 向下填充：同组空白行继承上行的 B/L No / 渠道 / 模板 / 船名航次等 ──
    FILL_DOWN_COLS = {
        'B/L No.', '引用模板', '渠道',
        'Ocean Vessel', 'Voy.No',
        'Place of receipt', 'Port of loading',
        'Port of discharge', 'Place of delivery',
        'Container no.', 'collect',
        'Place and date of issue', 'on board date',
        'Shipper', 'Consignee', 'Notify party',
    }
    fill_cache = {}

    shipments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        raw_jtt = _safe_str(d.get('JTT no.', '')).strip()
        if not raw_jtt:
            continue  # 完全空行直接跳过
        # 拆分单元格中的多个单号
        jtt_nos = _split_jtt_cell(raw_jtt)
        if not jtt_nos:
            continue
        # 跳过非 JTT 编号的行（如底部备注说明行）
        if not jtt_nos[0].startswith('JTT'):
            continue

        # 向下填充：属于填充列表且当前为 None 的列，从上一条缓存取值
        for col_name in FILL_DOWN_COLS:
            if col_name in d and d[col_name] is None and col_name in fill_cache:
                d[col_name] = fill_cache[col_name]
            elif col_name in d and d[col_name] is not None:
                fill_cache[col_name] = d[col_name]

        template = _safe_str(d.get('引用模板', '')).strip()
        if not template:
            continue
        if _safe_str(d.get('Place of receipt', '')).strip() == '查验':
            print(f'  ⏭ 跳过查验: {raw_jtt}')
            continue
        bl_no = _safe_str(d.get('B/L No.', '')).strip()
        if not bl_no:
            print(f'  ⏭ 跳过无提单号: {raw_jtt}')
            continue

        # 单元格含多个单号 → 拆分为多票；箱数/KGS/CBM 只在首个单号保留，
        # 避免同 B/L 合并时重复累加
        for i, jtt in enumerate(jtt_nos):
            item = dict(d)
            item['JTT no.'] = jtt
            if i > 0:
                item['cartons'] = 0
                item['KGS'] = 0
                item['CBM'] = 0
            shipments.append(item)

    wb.close()
    return shipments


def _get_template_path(template_type):
    mapping = {
        'by sea': BL_SEA_TEMPLATE,
        'by truck': BL_TRUCK_TEMPLATE,
        'by train': BL_TRAIN_TEMPLATE,
    }
    return mapping.get(template_type.lower().strip())


# ============================================================
# 1. 电放保函 (Telex Release) — .xlsx
# ============================================================
def generate_telex(shipment, jtt_part=None, total_cartons=None):
    if not os.path.exists(TELEX_TEMPLATE):
        print(f'  ❌ 找不到电放保函模板')
        return None
    jtt_no = jtt_part or _safe_str(shipment.get('JTT no.', ''))
    channel = _safe_str(shipment.get('渠道', ''))
    cartons = total_cartons if total_cartons is not None else (shipment.get('cartons', 0) or 0)
    bl_no = _safe_str(shipment.get('B/L No.', ''))
    vessel = _safe_str(shipment.get('Ocean Vessel', ''))
    voy = _safe_str(shipment.get('Voy.No', ''))
    container = _safe_str(shipment.get('Container no.', ''))
    collect_date = fmt_date(shipment.get('collect', ''))

    fname = f'{jtt_no}{sanitize_filename(channel)}{cartons}件电放保函.xlsx'
    out_path = os.path.join(OUTPUT_DIR, fname)

    wb = openpyxl.load_workbook(TELEX_TEMPLATE)
    ws = wb['sheet1']

    # 只写值，不改任何格式（字体/对齐/边框保留模板原样）
    ws['E10'] = bl_no
    vessel_str = f'{vessel}/{voy}' if voy else vessel
    ws['E12'] = vessel_str
    ws['E14'] = container

    # Shipper/Consignee — 固定标签文案（与 Web 模块一致，不写 shipper/consignee 变量）
    ws['A16'] = 'Shipper （发货人）                   :'
    ws['A18'] = 'Consignee （收货人）               :   '

    # 日期
    if isinstance(collect_date, datetime):
        m, d, y = collect_date.month, collect_date.day, collect_date.year
    else:
        m, d, y = '5', '30', '2026'
    ws['C31'] = m
    ws['D31'] = d
    ws['E31'] = y

    wb.save(out_path)
    wb.close()
    print(f'  ✅ 电放保函: {fname}')
    return out_path


# ============================================================
# 2. 提单 (Bill of Lading) — .pdf
# ============================================================
#
# 两阶段方法：
#   阶段1：用清除矩形覆盖旧数据区域 → 红划删除（白色填充）
#   阶段2：在指定插入点写入新文本
#
# 数据提取函数（接收 shipment 字典，返回文本）
# -----------------------------------------------------------

def _data_fns():
    """返回数据提取函数字典"""
    return {
        'bl_no':    lambda s: _safe_str(s.get('B/L No.', '')),
        'vessel':   lambda s: (_safe_str(s.get('Ocean Vessel', '')) + '/' +
                               _safe_str(s.get('Voy.No', ''))) if s.get('Voy.No') else _safe_str(s.get('Ocean Vessel', '')),
        'place_rcpt': lambda s: _safe_str(s.get('Place of receipt', '')),
        'port_load':  lambda s: _safe_str(s.get('Port of loading', '')),
        'port_disc':  lambda s: _safe_str(s.get('Port of discharge', '')),
        'place_delv': lambda s: _safe_str(s.get('Place of delivery', '')),
        'container':  lambda s: _safe_str(s.get('Container no.', '')),
        'marks':    lambda s: _safe_str(s.get('Marks & No.', '')) or 'N/M',
        'cartons':  lambda s: str(int(s.get('cartons', 0) or 0)) + ' CARTONS',
        'desc':     lambda s: _safe_str(s.get('Description of goods', '')),
        'kgs':      lambda s: (_safe_str(s.get('KGS', 0) or 0) + ' KGS') if s.get('KGS') else '',
        'cbm':      lambda s: (_safe_str(s.get('CBM', 0) or 0) + ' CBM') if s.get('CBM') else '',
        'pdi_date': lambda s: (fmt_date(s.get('Place and date of issue', '')).strftime('%Y/%m/%d')
                               if isinstance(fmt_date(s.get('Place and date of issue', '')), datetime)
                               else _safe_str(s.get('Place and date of issue', ''))),
        'ob_date':  lambda s: (fmt_date(s.get('on board date', '')).strftime('%Y/%m/%d')
                               if isinstance(fmt_date(s.get('on board date', '')), datetime)
                               else _safe_str(s.get('on board date', ''))),
    }


def _write_wrapped(page, text, point, fontsize, max_width=158, max_words=10):
    """写入可自动换行的多行文本（Description of goods）。

    换行规则（满足任一即换行）：
      1. 遇到逗号（, 或 ，）或数据中的换行 → 换行
      2. 每行不超过 max_words（10）个单词
      3. 行宽不超过 max_width（点）
    避免单行溢出页面。
    """
    if not text:
        return
    segments = [seg for seg in re.split(r'[,，\n]+', text) if seg.strip()]
    if not segments:
        return
    lines = []
    cur = []
    for seg in segments:
        for w in seg.split():
            if not cur:
                cur = [w]
                continue
            candidate = cur + [w]
            too_wide = fitz.get_text_length(' '.join(candidate), fontname='helv',
                                            fontsize=fontsize) > max_width
            if len(cur) >= max_words or too_wide:
                lines.append(cur)
                cur = [w]
            else:
                cur = candidate
        if cur:  # 逗号/换行分隔 → 另起一行
            lines.append(cur)
            cur = []
    if cur:
        lines.append(cur)
    x, y = point
    line_h = fontsize * 1.25
    for i, line in enumerate(lines):
        page.insert_text(fitz.Point(x, y + i * line_h), ' '.join(line),
                         fontsize=fontsize, fontname='helv', color=(0, 0, 0))


def _redact_and_insert(page, clear_rects, text_inserts, shipment, desc_max_width=158):
    """
    clear_rects: [(x0, y0, x1, y1), ...]  — 红划删除区域
    text_inserts: [((x, y), field_key, fontsize), ...] — 插入点 + 字段名
    """
    F = _data_fns()
    # 阶段1：删除旧数据
    for rect in clear_rects:
        if not rect:
            continue
        page.add_redact_annot(fitz.Rect(*rect), fill=(1, 1, 1))
    page.apply_redactions()
    # 阶段2：写入新数据
    for pt, field_key, fontsize in text_inserts:
        text = F[field_key](shipment)
        if not text:
            continue
        if field_key == 'desc':
            # Description of goods 多行自动换行（每行最多 10 词，超宽自动换行）
            _write_wrapped(page, text, pt, fontsize, desc_max_width)
        else:
            page.insert_text(fitz.Point(*pt), text, fontsize=fontsize,
                             fontname='helv', color=(0, 0, 0))


def sea_train_fields(is_train=False):
    """
    By sea / By train 模板的清除矩形 + 文本插入点

    蓝色边框线位置（By sea, 左栏 x=50-251）:
      水平线: y=272, 286, 300, 313, 327, 341, 354
      单元格按 y 区间分：
        y=272-286: 标签 "Pre-carriage by"
        y=286-300: ⭐ Place of receipt 数据 (161,287)-(238,298)
        y=300-313: 标签 "Ocean vessel"
        y=313-327: ⭐ Vessel 数据 (52,314)-(150,322)
        y=327-341: 标签 "Port of discharge"
        y=341-354: ⭐ Port discharge 数据 (52,342)-(153,352)
      右栏 x=251-513: 整个区域 y=272-354 为一个单元格

    货物表格（x=50-513, y=354-547）：
      水平线: y=354(表头底), y=382(数据行顶), y=533/546(底)
      垂直线: x=119, 210, 384, 457
    """
    # ── 清除矩形 — 精确到每个单元格，距蓝色边框 2px ──
    clear_rects = [
        # B/L No（蓝线 y=54，旧数据 y=54-64）
        (421, 55, 509, 66),
        # Place of receipt（单元格 y=286-300，旧数据 y=287-298）
        (156, 288, 249, 299),
        # Vessel（单元格 y=313-327，旧数据 y=314-322）
        (52, 315, 154, 326),
        # Port loading（同单元格 y=313-327，旧数据 y=314-325）
        (156, 315, 249, 326),
        # Port discharge（单元格 y=341-354，旧数据 y=342-352）
        (52, 343, 158, 353),
        # Place delivery（同单元格，旧数据 y=342-352）
        (156, 343, 268, 353),
        # Container No（表格下方 y=533-546）
        (52, 535, 118, 545),
        # ── 货物表逐列清除（蓝线 y=382 和 y=533 之间）──
        (52, 384, 117, 531),    # 列1 Marks
        (121, 384, 208, 531),   # 列2 Quantity
        (212, 384, 382, 531),   # 列3 Description
        (386, 384, 455, 531),   # 列4 KGS
        (459, 384, 511, 531),   # 列5 CBM
        # 底部日期（避开蓝色文字和蓝线）
        #   SHIPPED ON BOARD 标签 y=605-616, 蓝线 y=630, 旧日期 y=618-630
        (426, 618, 473, 629),
        #   Place and Date of Issue 蓝色标签 y=635-644, 蓝线 y=656
        (395, 645, 455, 654),
        #   On board date 蓝色标签 y=686-696, 底部蓝线 y=709
        (164, 697, 213, 708),
    ]
    # Train 有额外 "CHONGQING" 文字在 Vessel 区域
    if is_train:
        clear_rects.append((52, 314, 249, 326))

    # ── 文字写入位置 — baseline 与模板原始数据对齐 ──
    # 模板原始数据 bbox:
    #   B/L No "CHN3261876P5"  bbox y=54-64   → baseline y=64
    #   Place of receipt       bbox y=287-298 → baseline y=298
    #   Vessel                 bbox y=314-322 → baseline y=322
    #   Port loading           bbox y=314-325 → baseline y=325
    #   Port discharge         bbox y=342-352 → baseline y=352
    #   Place delivery         bbox y=342-352 → baseline y=352
    #   Container No           bbox y=533-544 → baseline y=544
    #   表格第一行             bbox y=383-395 → baseline y=395
    text_inserts = [
        ((428, 67),   'bl_no',    9),
        ((161, 298),  'place_rcpt', 11),
        ((52, 322),   'vessel',   8),
        ((161, 325),  'port_load', 11),
        ((52, 352),   'port_disc', 11),
        ((161, 352),  'place_delv', 11),
        ((56, 544),   'container', 10),
        # 货物表 — baseline 对齐模板 y=395
        ((77, 395),   'marks',    9),
        ((130, 395),  'cartons',  9),
        ((218, 395),  'desc',     10.5),
        ((382, 395),  'kgs',      9),
        ((459, 395),  'cbm',      9),
        # 底部日期
        ((428, 628),  'ob_date',  9),
        ((397, 654),  'pdi_date', 9),
        ((166, 706),  'ob_date',  9),
    ]

    return clear_rects, text_inserts


def truck_fields():
    """By truck 模板的清除矩形 + 文本插入点"""
    clear_rects = [
        (480, 14, 560, 32),        # B/L No
        (35, 288, 310, 310),       # Vessel + Place loading
        (35, 310, 310, 340),       # Place discharge + delivery
        (35, 355, 590, 455),       # 货物表整区
        (375, 665, 460, 695),      # Place/Date of Issue
        (180, 718, 240, 750),      # Date of departure
    ]
    text_inserts = [
        ((488, 18),    'bl_no',    10),
        ((37, 292),    'vessel',   9),
        ((202, 292),   'port_load', 9),
        ((37, 316),    'port_disc', 9),
        ((202, 316),   'place_delv', 9),
        ((42, 360),    'container', 8),
        ((46, 362),    'marks',    9),
        ((298, 362),   'desc',     10.5),
        ((435, 362),   'kgs',      8),
        ((435, 372),   'cbm',      8),
        ((380, 683),   'pdi_date', 9),
        ((182, 739),   'ob_date',  9),
    ]
    return clear_rects, text_inserts


def generate_bl(shipment, jtt_part=None, total_cartons=None):
    template_type = _safe_str(shipment.get('引用模板', ''))
    template_path = _get_template_path(template_type)
    if not template_path or not os.path.exists(template_path):
        print(f'  ⚠ 找不到提单模板: {template_type}')
        return None

    jtt_no = jtt_part or _safe_str(shipment.get('JTT no.', ''))
    channel = _safe_str(shipment.get('渠道', ''))
    cartons = total_cartons if total_cartons is not None else (shipment.get('cartons', 0) or 0)
    fname = f'{jtt_no}{sanitize_filename(channel)}{cartons}件提单.pdf'
    out_path = os.path.join(OUTPUT_DIR, fname)

    try:
        doc = fitz.open(template_path)
        page = doc[0]
        tt = template_type.lower().strip()

        if tt in ('by sea', 'by train'):
            is_train = (tt == 'by train')
            clear_rects, inserts = sea_train_fields(is_train)
        elif tt == 'by truck':
            clear_rects, inserts = truck_fields()
        else:
            print(f'  ⚠ 未知模板类型: {template_type}')
            doc.close()
            return None

        _redact_and_insert(page, clear_rects, inserts, shipment,
                           desc_max_width=130 if tt == 'by truck' else 158)
        doc.save(out_path, garbage=4, deflate=True)
        doc.close()
        print(f'  ✅ 提单: {fname}')
        return out_path
    except Exception as e:
        import traceback
        print(f'  ❌ 提单生成失败 {jtt_no}: {e}')
        traceback.print_exc()
        return None


# ============================================================
# 合并逻辑 — 同一 B/L No 的多票合并为一票
# ============================================================
def _fmt_jtts(jtt_list):
    """格式化 JTT 号列表用于文件名

    单票: JTT202605000307
    多票: JTT202605000364,353 （共享前缀 + 逗号分隔的序号）
    """
    if len(jtt_list) == 1:
        return jtt_list[0]
    # 所有 JTT 号格式: "JTT" + 年月日 + 序号，前12字符为公共前缀
    PREFIX_LEN = 12  # "JTT202605000"
    prefix = jtt_list[0][:PREFIX_LEN]
    if all(j.startswith(prefix) for j in jtt_list):
        suffixes = [j[PREFIX_LEN:] for j in jtt_list]
        return prefix + ','.join(suffixes)
    # fallback: 用 os.path.commonprefix
    prefix = os.path.commonprefix(jtt_list)
    if len(prefix) >= 6:
        suffixes = [j[len(prefix):] for j in jtt_list]
        return prefix + ','.join(suffixes)
    return ','.join(jtt_list)


def _merge_shipments(group):
    """合并同一 B/L 的多票货为一个 shipment dict

    多行数据用 \\n 拼接，cartons/KGS/CBM 累加
    """
    merged = dict(group[0])

    marks_list = []
    desc_list = []
    total_cartons = 0
    total_kgs = 0.0
    total_cbm = 0.0

    for s in group:
        m = str(s.get('Marks & No.', '') or '').strip()
        if m and m not in marks_list:
            marks_list.append(m)
        d = str(s.get('Description of goods', '') or '').strip()
        if d and d not in desc_list:
            desc_list.append(d)
        total_cartons += int(s.get('cartons', 0) or 0)
        try:
            total_kgs += float(s.get('KGS', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            total_cbm += float(s.get('CBM', 0) or 0)
        except (ValueError, TypeError):
            pass

    merged['Marks & No.'] = '\n'.join(marks_list) if marks_list else ''
    merged['Description of goods'] = '\n'.join(desc_list) if desc_list else ''
    merged['cartons'] = total_cartons
    merged['KGS'] = round(total_kgs, 2)
    merged['CBM'] = round(total_cbm, 3)

    merged['_jtt_list'] = [str(s.get('JTT no.', '') or '').strip() for s in group]
    return merged


# ============================================================
# Main
# ============================================================
def main():
    print(f'📂 数据文件: {DATA_FILE}')
    print(f'📂 模板目录: {TEMPLATE_DIR}')
    print(f'📂 输出目录: {OUTPUT_DIR}\n')

    shipments = load_shipments()
    print(f'📋 共 {len(shipments)} 票待生成（按 B/L No 合并输出）\n')

    # 按 B/L No 分组（同一提单的多票合并输出）
    bl_groups = defaultdict(list)
    for s in shipments:
        bl_no = _safe_str(s.get('B/L No.', '')).strip()
        bl_groups[bl_no].append(s)

    telex_ok = bl_ok = 0
    for i, (bl_no, group) in enumerate(bl_groups.items(), 1):
        if len(group) == 1:
            # 单票 — 直接生成
            s = group[0]
            jtt_no = _safe_str(s.get('JTT no.', ''))
            cartons = int(s.get('cartons', 0) or 0)
            print(f'[{i:02d}] {jtt_no} | {_safe_str(s.get("渠道", ""))} | {_safe_str(s.get("引用模板", ""))} | {cartons}箱')
            if generate_telex(s, jtt_part=jtt_no, total_cartons=cartons):
                telex_ok += 1
            if generate_bl(s, jtt_part=jtt_no, total_cartons=cartons):
                bl_ok += 1
            print()
        else:
            # 多票合并 — 汇总数据
            merged = _merge_shipments(group)
            jtt_part = _fmt_jtts(merged['_jtt_list'])
            total_cartons = merged['cartons']
            print(f'[{i:02d}] {jtt_part} | {_safe_str(merged.get("渠道", ""))} | {_safe_str(merged.get("引用模板", ""))} | {total_cartons}箱')
            if generate_telex(merged, jtt_part=jtt_part, total_cartons=total_cartons):
                telex_ok += 1
            if generate_bl(merged, jtt_part=jtt_part, total_cartons=total_cartons):
                bl_ok += 1
            print()

    print(f'={"=" * 50}')
    print(f'📊 完成统计:')
    print(f'   ✅ 电放保函: {telex_ok}/{len(bl_groups)}')
    print(f'   ✅ 提单:     {bl_ok}/{len(bl_groups)}')
    print(f'   📁 输出目录: {OUTPUT_DIR}')


if __name__ == '__main__':
    main()
