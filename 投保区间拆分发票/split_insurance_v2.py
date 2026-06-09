#!/usr/bin/env python3
"""
投保区间拆分工具 v2

将「下单 Excel」按每箱申报价值(RMB)拆分为多个独立 Excel 文件，
每个文件对应一个投保区间，用于分别投保。

规则详见：投保区间拆分规则说明_v2.md

用法：
  python3 split_insurance_v2.py <下单发票.xlsx>
  python3 split_insurance_v2.py <下单发票.xlsx> --out-dir ./输出目录
"""

import sys
import os
import re
import argparse
import zipfile
import io
import shutil
import tempfile
from copy import copy
from xml.etree import ElementTree as ET

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#  常量
# ═══════════════════════════════════════════════════════════════

CURRENCY_RATES = {
    '美金': 7, '美元': 7, 'USD': 7,
    '欧元': 8, 'EUR': 8,
    '英镑': 9, 'GBP': 9,
}

RANGES = [
    ('不足5000RMB',    0,  5000),
    ('5000-10000RMB',  5000, 10000),
    ('10000-20000RMB', 10000, 20000),
    ('20000-30000RMB', 20000, 30000),
    ('30000-40000RMB', 30000, 40000),
]

HEADER_END = 26
COL_HEADER_ROW = 27
DATA_START_ROW = 28
NUM_DATA_COLS = 19    # A-S
TOTAL_COLS = 20       # +T辅助列
IMG_COL = 16          # P列 = 产品图片

AUX_SHEETS = ['FBA地址库编码表', '服务名称', '换算']

THIN_SIDE = Side(style='thin')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE, bottom=THIN_SIDE)


# ═══════════════════════════════════════════════════════════════
#  辅助函数
# ═══════════════════════════════════════════════════════════════

def _copy_style(src_cell):
    return {
        'font': copy(src_cell.font) if src_cell.font is not None else Font(),
        'alignment': (copy(src_cell.alignment)
                      if src_cell.alignment is not None else Alignment()),
        'border': copy(src_cell.border) if src_cell.border is not None else Border(),
        'fill': copy(src_cell.fill) if src_cell.fill is not None else PatternFill(),
        'number_format': src_cell.number_format,
    }


def _apply_style(dst_cell, style):
    dst_cell.font = copy(style['font'])
    dst_cell.alignment = copy(style['alignment'])
    dst_cell.border = copy(style['border'])
    dst_cell.fill = copy(style['fill'])
    dst_cell.number_format = style['number_format']


def _normalize_box_no(box_no):
    """
    货箱编号统一格式：确保末尾数字前有 U0000 前缀。

    若箱号已含 "U0000" → 原样保留。
    若箱号为范围格式 "X-Y"（如 "1-3"）→ "U00001-3"。
    若箱号为纯数字 → "U00001"。
    若箱号已有字母前缀 → "FBA...U000001"。
    """
    s = str(box_no).strip()
    if 'U0000' in s:
        return s
    # 范围格式 "X-Y" → "U0000X-Y"
    if re.match(r'^\d+-\d+$', s):
        return f'U0000{s}'
    m = re.search(r'(\d+)$', s)
    if not m:
        return f'U0000{s}'
    prefix = s[:m.start()]
    digits = m.group(1)
    return f'{prefix}U0000{digits}'


def _parse_box_number(box_no):
    """
    解析箱号，提取前缀和数字部分。

    返回:
      - 简单数字 "1" → (prefix_upto_digits, digit_int)
      - 范围格式 "1-3" → (prefix_upto_digits, [(s, e)]) 特殊标记
      - 复杂格式 "FBA15...U000001" → (prefix, digit_int)
    """
    s = str(box_no).strip()
    # 范围格式
    m_range = re.match(r'^(\d+)-(\d+)$', s)
    if m_range:
        return ('__range__', int(m_range.group(1)), int(m_range.group(2)))
    m = re.search(r'(\d+)$', s)
    if m:
        return (s[:m.start()], int(m.group(1)))
    return (s, None)


def _calc_total_boxes(box_groups):
    """
    箱数计算：每行箱号各自计算箱数，再累加。

    - 范围格式 "X-Y"（如 "1-3"）→ Y-X+1 = 3
    - 简单数字（如 "U000001"）→ 1
    - 复杂格式（"FBA...U000001" 等）按前缀分组 max-min+1 后累加
    """
    total = 0
    # 用于处理复杂格式的分组
    complex_groups = {}

    for bg in box_groups:
        s = str(bg['box_no']).strip()

        # 范围格式 "X-Y" → 直接算 Y-X+1
        m_range = re.match(r'^\d+-\d+$', s)
        if m_range:
            parts = s.split('-')
            total += int(parts[1]) - int(parts[0]) + 1
            continue

        # 简单数字 → 1 箱
        if re.match(r'^\d+$', s):
            total += 1
            continue

        # 复杂格式（带字母前缀）：按前缀分组处理
        m = re.search(r'(\d+)$', s)
        if m:
            prefix = s[:m.start()]
            digit = int(m.group(1))
            complex_groups.setdefault(prefix, []).append(digit)
        else:
            # 无法解析 → 算 1 箱
            total += 1

    # 复杂格式每组内 max-min+1
    for digits in complex_groups.values():
        total += max(digits) - min(digits) + 1

    return total


# ═══════════════════════════════════════════════════════════════
#  图片基础设施提取 + 保留（支持 cellimages + 标准 drawing 两种模式）
# ═══════════════════════════════════════════════════════════════

# 需要从源文件复制的图片相关路径前缀
_IMG_INFRA_PREFIXES = [
    'xl/cellimages.xml',
    'xl/_rels/cellimages.xml.rels',
    'xl/drawings/',
    'xl/comments1.xml',
    'xl/media/',
    'xl/worksheets/_rels/sheet1.xml.rels',
]

# Content-Type 映射（PartName → ContentType）
# 用于补充输出文件中缺失的 Override 条目
_IMG_CONTENT_TYPES = {
    '/xl/cellimages.xml': 'application/vnd.openxmlformats-officedocument.spreadsheetml.cellImages+xml',
    '/xl/drawings/drawing1.xml': 'application/vnd.openxmlformats-officedocument.drawing+xml',
    '/xl/drawings/vmlDrawing1.vml': 'application/vnd.openxmlformats-officedocument.vmlDrawing',
    '/xl/comments1.xml': 'application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml',
}


def _get_image_data(src_path):
    """
    从源文件提取所有图片基础设施数据，供输出文件后处理使用。

    返回 dict:
      - files: {relative_path: bytes}  — 需要复制的所有文件
      - source_sheet_rels: [xml_element_dict]  — 源文件的 sheet rels 条目
    """
    data = {'files': {}, 'source_sheet_rels': [], 'sheet_xml_anchors': []}
    try:
        with zipfile.ZipFile(src_path, 'r') as z:
            all_names = z.namelist()
            for name in all_names:
                is_infra = any(
                    name.startswith(p) or name == p
                    for p in _IMG_INFRA_PREFIXES
                )
                if is_infra and not name.endswith('/'):
                    data['files'][name] = z.read(name)

            # 解析源文件 sheet1.xml.rels，记录所有关系类型
            rels_path = 'xl/worksheets/_rels/sheet1.xml.rels'
            if rels_path in data['files']:
                import xml.etree.ElementTree as ET
                NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                root = ET.fromstring(data['files'][rels_path])
                for child in root:
                    data['source_sheet_rels'].append({
                        'id': child.get('Id', ''),
                        'type': child.get('Type', ''),
                        'target': child.get('Target', ''),
                        'target_mode': child.get('TargetMode', ''),
                    })

            # 解析源文件 sheet1.xml，提取 <drawing>/<comments>/<legacyDrawing> 锚点
            if 'xl/worksheets/sheet1.xml' in all_names:
                NS_S = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                sheet_root = ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
                for child in sheet_root:
                    tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if tag in ('drawing', 'comments', 'legacyDrawing'):
                        data['sheet_xml_anchors'].append({
                            'tag': tag,
                            'rId': child.get(f'{{{NS_R}}}id', '') or child.get('r:id', ''),
                        })

            # 记录 Content_Types
            if '[Content_Types].xml' in all_names:
                data['content_types'] = z.read('[Content_Types].xml')
    except Exception:
        pass
    return data


def _embed_image_data(output_path, image_data, kept_src_rows=None):
    """
    后处理输出 xlsx：将源文件中的图片基础设施（drawing/comments/cellimages/media）
    复制到输出文件中，仅保留 kept_src_rows 中指定行的图片。

    参数:
      output_path: 输出 xlsx 路径
      image_data: _get_image_data() 返回的数据
      kept_src_rows: set 或 list，需要保留的源文件行号；None=保留全部
    """
    if not image_data or not image_data.get('files'):
        return

    if kept_src_rows is not None:
        kept_src_rows = set(kept_src_rows)

    tmp_dir = tempfile.mkdtemp()
    try:
        # 1. 解压输出文件
        with zipfile.ZipFile(output_path, 'r') as z:
            z.extractall(tmp_dir)

        # 声明命名空间
        NS_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        NS_S = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        NS_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
        NS_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'
        NS_CT = 'http://schemas.openxmlformats.org/package/2006/content-types'

        # ── 2. 过滤 drawing1.xml（按行号） ──
        used_img_rIds = set()
        drawing_blob = image_data['files'].get('xl/drawings/drawing1.xml')
        if drawing_blob and kept_src_rows is not None:
            # 解析 drawing1.xml，只保留指定行的 twoCellAnchor
            ET.register_namespace('xdr', NS_XDR)
            ET.register_namespace('a', NS_A)
            ET.register_namespace('r', NS_REL)
            ET.register_namespace('', NS_S)

            draw_tree = ET.fromstring(drawing_blob)
            # 注意 drawing1.xml 的根元素在 xdr 命名空间下
            root_tag = draw_tree.tag
            ns = {'xdr': NS_XDR, 'a': NS_A, 'r': NS_REL}
            # 根可能是 xdr:wsDr 或非命名空间
            # 使用通用的 XPath

            # 遍历所有 twoCellAnchor（可能在一级子元素或任意位置）
            # 手动迭代以便过滤
            anchors_to_keep = []
            for anchor in list(draw_tree):
                # 检查是否 twoCellAnchor
                local_tag = anchor.tag.split('}')[-1] if '}' in anchor.tag else anchor.tag
                if local_tag not in ('twoCellAnchor', 'oneCellAnchor', 'absoluteAnchor'):
                    continue

                # 提取行号
                from_elem = anchor.find(f'{{{NS_XDR}}}from')
                row_to_keep = True  # 默认保留
                if from_elem is not None:
                    row_el = from_elem.find(f'{{{NS_XDR}}}row')
                    if row_el is not None and row_el.text is not None:
                        drawing_row = int(row_el.text)  # 0-indexed
                        sheet_row = drawing_row + 1      # 转为1-indexed
                        if sheet_row not in kept_src_rows:
                            row_to_keep = False

                if row_to_keep:
                    anchors_to_keep.append(anchor)
                    # 提取该 anchor 中用到的图片 rId
                    for blip in anchor.iter(f'{{{NS_A}}}blip'):
                        embed = blip.get(f'{{{NS_REL}}}embed', '')
                        if embed:
                            used_img_rIds.add(embed)
                else:
                    draw_tree.remove(anchor)

            filtered_drawing = ET.tostring(draw_tree, xml_declaration=True, encoding='UTF-8')
            # 写入过滤后的 drawing1.xml
            draw_out = os.path.join(tmp_dir, 'xl', 'drawings', 'drawing1.xml')
            os.makedirs(os.path.dirname(draw_out), exist_ok=True)
            with open(draw_out, 'wb') as f:
                f.write(filtered_drawing)
        else:
            # 不过滤：直接复制
            for rel_path, blob in image_data['files'].items():
                if rel_path == 'xl/worksheets/_rels/sheet1.xml.rels':
                    continue
                if rel_path == 'xl/drawings/drawing1.xml' and drawing_blob:
                    continue  # 上面已处理
                dst = os.path.join(tmp_dir, rel_path)
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                with open(dst, 'wb') as f:
                    f.write(blob)

        # ── 3. 过滤 drawing1.xml.rels（只保留 used_img_rIds） ──
        draw_rels_blob = image_data['files'].get('xl/drawings/_rels/drawing1.xml.rels')
        if draw_rels_blob and kept_src_rows is not None and used_img_rIds:
            draw_rels_root = ET.fromstring(draw_rels_blob)
            for child in list(draw_rels_root):
                rid = child.get('Id', '')
                if rid not in used_img_rIds:
                    draw_rels_root.remove(child)
            filtered_draw_rels = ET.tostring(draw_rels_root, xml_declaration=True, encoding='UTF-8')
            draw_rels_out = os.path.join(tmp_dir, 'xl', 'drawings', '_rels', 'drawing1.xml.rels')
            os.makedirs(os.path.dirname(draw_rels_out), exist_ok=True)
            with open(draw_rels_out, 'wb') as f:
                f.write(filtered_draw_rels)
        elif draw_rels_blob:
            # 全部保留
            draw_rels_out = os.path.join(tmp_dir, 'xl', 'drawings', '_rels', 'drawing1.xml.rels')
            os.makedirs(os.path.dirname(draw_rels_out), exist_ok=True)
            with open(draw_rels_out, 'wb') as f:
                f.write(draw_rels_blob)

        # ── 4. 过滤 comments1.xml（只保留 kept_src_rows） ──
        comments_blob = image_data['files'].get('xl/comments1.xml')
        if comments_blob and kept_src_rows is not None:
            comm_tree = ET.fromstring(comments_blob)
            # comments1.xml: <comments><authors>...</authors><commentList><comment ref="P28"...>
            comm_list = comm_tree.find(f'{{{NS_S}}}commentList')
            if comm_list is not None:
                for child in list(comm_list):
                    ref = child.get('ref', '')
                    # 从 ref 中提取行号，如 "P28" → 28
                    import re as _re
                    m = _re.search(r'(\d+)$', ref)
                    if m and int(m.group(1)) not in kept_src_rows:
                        comm_list.remove(child)
            filtered_comments = ET.tostring(comm_tree, xml_declaration=True, encoding='UTF-8')
            comm_out = os.path.join(tmp_dir, 'xl', 'comments1.xml')
            with open(comm_out, 'wb') as f:
                f.write(filtered_comments)
        elif comments_blob:
            comm_out = os.path.join(tmp_dir, 'xl', 'comments1.xml')
            os.makedirs(os.path.dirname(comm_out), exist_ok=True)
            with open(comm_out, 'wb') as f:
                f.write(comments_blob)

        # ── 5. 复制 cellimages.xml（如果有） ──
        ci_blob = image_data['files'].get('xl/cellimages.xml')
        if ci_blob:
            ci_out = os.path.join(tmp_dir, 'xl', 'cellimages.xml')
            os.makedirs(os.path.dirname(ci_out), exist_ok=True)
            with open(ci_out, 'wb') as f:
                f.write(ci_blob)
        ci_rels_blob = image_data['files'].get('xl/_rels/cellimages.xml.rels')
        if ci_rels_blob:
            ci_rels_out = os.path.join(tmp_dir, 'xl', '_rels', 'cellimages.xml.rels')
            os.makedirs(os.path.dirname(ci_rels_out), exist_ok=True)
            with open(ci_rels_out, 'wb') as f:
                f.write(ci_rels_blob)

        # ── 6. 复制 media 图片文件（只复制被引用的） ──
        media_files = image_data.get('files', {})
        # 如果有过滤，只复制 used_img_rIds 相关的 media 文件
        if kept_src_rows is not None and used_img_rIds:
            # 从 drawing1.xml.rels 找出 used rIds 对应的 media 文件路径
            # targets 是相对于源文件 xl/drawings/drawing1.xml 的位置
            rels_data = image_data['files'].get('xl/drawings/_rels/drawing1.xml.rels', b'')
            if rels_data:
                rels_root = ET.fromstring(rels_data)
                used_media_paths = set()
                draw_base = os.path.dirname('xl/drawings/drawing1.xml')  # → 'xl/drawings'
                for child in rels_root:
                    if child.get('Id', '') in used_img_rIds:
                        target = child.get('Target', '')
                        if target:
                            used_media_paths.add(os.path.normpath(os.path.join(
                                draw_base, target
                            )))
                for media_path in used_media_paths:
                    if media_path in media_files:
                        dst = os.path.join(tmp_dir, media_path)
                        os.makedirs(os.path.dirname(dst), exist_ok=True)
                        with open(dst, 'wb') as f:
                            f.write(media_files[media_path])
        else:
            # 全部复制
            for rel_path, blob in media_files.items():
                if rel_path.startswith('xl/media/') and not rel_path.endswith('/'):
                    dst = os.path.join(tmp_dir, rel_path)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    with open(dst, 'wb') as f:
                        f.write(blob)

        # ── 7. 合并 sheet1.xml.rels ──
        rels_dir = os.path.join(tmp_dir, 'xl', 'worksheets', '_rels')
        os.makedirs(rels_dir, exist_ok=True)
        rels_path = os.path.join(rels_dir, 'sheet1.xml.rels')

        if os.path.exists(rels_path):
            rels_tree = ET.parse(rels_path)
            rels_root = rels_tree.getroot()
        else:
            rels_root = ET.Element(f'{{{NS_REL}}}Relationships')
            rels_tree = ET.ElementTree(rels_root)

        existing_types = set()
        next_rId = 1
        for child in rels_root:
            rid = child.get('Id', '')
            if rid.startswith('rId'):
                try:
                    next_rId = max(next_rId, int(rid[3:]) + 1)
                except ValueError:
                    pass
            existing_types.add(child.get('Type', ''))

        IMG_REL_TYPES = {
            f'{NS_REL}/drawing',
            f'{NS_REL}/comments',
            f'{NS_REL}/vmlDrawing',
            f'{NS_REL}/cellImages',
        }

        for src_rel in image_data.get('source_sheet_rels', []):
            rel_type = src_rel['type']
            if rel_type in IMG_REL_TYPES and rel_type not in existing_types:
                new_rel = ET.SubElement(rels_root, 'Relationship')
                new_rel.set('Id', f'rId{next_rId}')
                new_rel.set('Type', rel_type)
                new_rel.set('Target', src_rel['target'])
                if src_rel.get('target_mode'):
                    new_rel.set('TargetMode', src_rel['target_mode'])
                next_rId += 1
                existing_types.add(rel_type)

        rels_tree.write(rels_path, xml_declaration=True, encoding='UTF-8')

        # ── 8. 更新 sheet1.xml 锚点 ──
        sheet_path = os.path.join(tmp_dir, 'xl', 'worksheets', 'sheet1.xml')
        if os.path.exists(sheet_path) and image_data.get('sheet_xml_anchors'):
            sheet_tree = ET.parse(sheet_path)
            sheet_root = sheet_tree.getroot()

            type_to_rid = {}
            for child in rels_root:
                rt = child.get('Type', '')
                rid = child.get('Id', '')
                if rt and rid:
                    short_type = rt.split('/')[-1]
                    type_to_rid[short_type] = rid

            existing_anchors = set()
            for child in list(sheet_root):
                tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                if tag in ('drawing', 'comments', 'legacyDrawing'):
                    existing_anchors.add(tag)

            for anchor in image_data.get('sheet_xml_anchors', []):
                tag = anchor['tag']
                if tag not in existing_anchors:
                    src_rid = anchor['rId']
                    src_type = None
                    for src_rel in image_data.get('source_sheet_rels', []):
                        if src_rel['id'] == src_rid:
                            src_type = src_rel['type'].split('/')[-1]
                            break
                    if src_type and src_type in type_to_rid:
                        el = ET.SubElement(sheet_root, f'{{{NS_S}}}{tag}')
                        el.set(f'{{{NS_REL}}}id', type_to_rid[src_type])

            sheet_tree.write(sheet_path, xml_declaration=True, encoding='UTF-8')

        # ── 9. 补充 Content_Types ──
        ct_path = os.path.join(tmp_dir, '[Content_Types].xml')
        if os.path.exists(ct_path):
            ct_tree = ET.parse(ct_path)
            ct_root = ct_tree.getroot()

            existing_exts = set()
            for child in ct_root.findall(f'{{{NS_CT}}}Default'):
                ext = child.get('Extension', '')
                if ext:
                    existing_exts.add(ext.lower())
            needed_exts = {'png': 'image/png', 'jpeg': 'image/jpeg',
                           'jpg': 'image/jpeg', 'JPG': 'image/jpeg'}
            for ext, ctype in needed_exts.items():
                if ext not in existing_exts:
                    el = ET.SubElement(ct_root, f'{{{NS_CT}}}Default')
                    el.set('Extension', ext)
                    el.set('ContentType', ctype)

            if 'content_types' in image_data:
                src_ct_root = ET.fromstring(image_data['content_types'])
                existing_overrides = set()
                for child in ct_root.findall(f'{{{NS_CT}}}Override'):
                    pn = child.get('PartName', '')
                    if pn:
                        existing_overrides.add(pn)
                for child in src_ct_root.findall(f'{{{NS_CT}}}Override'):
                    pn = child.get('PartName', '')
                    ct_val = child.get('ContentType', '')
                    is_img_related = any(
                        kw in ct_val.lower()
                        for kw in ['drawing', 'comment', 'cellimage', 'vml']
                    )
                    if is_img_related and pn not in existing_overrides:
                        override = ET.SubElement(ct_root, f'{{{NS_CT}}}Override')
                        override.set('PartName', pn)
                        override.set('ContentType', ct_val)

            ct_tree.write(ct_path, xml_declaration=True, encoding='UTF-8')

        # ── 10. 重新打包 ──
        tmp_out = output_path + '.tmp'
        with zipfile.ZipFile(tmp_out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for dirpath, _, filenames in os.walk(tmp_dir):
                for fn in filenames:
                    full = os.path.join(dirpath, fn)
                    arcname = os.path.relpath(full, tmp_dir)
                    zout.write(full, arcname)

        shutil.move(tmp_out, output_path)

    except Exception as e:
        print(f'  ⚠️ 图片基础设施复制失败: {e}')
        import traceback
        traceback.print_exc()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ═══════════════════════════════════════════════════════════════
#  源文件解析
# ═══════════════════════════════════════════════════════════════

def parse_source(src_path):
    """
    解析源文件，提取所有需要的数据。

    返回 dict:
      - currency, rate, service_name
      - header_styles, col_header_styles, data_styles
      - box_groups: [{box_no, rows[], total_price, rmb}]
      - cellimages_data: 供后处理使用
      - merged_cells, col_widths, row_heights, aux_sheets
    """
    wb = load_workbook(src_path, data_only=True)

    if '发票' not in wb.sheetnames:
        print(f'❌ 源文件缺少「发票」工作表，找到: {wb.sheetnames}')
        sys.exit(1)

    ws = wb['发票']
    max_row = ws.max_row

    # ── 元信息 ──
    currency = str(ws.cell(24, 2).value or '').strip()
    rate = CURRENCY_RATES.get(currency, 9)
    service_name = str(ws.cell(1, 2).value or '').strip()
    print(f'  币种: {currency} → 汇率: {rate}')
    print(f'  公式: 每箱RMB = 单箱子货值 × 汇率({rate}) × 1.1')
    print(f'  服务: {service_name}')

    # ── 头部 Row 1-26 ──
    header_styles = {}
    for r in range(1, HEADER_END + 1):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(r, c)
            header_styles[(r, c)] = {'value': cell.value, **_copy_style(cell)}

    # ── 列头 Row 27 ──
    col_header_styles = {}
    for c in range(1, TOTAL_COLS + 1):
        cell = ws.cell(COL_HEADER_ROW, c)
        col_header_styles[c] = {'value': cell.value, **_copy_style(cell)}

    # ── 数据行 ──
    data_styles = {}
    data_rows = []

    for r in range(DATA_START_ROW, max_row + 1):
        box_no = ws.cell(r, 1).value
        if box_no is None or str(box_no).strip() == '':
            break
        data_rows.append(r)
        for c in range(1, NUM_DATA_COLS + 1):
            cell = ws.cell(r, c)
            data_styles[(r, c)] = {'value': cell.value, **_copy_style(cell)}

    # ── 箱组分组（按货箱编号） ──
    box_groups = []
    cur_no = None
    cur_group = None

    for r in data_rows:
        bno = str(ws.cell(r, 1).value).strip()
        qty = ws.cell(r, 4).value or 0
        uprice = ws.cell(r, 5).value or 0

        if bno != cur_no:
            cur_no = bno
            cur_group = {
                'box_no': bno,
                'rows': [r],
                'total_price': qty * uprice,
            }
            box_groups.append(cur_group)
        else:
            cur_group['rows'].append(r)
            cur_group['total_price'] += qty * uprice

    # 每箱 RMB = 货值 × 汇率 × 1.1
    for bg in box_groups:
        bg['rmb'] = round(bg['total_price'] * rate * 1.1, 2)

    # ── 提取图片基础设施数据（供后处理） ──
    image_data = _get_image_data(src_path)
    file_count = len(image_data.get('files', {}))
    if file_count:
        print(f'  提取图片基础设施: {file_count} 个文件')

    # ── 合并单元格（头部） ──
    merged_cells = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row < DATA_START_ROW:
            merged_cells.append(str(mc))

    # ── 列宽 ──
    col_widths = {}
    for c in range(1, TOTAL_COLS + 1):
        letter = get_column_letter(c)
        if letter in ws.column_dimensions:
            w = ws.column_dimensions[letter].width
            if w:
                col_widths[c] = w

    # ── 行高（头部） ──
    row_heights = {}
    for r in range(1, DATA_START_ROW):
        if r in ws.row_dimensions:
            h = ws.row_dimensions[r].height
            if h:
                row_heights[r] = h

    # ── 辅助 Sheet ──
    aux_sheets = {}
    for sn in AUX_SHEETS:
        if sn in wb.sheetnames:
            aws = wb[sn]
            rows = []
            for row in aws.iter_rows(min_row=1, max_row=aws.max_row,
                                     max_col=aws.max_column):
                rows.append([cell.value for cell in row])
            aux_sheets[sn] = rows

    wb.close()

    return {
        'currency': currency,
        'rate': rate,
        'service_name': service_name,
        'header_styles': header_styles,
        'col_header_styles': col_header_styles,
        'data_styles': data_styles,
        'data_rows': data_rows,
        'box_groups': box_groups,
        'image_data': image_data,
        'merged_cells': merged_cells,
        'col_widths': col_widths,
        'row_heights': row_heights,
        'aux_sheets': aux_sheets,
    }


# ═══════════════════════════════════════════════════════════════
#  区间分配
# ═══════════════════════════════════════════════════════════════

def assign_ranges(box_groups):
    """按每箱 RMB 将箱组分配到各投保区间"""
    assigned = {r[0]: [] for r in RANGES}
    for bg in box_groups:
        placed = False
        for r_name, r_low, r_high in RANGES:
            if r_low <= bg['rmb'] < r_high:
                assigned[r_name].append(bg)
                placed = True
                break
        if not placed:
            assigned[RANGES[-1][0]].append(bg)
    return assigned


# ═══════════════════════════════════════════════════════════════
#  输出文件生成
# ═══════════════════════════════════════════════════════════════

def create_range_output(src_data, range_name, box_groups, output_path):
    """
    为一个投保区间生成独立 .xlsx 文件。
    保存后会自动后处理嵌入 WPS cellimages 图片。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '发票'

    # 箱数 = 末尾数字 max-min+1
    total_boxes = _calc_total_boxes(box_groups)

    # ═══════════════════════════════════════════════
    #  1. 头部 Row 1-26
    # ═══════════════════════════════════════════════
    for r in range(1, HEADER_END + 1):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(r, c)
            info = src_data['header_styles'].get((r, c), {})
            cell.value = info.get('value')
            if 'font' in info:
                _apply_style(cell, info)

    for mc_str in src_data['merged_cells']:
        ws.merge_cells(mc_str)

    ws.cell(1, 1).value = (
        f'{src_data["service_name"]} - {range_name} ({total_boxes}箱)'
    )
    ws.cell(26, 2).value = total_boxes

    # ═══════════════════════════════════════════════
    #  2. 列头 Row 27
    # ═══════════════════════════════════════════════
    for c in range(1, TOTAL_COLS + 1):
        cell = ws.cell(COL_HEADER_ROW, c)
        info = src_data['col_header_styles'].get(c, {})
        cell.value = info.get('value')
        if 'font' in info:
            _apply_style(cell, info)

    t_hdr = ws.cell(COL_HEADER_ROW, 20)
    t_hdr.value = '每箱RMB'
    t_hdr.font = Font(name='微软雅黑', size=10, bold=True)
    t_hdr.alignment = Alignment(horizontal='center', vertical='center')
    t_hdr.border = THIN_BORDER
    t_hdr.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3',
                             fill_type='solid')

    # ═══════════════════════════════════════════════
    #  3. 数据行 Row 28+
    # ═══════════════════════════════════════════════
    current_row = DATA_START_ROW

    for bg in box_groups:
        for src_row in bg['rows']:
            # 统一行高 16
            ws.row_dimensions[current_row].height = 16

            for c in range(1, NUM_DATA_COLS + 1):
                cell = ws.cell(current_row, c)
                info = src_data['data_styles'].get((src_row, c), {})

                if c == 1:
                    # A列：货箱编号 → 统一加 U0000 前缀
                    raw = info.get('value', '')
                    cell.value = _normalize_box_no(raw)
                elif c == IMG_COL:
                    # P列：保留原始 DISPIMG 公式（WPS 通过 cellimages.xml 渲染）
                    cell.value = info.get('value')
                else:
                    cell.value = info.get('value')

                if 'font' in info:
                    _apply_style(cell, info)

            # T列：每箱 RMB
            t_cell = ws.cell(current_row, 20)
            t_cell.value = bg['rmb']
            t_cell.font = Font(name='微软雅黑', size=10)
            t_cell.alignment = Alignment(horizontal='center', vertical='center')
            t_cell.number_format = '#,##0.00'
            t_cell.border = THIN_BORDER

            current_row += 1

    # ═══════════════════════════════════════════════
    #  4. 列宽 & 头部行高
    # ═══════════════════════════════════════════════
    for c, w in src_data['col_widths'].items():
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = w
    ws.column_dimensions['T'].width = 15

    for r, h in src_data['row_heights'].items():
        if r <= COL_HEADER_ROW:
            ws.row_dimensions[r].height = h

    # ═══════════════════════════════════════════════
    #  5. 辅助 Sheet
    # ═══════════════════════════════════════════════
    for sn, rows in src_data['aux_sheets'].items():
        ws_aux = wb.create_sheet(title=sn)
        for r_idx, row_vals in enumerate(rows, 1):
            for c_idx, val in enumerate(row_vals, 1):
                ws_aux.cell(r_idx, c_idx).value = val

    # ═══════════════════════════════════════════════
    #  6. 保存
    # ═══════════════════════════════════════════════
    wb.save(output_path)
    wb.close()

    # ═══════════════════════════════════════════════
    #  7. 后处理：嵌入图片基础设施（仅保留当前区间的行）
    # ═══════════════════════════════════════════════
    image_data = src_data.get('image_data', {})
    if image_data:
        kept_rows = set()
        for bg in box_groups:
            for r in bg['rows']:
                kept_rows.add(r)
        _embed_image_data(output_path, image_data, kept_src_rows=kept_rows)


# ═══════════════════════════════════════════════════════════════
#  Web 集成接口
# ═══════════════════════════════════════════════════════════════

def split_invoice_to_ranges(src_path, output_dir=None):
    """
    供 JTT电商AI助手 Web 应用调用的高层接口。
    返回: dict {range_name: output_file_path}
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(src_path), 'output')
    os.makedirs(output_dir, exist_ok=True)

    src_data = parse_source(src_path)
    range_boxes = assign_ranges(src_data['box_groups'])

    out_files = {}
    for r_name, _, _ in RANGES:
        boxes = range_boxes[r_name]
        total = _calc_total_boxes(boxes)
        out_name = f'{r_name} ({total}箱).xlsx'
        out_path = os.path.join(output_dir, out_name)
        create_range_output(src_data, r_name, boxes, out_path)
        out_files[r_name] = out_path

    return out_files


# ═══════════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='下单发票 → 按投保区间拆分 v2',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('input', help='下单发票 .xlsx 文件路径')
    parser.add_argument('--out-dir', '-o',
                        help='输出目录（默认: 源文件所在目录下的 output 目录）')

    args = parser.parse_args()

    src_path = os.path.abspath(args.input)
    if not os.path.exists(src_path):
        print(f'❌ 文件不存在: {src_path}')
        sys.exit(1)

    if args.out_dir:
        output_dir = os.path.abspath(args.out_dir)
    else:
        output_dir = os.path.join(os.path.dirname(src_path), 'output')
    os.makedirs(output_dir, exist_ok=True)

    print(f'📂 读取源文件: {os.path.basename(src_path)}')
    src_data = parse_source(src_path)

    print(f'  总箱数: {_calc_total_boxes(src_data["box_groups"])}')
    for bg in src_data['box_groups']:
        print(f'  📦 {_normalize_box_no(bg["box_no"])}: {len(bg["rows"])} 品名行, '
              f'原币 {bg["total_price"]:.2f} → ¥{bg["rmb"]:.2f}')

    range_boxes = assign_ranges(src_data['box_groups'])

    print(f'\n📊 按区间拆分:')
    out_files = []
    for r_name, _, _ in RANGES:
        boxes = range_boxes[r_name]
        total = _calc_total_boxes(boxes)
        out_name = f'{r_name} ({total}箱).xlsx'
        out_path = os.path.join(output_dir, out_name)
        create_range_output(src_data, r_name, boxes, out_path)
        print(f'  ✅ {out_name}')
        out_files.append(out_path)

    print(f'\n✅ 拆分完成! 共 {len(out_files)} 个文件')
    print(f'   保存目录: {output_dir}')


if __name__ == '__main__':
    main()
